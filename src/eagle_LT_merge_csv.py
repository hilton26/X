#!/usr/bin/env python
# coding: utf-8

# # Merge Excel sheets

# https://stackoverflow.com/questions/20908018/import-multiple-excel-files-into-python-pandas-and-concatenate-them-into-one-dat

# Result: An .xlsx sheet in the py-test folder with CS1 fund PARN holdings as one sheet and CS1 fund NAVs as the other

# libraries
import time

start_time_merge = time.time()
start_time = time.time()

print(f"Importing libraries and setting paths ...")
from datetime import datetime
import pandas as pd
import os
from pathlib import Path
import openpyxl
from tqdm import tqdm, notebook  # notebook version of tqdm
import subprocess
from utilities import timediff, osprey

# set paths
pthDl = str(Path.home() / "Downloads")
pthPy = r"P:\Investment Operations\GRC\Compliance\Daily\py_reports.xlsm"
pthReports = r"\\PIM-CPT-FS.prescient.local\PIM-Documents$\Investment Operations\GRC\Compliance\Reg28 and Reg30 Reporting"
pthTest = r"P:\Working Folders\Hilton\W\Reg_Tests"

print(
    f"{timediff(start_time, time.time())} importing libraries and setting paths", "\n"
)

# get report date from the py_report.xlsm sheet
start_time = time.time()
print(f"Getting fund codes and report date ...")

# get credentials
holdings = pd.read_excel(pthPy, sheet_name="creds", header=None, usecols="A", nrows=2)
aladdin = holdings.iloc[0, 0]
sesame = holdings.iloc[1, 0]

# get report date
holdings1 = pd.read_excel(pthPy, sheet_name="downloader", usecols="D", nrows=1).dropna()
rptDate = list(holdings1)[0].date()  # convert dataframe header to a date

# get fund codes
df2 = pd.read_excel(pthPy, sheet_name="downloader", usecols="A").dropna()
f_codes = (",").join(df2.iloc[:, 0].tolist())

print(
    f" {datetime.strftime(rptDate, '%A %d %B %Y')} for {len(df2)} funds:",
    "\n",
    f"{f_codes}",
)

print(f"{timediff(start_time, time.time())} getting fund codes and report date", "\n")

# list and then pick out the .csv files from the Downloads folder and then ...
start_time = time.time()
print(f"Identifying the <R28I ... {rptDate.strftime('%d%m%Y')}.csv> files ...")

# pick out the csv files downloaded with osprey()
import re

pattern = r"^R28I.*\(1\) " + f"{datetime.strftime(rptDate, '%d%b%Y')}" + "\.csv$"
files_csv = [s for s in os.listdir(pthDl) if re.match(pattern, s)]

print(f" Number of lookthrough csv: {len(files_csv)}")
print(f" {', '.join(files_csv)}")

print(
    f"{timediff(start_time, time.time())} identifying the <R28I ... {rptDate.strftime('%d%m%Y')}.csv> files",
    "\n",
)

# ... merge the fund holdings by looping over their files and appending to an initially empty dataframe
start_time = time.time()
print(f"Merging {len(files_csv)} csv files into a dataframe ...")

# https://stackoverflow.com/questions/13784192/creating-an-empty-pandas-dataframe-and-then-filling-it
holdings = pd.DataFrame()  # initialise an empty dataframe
for filename in notebook.tqdm(files_csv):
    data = pd.read_csv(os.path.join(pthDl, filename))
    holdings = pd.concat([holdings, data])

print(
    f"{len(files_csv) - len(holdings['Entity Name'].unique())} difference: {len(files_csv)} holdings, {len(holdings['Entity Name'].unique())} NAVs"
)

print(
    f"{timediff(start_time, time.time())} merging {len(files_csv)} csv files into a dataframe",
    "\n",
)

# get the fund NAVs
start_time = time.time()
print(
    f"Getting the {len(df2)} fund NAVs as at {rptDate.strftime('%A %d %B %Y')} with osprey() ..."
)

df = pd.read_excel(pthPy, sheet_name="downloader", usecols="A").dropna()
f_codes = (",").join(df["Entity Name"].tolist())

# def osprey(rpt_type = 'r28i', funds = 'PABS,PIMBAL', d_from as datetime, d_to as datetime, sfx = 'csv', al, xe):
osprey("fnav", f_codes, rptDate, rptDate, "", "csv", aladdin, sesame)

# dataframe the fund NAVs
fund_navs = pd.read_csv(
    os.path.join(
        Path.home(), "Downloads", f"FNAV ({len(df2)}) {rptDate.strftime('%d%b%Y')}.csv"
    )
)

# converting NAVs from object to float
fund_navs["Total Net Assets"] = [
    str(x).replace(",", "").replace("-", "-") for x in fund_navs["Total Net Assets"]
]
fund_navs["Total Net Assets"] = fund_navs["Total Net Assets"].astype(float)

# holdings date column from type string to type datetime
fund_navs["Effective Date"] = pd.to_datetime(fund_navs["Effective Date"])

# determine which fund NAVs had not been downloaded   https://stackoverflow.com/questions/3462143/get-difference-between-two-lists-with-unique-entries
missing = (", ").join(list(set(fund_navs["NAV Entity ID"]) ^ set(df2["Entity Name"])))

print("", f"Fund NAVS not downloaded: {missing}", "\n")

m = len(fund_navs["NAV Entity ID"].unique())
print(
    f"{timediff(start_time, time.time())} getting the {m} fund NAV{'s' if m != 1 else ''} as at {rptDate.strftime('%A %d %B %Y')} with osprey()",
    "\n",
)

# convert the MV and EE columns from str to float
# https://stackoverflow.com/questions/45027400/reading-csv-file-to-pandas-dataframe-as-float
start_time = time.time()
print(f"Converting numerical columns to float ...")

# convert the MV and EE columns from str to float
cols_obj_to_float = ["End Market Value", "Closing Exposure PA"]
for col in cols_obj_to_float:
    holdings[col] = [str(x).replace(",", "").replace("-", "-") for x in holdings[col]]
    holdings[col] = holdings[col].astype(float)

# rename some columns
fund_navs.rename(
    columns={"Entity Name": "Long Name", "Total Net Assets": "Fund NAV"}, inplace=True
)

# add a report date column
holdings[datetime.strftime(rptDate, "%d %b %Y")] = ""  # add a date column to holdings

# group holdings by MV and EE totals
cols_to_drop = [
    "Investment Type",
    "i Issue Name",
    "Primary Asset ID",
    "CCY",
    "Reg28 Classification",
]
totals = (
    holdings.groupby("Entity Name", as_index=False).sum().drop(cols_to_drop, axis=1)
)
totals

print(
    f"{timediff(start_time, time.time())} converting numerical columns to float", "\n"
)

# compare holdings and NAVs
start_time = time.time()
print(f"Comparing dataframe holdings to NAVs ...")

# merge the holdings and NAVs dataframes in order to compare
df3 = totals.merge(
    fund_navs,
    how="left",
    left_on="Entity Name",
    right_on="NAV Entity ID",
    suffixes=("_x", None),
)
df3["MV - EE"] = df3["End Market Value"] - df3["Closing Exposure PA"]
df3["MV - NAV"] = df3["End Market Value"] - df3["Fund NAV"]

# order the columns and determine number of NaNs
order_cols = [
    "Effective Date",
    "Entity Name",
    "End Market Value",
    "Closing Exposure PA",
    "Fund NAV",
    "Percentage of Market Value",
    "MV - EE",
    "MV - NAV",
]
df3 = df3[order_cols]
nans = df3["Fund NAV"].isnull().sum()  # number of missing funds

# sort to show NaNs first
df3.sort_values(
    by="Fund NAV", axis=0, ascending=True, na_position="first", inplace=True
)

# format values with comma separator and two decimal places   https://www.geeksforgeeks.org/formatting-integer-column-of-dataframe-in-pandas/
cols_to_format = [
    "End Market Value",
    "Closing Exposure PA",
    "Fund NAV",
    "Percentage of Market Value",
    "MV - EE",
    "MV - NAV",
]
for col in cols_to_format:
    df3[col].float_format = "{:,.2f}".format

print(f"{timediff(start_time, time.time())} comparing dataframe holdings to NAVs", "\n")

# write the dataframe to review it as a worksheet
start_time = time.time()
print("Writing the dataframe to a sheet ...")

s = "s" if len(files_csv) != 1 else ""
lthr_name = os.path.join(
    pthTest,
    f"Lookthroughs ({len(files_csv)} fund{s}) {datetime.strftime(rptDate, '%d%b%Y')}.xlsx",
)

writer = pd.ExcelWriter(lthr_name, engine="xlsxwriter")  # instantiate a sheet writer
holdings.to_excel(writer, index=False, sheet_name="All")  # write the NAV sheet
df3.to_excel(
    writer, index=False, sheet_name=f"NAVs ({nans} missing)"
)  # write the missing NAVs sheet

writer.close()  # https://pandas.pydata.org/docs/reference/api/pandas.ExcelWriter.html   class for writing DataFrame objects into excel sheets
print("", lthr_name)

print(f"{timediff(start_time, time.time())} writing the dataframe to a sheet", "\n")

# (11) prettify the sheets using openpyxl
start_time = time.time()
print("Giving the review sheet structure with openpyxl ...")

# utilise openpyxl tools to add excel features to results sheet
import openpyxl as px  # for adding sort filters to the excel sheet
from openpyxl.cell import Cell  # to format cells
from openpyxl.styles import (
    Alignment,
    Color,
    PatternFill,
    Font,
    Border,
)  # to format cells

wb = px.load_workbook(lthr_name)
for sheet in wb.worksheets:
    sheet.auto_filter.ref = sheet.dimensions
    sheet.freeze_panes = sheet["D2"]
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.fill = PatternFill(
                start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
            )

# add the report date to the holdings sheet
sht = wb[f"All"]
sht["J1"].value = rptDate
sht["J1"].alignment = Alignment(horizontal="left", vertical="top")
sht.column_dimensions["J"].width = 12

# add a hyperlink to the 'NAVs' sheet
sht = wb[f"NAVs ({nans} missing)"]
sht["I1"].value = rptDate
sht["I1"].alignment = Alignment(horizontal="left", vertical="top")
sht["J1"].hyperlink = r"P:\Working Folders\Hilton\W\structures.xlsm"
sht["J1"].font = Font(bold=True, underline="single", color="0000EE")
sht["J1"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

# column width   https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size
dims = {}
for row in sht.rows:
    for cell in row:
        if cell.value:
            # dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
            dims[cell.column_letter] = max(
                (dims.get(cell.column_letter, 0), len(str(cell.value)))
            )
for col, value in dims.items():
    sht.column_dimensions[col].width = value
sht.column_dimensions["I"].width = 12

wb.save(lthr_name)  # save the file to the W folder
wb.close()

print("", lthr_name)
print(
    f"{timediff(start_time, time.time())} giving the review sheet structure with openpyxl",
    "\n",
)

# update the classifier sheet in py_reports.xlsm in preparation for eagle_r28_r30_classifications.ipynb

start_time = time.time()
print(
    'Updating py_reports.xlsm "classifier" sheet with path to the look-through csv holdings file ...'
)
# https://stackoverflow.com/questions/13381384/modify-an-existing-excel-file-using-openpyxl-in-python
# pthTest    = r'P:\Working Folders\Hilton\W\Reg_Tests'
# pthPy      = r'P:\Investment Operations\GRC\Compliance\Daily\py_reports.xlsm'
# len(holdings["Entity Name"].unique())
# lthr_name  = os.path.join(pthTest, f'Lookthroughs ({len(holdings["Entity Name"].unique())} funds) {rptDate.strftime("%d%b%Y")}.xlsx')

import xlwings as xw

wb = xw.Book(pthPy)
ws = wb.sheets("classifier")
ws.range("L2").value = lthr_name
ws.range("M1").value = "Reg 28 and Reg 30 only"  # alternative: 'CS1 format only'
wb.save()
wb.close()

print("", lthr_name)
print(
    f'{timediff(start_time, time.time())} updating py_reports.xlsm "classifier" sheet with path to the look-through csv holdings file',
    "\n",
)

print(
    f"{timediff(start_time_merge, time.time())} roundtrip time to merge the lookthrough csv files",
    "\n",
)

# commence classifications, i.e., issuers_1.py,  script

start_time_r28_r30_classifications = time.time()
print("Executing issuer_1 ...", "\n")

subprocess.run(
    ["python", r"C:/Users/hilton.netta/OneDrive - Prescient/py/gitrepo/issuers_1.py"]
)

print(
    f"Executing issuer_1 completed: {timediff(start_time_r28_r30_classifications, time.time())}"
)
