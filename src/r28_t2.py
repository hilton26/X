#!/usr/bin/env python
# coding: utf-8

# # Reg 28 Table 2 (Infrastructure)
#
# ### To generate multiple Reg 28 Table 2 reports

print("\n\n##############################################")
print("#                                            #")
print("#              START r28_t2.py  X            #")
print("#                                            #")
print("##############################################\n\n")

# libraries, libraries!
import time

start_time0 = time.time()
start_time = time.time()
print(f"Importing libraries ...")

import pandas as pd
import openpyxl, os
from datetime import datetime
from tqdm import tqdm 
from constants import (
    pthPy,
    pthSttlmnt,
    pth_tbl2_tmpl,
    pth_tbl2_static,
    pthReports,
    pthTest,
)
from utilities import timediff, item_row, prior_month_end

print(f" {timediff(start_time, time.time())} importing libraries completed \n")

# get report variables
start_time = time.time()
print(f"Getting report variables ...")

# Table 1 to Table 2 category translation
t1t2 = pd.read_excel(pth_tbl2_static, sheet_name="Static", usecols="A,E").dropna(
    subset=["Table 2"]
)

# get report fund codes from the py_report.xlsm 'arc' sheet
df_funds = pd.read_excel(pthPy, sheet_name="arc", usecols="N").dropna()  # funds
df_funds.iloc[:, 0] = df_funds.iloc[:, 0].str.upper()  # capitalise fund codes
df_funds.columns = ["Fund"]  # rename column
print(df_funds)

# get report date
df = pd.read_excel(pthPy, sheet_name="arc", usecols="S", nrows=2)
k = df.iloc[1, 0]  # report date
rptDate = k if k == k else prior_month_end(datetime.today().date())
# print(rptDate, k)

# get list of completed Reg28 classification files in the reporting folder
r28_files = [
    f
    for f in os.listdir(pthReports)
    if f[-20:] == f"Reg28 {rptDate.strftime('%d%b%Y')}.xlsx"
]
lst_r28 = [file[: file.find(" ")] for file in r28_files]
df_r28 = pd.DataFrame(lst_r28, columns=["Fund R28"])
files_r28 = ", ".join([file[: file.find(" ")] for file in r28_files])
print(
    f"{len(r28_files)} Reg28 files for {rptDate.strftime('%d %b %Y')} in the reporting folder",
    "\n",
)

# fund long name lookup
names = pd.read_excel(pthSttlmnt, sheet_name="Funds", usecols="A,B").dropna(
    subset=["Fund Code"]
)
mrg_w_rpt = df_funds.merge(names, left_on="Fund", right_on="Fund Code", how="left")
have_names = mrg_w_rpt.dropna(subset=["Fund Code"])
print(
    f"{len(mrg_w_rpt[mrg_w_rpt['Fund Code'].isnull()])} funds without corresponding long names: \
      {', '.join(mrg_w_rpt[mrg_w_rpt['Fund Code'].isnull()]['Fund'])}",
    "\n",
)

# identify fuund in the Table 2 list without completed Reg 28 reports
mrg_w_r28 = have_names.merge(df_r28, left_on="Fund", right_on="Fund R28", how="left")
print(
    f"{len(mrg_w_r28[mrg_w_r28['Fund R28'].isnull()])} funds without a completed Reg28 file: \
      {', '.join(mrg_w_r28[mrg_w_r28['Fund R28'].isnull()]['Fund'])}",
    "\n",
)

nl = mrg_w_r28.dropna(subset=["Fund R28"])

# get fund names
funds = df_funds.iloc[:, 0].values.tolist()
funds
s = "" if len(funds) == 1 else "s"
print(f" Report date: {rptDate.strftime('%d %b %Y')}")
print(
    f" Fund{s} ({len(nl)}, i.e., {len(funds)} in reporting list less \
{len(mrg_w_rpt[mrg_w_rpt['Fund Code'].isnull()])} unnamed less {len(mrg_w_r28[mrg_w_r28['Fund R28'].isnull()])} \
without Reg 28 sheets): {', '.join(nl['Fund'].values.tolist())}",
    "\n",
)

print(f"{timediff(start_time, time.time())} getting report variables completed \n")

# the loop
start_time_loop = time.time()
print(f"Preparing Reg 28 Table 2 report for {len(nl)} fund{s} ...")

# get Table 2 categories
z = pd.read_excel(
    pth_tbl2_tmpl, sheet_name="Tbl2", index_col=None, header=6, usecols="E"
).dropna()

# get style formatting methods from openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side


# tqdm(range(start, n), desc="Processing batches"):
for fund in tqdm(
    nl["Fund"],
    desc=f"Processing the Reg 28 Table 2 reports as at {rptDate.strftime('%d%b%Y')} ...",
):
    # set fund report Table 2 file name
    fnm = os.path.join(
        pthTest, f"{fund} Reg28 Table2 {rptDate.strftime('%d%b%Y')}.xlsx"
    )

    # get Table 2 items
    shtr = pd.read_excel(
        os.path.join(pthReports, f"{fund} Reg28 {rptDate.strftime('%d%b%Y')}.xlsx"),
        index_col=None,
        header=0,
        usecols="A:K",
    ).dropna(subset=["Entity Name"])  # reg categorised sheet
    shtr["Tbl2"] = shtr.apply(
        lambda row: (
            t1t2[t1t2["Reg 28 Classification"] == row["Reg 28 Classification"]].iat[
                0, 1
            ]
            if row["Infrastructure"] == "11(b)"
            else float("nan")
        ),
        axis=1,
    )  # add a translated Table 2 column
    shtr["Instr"] = shtr.apply(
        lambda row: "   ~ " + row["Primary Asset ID"] + " - " + row["i Issue Name"],
        axis=1,
    )  # instrument ID and name

    # populate the Table 2 sheet
    wb1 = openpyxl.load_workbook(pth_tbl2_tmpl)  # open the Table 2 template
    sh1 = wb1.active
    sh1["A2"] = f"{nl[nl['Fund Code'] == fund].iat[0, 2]} ({fund})"
    sh1["A5"] = f"As at {rptDate.strftime('%d %B %Y')}"
    sh1[item_row(sh1, "Total", 6)][1].value = shtr[shtr["Infrastructure"] == "11(b)"][
        "Percentage of Market Value"
    ].sum()
    sh1[item_row(sh1, "Total", 6)][2].value = shtr[shtr["Infrastructure"] == "11(b)"][
        "End Market Value"
    ].sum()
    sh1[item_row(sh1, "Fund Net Asset Value", 6)][2].value = shtr[
        "End Market Value"
    ].sum()
    sh1[item_row(sh1, "2.2", 6)][1].value = (
        shtr[shtr["Tbl2"] == "2.2.1"]["Percentage of Market Value"].sum()
        + shtr[shtr["Tbl2"] == "2.2.2"]["Percentage of Market Value"].sum()
    )
    sh1[item_row(sh1, "2.2", 6)][2].value = (
        shtr[shtr["Tbl2"] == "2.2.1"]["End Market Value"].sum()
        + shtr[shtr["Tbl2"] == "2.2.2"]["End Market Value"].sum()
    )
    sh1[item_row(sh1, "2.", 6)][1].value = (
        shtr[shtr["Tbl2"] == "2.2.1"]["Percentage of Market Value"].sum()
        + shtr[shtr["Tbl2"] == "2.2.2"]["Percentage of Market Value"].sum()
    )
    sh1[item_row(sh1, "2.", 6)][2].value = (
        shtr[shtr["Tbl2"] == "2.2.1"]["End Market Value"].sum()
        + shtr[shtr["Tbl2"] == "2.2.2"]["End Market Value"].sum()
    )
    sh1[item_row(sh1, "3.", 6)][1].value = (
        shtr[shtr["Tbl2"] == "3.1"]["Percentage of Market Value"].sum()
        + shtr[shtr["Tbl2"] == "3.2"]["Percentage of Market Value"].sum()
    )
    sh1[item_row(sh1, "3.", 6)][2].value = (
        shtr[shtr["Tbl2"] == "3.1"]["End Market Value"].sum()
        + shtr[shtr["Tbl2"] == "3.2"]["End Market Value"].sum()
    )
    sh1[item_row(sh1, "3.", 6)][1].value = (
        shtr[shtr["Tbl2"] == "3.1"]["Percentage of Market Value"].sum()
        + shtr[shtr["Tbl2"] == "3.2"]["Percentage of Market Value"].sum()
    )
    sh1[item_row(sh1, "3.", 6)][2].value = (
        shtr[shtr["Tbl2"] == "3.1"]["End Market Value"].sum()
        + shtr[shtr["Tbl2"] == "3.2"]["End Market Value"].sum()
    )
    sh1[item_row(sh1, "4.", 6)][1].value = (
        shtr[shtr["Tbl2"] == "4."]["Percentage of Market Value"].sum()
        + shtr[shtr["Tbl2"] == "4."]["Percentage of Market Value"].sum()
    )
    sh1[item_row(sh1, "4.", 6)][2].value = (
        shtr[shtr["Tbl2"] == "4."]["End Market Value"].sum()
        + shtr[shtr["Tbl2"] == "4."]["End Market Value"].sum()
    )
    sh1[item_row(sh1, "5.", 6)][1].value = (
        shtr[shtr["Tbl2"] == "5."]["Percentage of Market Value"].sum()
        + shtr[shtr["Tbl2"] == "5."]["Percentage of Market Value"].sum()
    )
    sh1[item_row(sh1, "5.", 6)][2].value = (
        shtr[shtr["Tbl2"] == "5."]["End Market Value"].sum()
        + shtr[shtr["Tbl2"] == "5."]["End Market Value"].sum()
    )
    sh1[item_row(sh1, "6.", 6)][1].value = (
        shtr[shtr["Tbl2"] == "6."]["Percentage of Market Value"].sum()
        + shtr[shtr["Tbl2"] == "6."]["Percentage of Market Value"].sum()
    )
    sh1[item_row(sh1, "6.", 6)][2].value = (
        shtr[shtr["Tbl2"] == "6."]["End Market Value"].sum()
        + shtr[shtr["Tbl2"] == "6."]["End Market Value"].sum()
    )
    sh1[item_row(sh1, "8.", 6)][1].value = (
        shtr[shtr["Tbl2"] == "8."]["Percentage of Market Value"].sum()
        + shtr[shtr["Tbl2"] == "8."]["Percentage of Market Value"].sum()
    )
    sh1[item_row(sh1, "8.", 6)][2].value = (
        shtr[shtr["Tbl2"] == "8."]["End Market Value"].sum()
        + shtr[shtr["Tbl2"] == "8."]["End Market Value"].sum()
    )
    sh1[item_row(sh1, "9.", 6)][1].value = (
        shtr[shtr["Tbl2"] == "9."]["Percentage of Market Value"].sum()
        + shtr[shtr["Tbl2"] == "9."]["Percentage of Market Value"].sum()
    )
    sh1[item_row(sh1, "9.", 6)][2].value = (
        shtr[shtr["Tbl2"] == "9."]["End Market Value"].sum()
        + shtr[shtr["Tbl2"] == "9."]["End Market Value"].sum()
    )
    sh1[item_row(sh1, "10.", 6)][1].value = (
        shtr[shtr["Tbl2"] == "10."]["Percentage of Market Value"].sum()
        + shtr[shtr["Tbl2"] == "10."]["Percentage of Market Value"].sum()
    )
    sh1[item_row(sh1, "10.", 6)][2].value = (
        shtr[shtr["Tbl2"] == "10."]["End Market Value"].sum()
        + shtr[shtr["Tbl2"] == "10."]["End Market Value"].sum()
    )
    sh1[item_row(sh1, "11.2", 6)][1].value = (
        shtr[shtr["Tbl2"] == "11.2"]["Percentage of Market Value"].sum()
        + shtr[shtr["Tbl2"] == "11.2"]["Percentage of Market Value"].sum()
    )
    sh1[item_row(sh1, "11.2", 6)][2].value = (
        shtr[shtr["Tbl2"] == "11.2"]["End Market Value"].sum()
        + shtr[shtr["Tbl2"] == "11.2"]["End Market Value"].sum()
    )
    sh1.title = f"{fund} Table2 {rptDate.strftime('%d%b%Y')}"  # tab name

    # update sheet with Table 2 values
    for item in z["Item"]:
        k = shtr.loc[
            (shtr["Infrastructure"] == "11(b)") & (shtr["Tbl2"] == item),
            ["Percentage of Market Value", "End Market Value", "Instr", "Issuer"],
        ].sort_values("Issuer", axis=0)  # infrastructure instruments and issuers

        # get issuer summaries for the item
        # https://stackoverflow.com/questions/32059397/pandas-groupby-without-turning-grouped-by-column-into-index
        m = k.groupby("Issuer", as_index=False)[
            ["Percentage of Market Value", "End Market Value"]
        ].sum()  # Issuers

        # https://www.boardinfinity.com/blog/learn-about-reset-index-pandas/#:~:text=To%20reset%20the%20index%20on,causes%20it%20to%20return%20Nothing.
        k.reset_index(inplace=True, drop=True)

        if len(m) > 0:
            sh1[item_row(sh1, item, 6)][1].value = shtr[shtr["Tbl2"] == item][
                "Percentage of Market Value"
            ].sum()
            sh1[item_row(sh1, item, 6)][2].value = shtr[shtr["Tbl2"] == item][
                "End Market Value"
            ].sum()
            spaces = 1  # number of rows below the item row at which to insert spaces
            sh1.insert_rows(
                item_row(sh1, item, 6), spaces
            )  # insert row location, number of rows to insert
            for index, row in m.iterrows():
                row_insert = item_row(sh1, item, 6) + 1 + index + spaces * index
                # print(f'insert rows from row {row_insert} and insert text at row {row_insert + spaces}')
                sh1.insert_rows(
                    row_insert, amount=spaces + 1
                )  # insert row location, number of rows to insert
                sh1[row_insert + spaces][1].value = row["Percentage of Market Value"]
                sh1[row_insert + spaces][2].value = row["End Market Value"]
                sh1[row_insert + spaces][3].value = row["Issuer"]

    row1 = item_row(sh1, "1.", 6)
    rows = item_row(sh1, "TTLb", 6)
    for row in range(row1, rows):
        # https://stackoverflow.com/questions/49525545/openpyxl-formatting-cell-with-decimal
        sh1[
            "B{}".format(row)
        ].number_format = '_-* #,##0.00_-;-* #,##0.00_-;_-* "-"_-;_-@_-'
        sh1[
            "C{}".format(row)
        ].number_format = '_-* #,##0.00_-;-* #,##0.00_-;_-* "-"_-;_-@_-'

    for cell in sh1[
        "A"
    ]:  # https://stackoverflow.com/questions/24917201/applying-borders-to-a-cell-in-openpyxl
        for row in range(row1, rows):
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"))

    for cell in sh1[
        "B"
    ]:  # https://stackoverflow.com/questions/26671581/horizontal-text-alignment-in-openpyxl
        for row in range(row1, rows):
            cell.alignment = Alignment(vertical="top")
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"))

    for cell in sh1[
        "C"
    ]:  # https://stackoverflow.com/questions/26671581/horizontal-text-alignment-in-openpyxl
        for row in range(row1, rows):
            cell.alignment = Alignment(vertical="top")
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"))

    for cell in sh1[
        "D"
    ]:  # https://stackoverflow.com/questions/38619471/iterate-through-all-rows-in-specific-column-openpyxl
        for row in range(row1, rows):
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"))

    # table titles alignment
    for col in range(1, 5):
        sh1.cell(item_row(sh1, "Item", 6), column=col).alignment = Alignment(
            wrap_text=True, vertical="top", horizontal="center"
        )

    for col in range(2, 4):
        sh1.cell(item_row(sh1, "TTLb", 6), column=col).alignment = Alignment(
            wrap_text=True, vertical="top", horizontal="center"
        )

    # borders
    item = pd.read_excel(
        pth_tbl2_tmpl,
        sheet_name="Tbl2",
        index_col=None,
        header=None,
        names=["Items"],
        usecols="F",
    ).dropna()
    item["RwN"] = item["Items"].apply(lambda row: item_row(sh1, row, 6))
    rwn = item["RwN"].tolist()  # list of row numbers requiring top borders
    for itm in rwn:
        for col in range(1, 5):
            sh1.cell(row=itm, column=col).border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
            )

    for col in range(1, 4):  # 'Total' bottom border
        sh1.cell(row=item_row(sh1, "Ttlb", 6), column=col).border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    for col in range(1, 5):
        for rw in range(item_row(sh1, "Ttlb", 6) + 1, item_row(sh1, "Ttlb", 6) + 3):
            sh1.cell(row=rw, column=col).border = Border(
                left=Side(border_style=None),
                right=Side(border_style=None),
                top=Side(border_style=None),
                bottom=Side(border_style=None),
            )

    for col in range(2, 4):  # 'Fund Net Asset Value' top and bottom borders
        sh1.cell(
            row=item_row(sh1, "Fund Net Asset Value", 6), column=col
        ).border = Border(top=Side(style="thin"), bottom=Side(style="double"))
    # https://openpyxl.readthedocs.io/en/latest/api/openpyxl.styles.borders.html

    for rw in range(item_row(sh1, "Ttlb", 6), item_row(sh1, "Ttlb", 6) + 3):
        sh1.cell(row=rw, column=4).border = Border(
            right=Side(border_style=None), top=Side(style="thin")
        )

    for rw in range(item_row(sh1, "Ttlb", 6) + 1, item_row(sh1, "Ttlb", 6) + 3):
        sh1.cell(row=rw, column=4).border = Border(
            top=Side(border_style=None), bottom=Side(border_style=None)
        )

    # remove guide columns - https://openpyxl.readthedocs.io/en/stable/editing_worksheets.html
    sh1.delete_cols(5, 2)  # delete columns E:F

    wb1.save(fnm)
    wb1.close

    # open_xl_file(fnm)
    # print(f' Tbl2 {fund}, {timediff(start_time, time.time())}')

print(
    f"{timediff(start_time0, time.time())} for {len(nl['Fund'])} Reg 28 Table 2 reports \
at {rptDate.strftime('%d%b%Y')} completed and stored in the test folder"
)

print("\n\n##############################################")
print("#                                            #")
print("#               END r28_t2.py   X            #")
print("#                                            #")
print("##############################################\n\n")
