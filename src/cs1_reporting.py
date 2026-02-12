#!/usr/bin/env python
# coding: utf-8

# # Prepare the CS1 reports

# In[12]:


# libraries, libraries!
import time

start_time_cs1 = time.time()
start_time = time.time()
from datetime import datetime, timedelta
import pandas as pd
from pathlib import Path
import os
from tqdm import tqdm
from constants import pthPy, pthSttlmnt, pth_dl, pthReports, pthTest, pth_r28_lmts
from utilities import (
    timediff,
    prior_month_end,
    item_row,
    range_border,
    rows_align_height,
)

# import openpyxl and certain of its functions

start_time = time.time()
print("Importing openpyxl and some of its functions")

import openpyxl

# import copy  # "AttributeError: Style objects are immutable and cannot be changed. Reassign the style with a copy"
from openpyxl.styles import (
    NamedStyle,
    Alignment,
    Font,
    PatternFill,
)  # https://openpyxl.readthedocs.io/en/stable/styles.html
from openpyxl.styles.borders import Border, Side
# from openpyxl.cell import (Cell)  # https://stackoverflow.com/questions/42215933/apply-wrap-text-to-all-cells-using-openpyxl

# cell number and border formats
cell_format = '#,##0.00_);(#,##0.00);"-"'  # e.g., sh["B15"].number_format = cell_format
thin_border = Side(style="thin", color="000000")  # black color
cell_border = Border(
    left=thin_border, right=thin_border, top=thin_border, bottom=thin_border
)  # e.g., sh["B15"].border = cell_border
cell_font_calibri_11 = Font(
    name="Calibri", size=11, color="000000"
)  # https://www.youtube.com/watch?v=fzUbfI8z1uc at 4:56
cell_font_calibri_11_bold = Font(
    name="Calibri", size=11, color="000000", bold=True
)  # https://www.youtube.com/watch?v=fzUbfI8z1uc at 4:56

print(
    f" {timediff(start_time, time.time())} importing openpyxl and some of its functions",
    "\n",
)

# # get report data
# def get_inputs():
start_time_inputs = time.time()
print(f"\nGetting the CS1 report inputs ...")

# get list of funds
# df     = pd.read_excel(pthPy, sheet_name = "r28_cs1", usecols = "A,D,F:G,J").dropna(subset = ['Fund'])
df = pd.read_excel(pthPy, sheet_name="arc", usecols="N").dropna()
df.columns = ["Funds"]  # rename column
funds = df["Funds"].apply(str.upper)  # cconvert to series

df2 = pd.read_excel(pthPy, sheet_name="arc", usecols="S", nrows=10)
cln = df2.iloc[8, 0]  # count CLNs as derivatives or not
k = df2.iloc[1, 0]  # get report date
rptDate = k if isinstance(k, datetime) else prior_month_end(datetime.today()).date()
ZARUSD = df2.iloc[9, 0]  # get exchange rate

s = "" if len(funds) == 1 else "s"
print(
    f"\n Reg 28 CS1 report{s} \n {rptDate.strftime('%A %d %b %Y')} \n ZAR/USD = {ZARUSD:.4f}\n \
{len(funds)} fund{s}: \n  {(',').join(funds)}\n"
)
print(
    f" {timediff(start_time_inputs, time.time())}: getting the CS1 report inputs completed\n"
)


# set constants

start_time = time.time()
print(f"\nSetting constants ...")

# get fund long names and types (UT, UCITS, ...)
nl = pd.read_excel(
    pthSttlmnt, sheet_name="Funds", index_col=None, header=0, usecols="A,B,E"
).dropna(subset=["Fund Code"])

# get Reg 28 classification codes and descriptions
static = pd.read_excel(pth_r28_lmts, sheet_name="Static", usecols="A,D").dropna()

# 13-months to maturity, convert it from datetime.date to datetime.datetime
td = rptDate + timedelta(days=397)
td = datetime.combine(td, datetime.min.time())

print(f" {timediff(start_time, time.time())} setting constants completed\n")

# dataframe the CS1 PARN holdings sheet with All, PARN, and NAV tabs
cs1_fname = os.path.join(
    pthTest, f"CS1 PARN holdings ({len(funds)}) {rptDate.strftime('%d%b%Y')}.xlsx"
)
holdings = pd.read_excel(cs1_fname, sheet_name="PARN", usecols="A:AQ")
navs = pd.read_excel(cs1_fname, sheet_name="NAVs", usecols="A:I")

# convert 'Maturity Date' column to datetime for the 13-month check to work
holdings["Maturity Date"] = pd.to_datetime(holdings["Maturity Date"])

# dataframe the Reg 28 CS1 reports
cs1_rpt_name = os.path.join(
    pthTest, f"Reg28 CS1 reports {rptDate.strftime('%d%b%Y')}.xlsx"
)
parn_r28 = pd.read_excel(cs1_rpt_name, sheet_name="CS1_All", usecols="A:M")

# identify funds not common to both dataframes
funds_not_classified = set(holdings["Entity ID"].unique()) ^ set(
    parn_r28["Entity Name"].unique()
)
none = "<empty>" if len(funds_not_classified) == 0 else (",").join(funds_not_classified)
print(
    f" {len(holdings['Entity ID'].unique()) - len(funds_not_classified)} \
out of the {len(holdings['Entity ID'].unique())} funds were classified \
with issuers_1.ipynb as at {rptDate.strftime('%d %B %Y')}.\n  Funds not \
classified ({len(funds_not_classified)}):\n  {none}"
)


# In[64]:


# loop through each fund
start_time_fund_loop = time.time()

# # TEST portfolios for the loop =======
# funds = pd.DataFrame({'Fund': ['HOLYPF','PGPGIF_C','HOLDINC'], 'Include CLNs?': ['Exclude CLNs','Exclude CLNs','Exclude CLNs']})
# # END OF TEST

open = 0  # open the CS1 file after saving it

no_holdings = []
no_derivatives = []

for fund in tqdm(funds):
    filename = os.path.join(
        pthTest, f"{fund} Reg28 CS1 Derivative Report {rptDate.strftime('%d%b%Y')}.xlsx"
    )

    # make a mini version of the holdings and PARN reports specific to the fund
    hold = holdings[holdings["Entity ID"] == fund]
    parn = parn_r28[parn_r28["Entity Name"] == fund]
    df = nl[nl["Fund Code"] == fund]
    fname = df.iloc[0, 1]
    ftyp = df.iloc[0, 2]
    zarusd = ZARUSD if (len(df) > 0) and (df.iloc[0, 2] in ["UCITS", "ICAV"]) else 1

    # determine the types of derivatives, in/excluding CLNs, in the fund
    if cln == "Include CLNs":
        list_of_dervs = parn["Derivative"].dropna().unique()
    else:
        list_of_dervs = (
            parn[parn["Derivative"] != "Credit-linked Note"]["Derivative"]
            .dropna()
            .unique()
        )

    # test for no holdings in the fund ...
    if (
        len(parn) == 0
    ):  # if the fund has no holdings (didnt exist at the time or was terminated)
        start_time_no_holdings = time.time()
        no_holdings.append(fund)  # collect the names of all funds with no holdings
        print(
            f"{fund}: ",
            f"{timediff(start_time_no_holdings, time.time())} no holdings in {fund} at {rptDate.strftime('%A %d %b %Y')}",
        )

    # # test for no derivatives in the fund ...
    # elif len(list_of_dervs) == 0: # if the fund has no derivatives
    #     start_time_no_derivatives = time.time()
    #     no_derivatives.append(fund) # collect the names of all funds with no derivatives
    #     print(''f'{fund}: ',f'{timediff(start_time_no_derivatives, time.time())} no derivatives in {fund} at {rptDate.strftime("%A %d %b %Y")}')

    # ... else produce the CS1 report
    else:
        start_time_cs1 = time.time()

        # (1) FUND SET-UP =====

        # get a new CS1 template
        wb = openpyxl.load_workbook(pth_r28_lmts)  # open the template
        sh = wb["CS1"]  # assign the sheet to be worked on

        # delete non-CS1 sheets from the template workbook
        shts = wb.sheetnames
        shts.remove("CS1")
        for sht in shts:
            del wb[sht]

        # make the sheet unique to the fund
        sh.title = f"{fund} Reg28 CS1 {rptDate.strftime('%d%b%Y')}"  # sheet name
        sh["A1"] = f"{fname} ({fund})"  # fund name and short code on the sheet
        sh["A3"] = (
            f"Regulation 28 Derivatives Report (Conduct Standard 1 of 2023) as at {rptDate.strftime('%d %b %Y')}"  # sheet title
        )

        # END OF (1) FUND SET-UP =====

        # (2) TEST FOR NO DERIVATIVES IN THE FUND =====
        if len(list_of_dervs) == 0:  # if the fund has no derivatives
            start_time_no_derivatives = time.time()
            no_derivatives.append(
                fund
            )  # collect the names of all funds with no derivatives
            sh[
                f"E{item_row(sh, 'first_row', 16)}"
            ].value = "NO DERIVATIVES HELD IN THE FUND"
            sh.cell(
                row=item_row(sh, "no_derivatives", 16), column=5
            ).font = cell_font_calibri_11_bold
            print(
                f"{fund}: ",
                f"{timediff(start_time_no_derivatives, time.time())} no derivatives in {fund} at {rptDate.strftime('%A %d %b %Y')}",
            )

        # END OF (2) TEST FOR NO DERIVATIVES IN THE FUND =====

        # (3) DERIVATIVE COVER METRICS from derv_checker_compiling_csv_new.ipynb =====
        # calculate the derivative cover metrics
        dervs = (
            (hold["Valuation First Level"] == "DERIVATIVES").sum()
            + (hold["Valuation First Level"] == "FORWARDS").sum()
            + (hold["Valuation First Level"] == "FORWARD RATE AGREEMENT").sum()
            + (hold["Security Type"] == "SWP").sum() / 3
            + (hold["Security Type"] == "SWAP").sum() / 3
        )  # number of derivatives, catering for the three legs of a FRA

        nav = hold["Sum of Market Value Income"].sum()  # fund NAV

        cash = hold[
            (hold["Valuation First Level"] == "CASH")
            & (hold["Security Type"] != "CIS")
            & (hold["Sub Security Type"] != "REPO")
            & (~hold["i Issue Name"].str.upper().str.contains("MARGIN"))
        ][
            "Current Exposure"
        ].sum()  # cash incl accruals, excl MMFs, excl margins, excl repo P/L

        mmfs = hold[
            (hold["PrimaryAssetID"] == "PRMFB3") | (hold["PrimaryAssetID"] == "PCMMB3")
        ]["Current Exposure"].sum()  # + \
        # hold[hold['PrimaryAssetID'] == 'PCMMB3']['Current Exposure'].sum() # money market funds, i.e., PMMF and PCMMF

        mmis = (
            hold[
                (hold["Valuation First Level"] == "MONEY MARKET")
                & (hold["Maturity Date"] < td)
                # #TypeError: Cannot compare Timestamp with datetime. Use ts == pd.Timestamp(date) or ts.date() == date instead.
                & (hold["Sub Security Type"] != "CLN")
            ]["Current Exposure"].sum()
            if ftyp == "UT" or ftyp == "ETF"
            else hold[hold["Valuation First Level"] == "MONEY MARKET"][
                "Current Exposure"
            ].sum()
        )  # money market instruments

        bonds = (
            hold[
                (hold["Valuation First Level"] == "BONDS")
                & (hold["Maturity Date"] < td)
                & (hold["Sub Security Type"] != "CLN")
            ]["Current Exposure"].sum()
            if ftyp == "UT" or ftyp == "ETF"
            else hold[hold["Valuation First Level"] == "BONDS"][
                "Current Exposure"
            ].sum()
        )  # bonds excl CLNs, excl > 13 months, else incl CLNs

        marg_jse = (
            hold[hold["PrimaryAssetID"] == "SAFEX"]["Current Exposure"].sum()
            + hold[hold["PrimaryAssetID"] == "VARMARG"]["Current Exposure"].sum()
        )  # JSE SAFEX and JSE VARMARG

        marg_otc = hold[
            (hold["i Issue Name"].str.upper().str.contains("MARGIN"))
            & (hold["PrimaryAssetID"] != "SAFEX")
            & (hold["PrimaryAssetID"] != "VARMARG")
        ]["Current Exposure"].sum()  # other margins, non-JSE margin accounts

        repo = (
            hold[hold["PrimaryAssetID"].str.upper().str.startswith("RPCO")][
                "Current Exposure"
            ].sum()
            + hold[hold["PrimaryAssetID"].str.upper().str.startswith("RPCA")][
                "Current Exposure"
            ].sum()
        )

        repo_gain = max(0, repo)  # net profit on repos

        repo_loss = min(0, repo)  # net loss on repos

        crry_derv = hold[
            hold["Valuation Second Level"]
            .str.upper()
            .str.contains("CURRENCY DERIVATIVES")
        ]["Current Exposure"].sum()

        fwds = min(
            0,
            hold[hold["Valuation Second Level"].str.upper().str.contains("FORWARDS")][
                "Current Exposure"
            ].sum(),
        )

        other_UTs = (
            0
            if ftyp == "UT"
            else hold[
                (hold["Security Type"] == "CIS")
                & (hold["Sub Security Type"] != "CSH")
                & (~hold["PrimaryAssetID"].isin(["PIMEVOA", "PIMIDFA"]))
            ]["Current Exposure"].sum()
        )  # other, non-MMF, UTs

        count_other_UTs = (
            0
            if ftyp == "UT"
            else hold[
                (hold["Security Type"] == "CIS")
                & (hold["Sub Security Type"] != "CSH")
                & (~hold["PrimaryAssetID"].isin(["PIMEVOA", "PIMIDFA"]))
            ]["Current Exposure"].count()
        )  # other, non-MMF, UTs

        other_ETFs = (
            0
            if ftyp == "UT"
            else hold[
                (hold["Security Type"] == "ETF") & (hold["Sub Security Type"] != "CSH")
            ]["Current Exposure"].sum()
        )  # other, non-MMF, ETFs

        count_other_ETFs = (
            0
            if ftyp == "UT"
            else hold[
                (hold["Security Type"] == "ETF") & (hold["Sub Security Type"] != "CSH")
            ]["Current Exposure"].count()
        )  # other, non-MMF, ETFs

        fras = min(
            0,
            hold[hold["Valuation First Level"] == "FORWARD RATE AGREEMENT"][
                "Original Nominal"
            ]
            .fillna(0)
            .dot(
                hold[hold["Valuation First Level"] == "FORWARD RATE AGREEMENT"][
                    r"Market Price /Yield"
                ].fillna(0)
            ),
        )

        ailf = (
            cash + mmfs + mmis + bonds + repo_gain + marg_jse + marg_otc
        )  # total assets in liquid form

        eqty_fut_frgn_mtm = -hold[
            (hold["Valuation Second Level"] == "Equity Derivatives")
            & (hold["CCY"] != "ZAR")
        ]["Sum of Market Value Income"].sum()  # mark-to-market p&l on equity futures

        bond_fut_frgn_mtm = -hold[
            (hold["Valuation Second Level"] == "Bond Derivatives")
            & (hold["CCY"] != "ZAR")
        ]["Sum of Market Value Income"].sum()  # mark-to-market p&l on bond futures

        trs_neg_mtm = -min(
            0,
            hold[(hold["Sub Security Type"] == "TRS")][
                "Sum of Market Value Income"
            ].sum(),
        )  # net negative mtm on OTC derivativesotcs      = fwds + min(0, trs_neg_mtm) + repo_loss + fras    # total OTC derivatives

        eqty_futs = max(
            0,
            hold[(hold["Valuation Second Level"] == "Equity Derivatives")][
                "Current Exposure"
            ].sum(),
        )  # equity futures

        bond_futs = max(
            0,
            hold[(hold["Valuation Second Level"] == "Bond Derivatives")][
                "Current Exposure"
            ].sum(),
        )  # bond futures

        frgn_futs = 0

        otcs = max(0, -fwds) + max(0, trs_neg_mtm) + repo_loss + fras

        # frgn_trs  = hold[(hold['Sub Security Type'] == 'TRS') & (hold['CCY'] != 'ZAR')]['Current Exposure'].sum() # profit or loss on total return swaps

        lstd_drvs = (
            eqty_futs + bond_futs + eqty_fut_frgn_mtm + bond_fut_frgn_mtm + trs_neg_mtm
        )

        eqty_fut_frgn_mtm = -hold[
            (hold["Valuation Second Level"] == "Equity Derivatives")
            & (hold["CCY"] != "ZAR")
        ]["Sum of Market Value Income"].sum()  # mark-to-market p&l on equity futures

        bond_fut_frgn_mtm = -hold[
            (hold["Valuation Second Level"] == "Bond Derivatives")
            & (hold["CCY"] != "ZAR")
        ]["Sum of Market Value Income"].sum()  # mark-to-market p&l on bond futures

        # leverage
        net_eff_exp = -min(
            0, hold[hold["Investment Type"] == "SYTH"]["Current Exposure"].sum() / nav
        )
        short_cash = -min(
            0,
            hold[
                (
                    (hold["Valuation First Level"] == "CASH")
                    & (~hold["PrimaryAssetID"].isin(["PRMFB3", "PCMMB3"]))
                )
            ]["Current Exposure"].sum()
            / nav,
        )  # cash excl MMFs < 0

        # END OF (3) DERIVATIVE COVER METRICS ==================================================

        # (4) HEADING header_asset_allocation "Asset allocation summary" =====

        # Synthetic cash market value, row 9 of CS1 report
        sh["B9"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "1.1")
                & (parn["CCY"] == "ZAR")
                & (parn["Investment Type"] == "SYTH")
            ]["End Market Value"].sum()
            / nav
            * 100
        )
        sh["C9"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "1.2")
                & (parn["CCY"] != "ZAR")
                & (parn["Investment Type"] == "SYTH")
            ]["End Market Value"].sum()
            / nav
            * 100
        )
        sh["D9"] = sh["B9"].value + sh["C9"].value

        # Synthetic cash effective exposure, row 9 of CS1 report
        sh["F9"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "1.1")
                & (parn["CCY"] == "ZAR")
                & (parn["Investment Type"] == "SYTH")
            ]["Closing Exposure PA"].sum()
            / nav
            * 100
        )
        sh["G9"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "1.2")
                & (parn["CCY"] != "ZAR")
                & (parn["Investment Type"] == "SYTH")
            ]["Closing Exposure PA"].sum()
            / nav
            * 100
        )
        sh["H9"] = sh["F9"].value + sh["G9"].value

        # Cash and money market instruments market value, row 10 of CS1 report
        sh["B10"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "1.1")
                & (parn["Investment Type"] != "FWD")
                & (parn["Investment Type"] != "SYTH")
            ]["End Market Value"].sum()
            / nav
            * 100
        )
        sh["C10"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "1.2")
                & (parn["Investment Type"] != "FWD")
                & (parn["Investment Type"] != "SYTH")
            ]["End Market Value"].sum()
            / nav
            * 100
        )
        sh["D10"] = sh["B10"].value + sh["C10"].value

        # Cash and money market instruments effective exposure, row 10 of CS1 report
        sh["F10"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "1.1")
                & (parn["Investment Type"] != "FWD")
                & (parn["Investment Type"] != "SYTH")
            ]["Closing Exposure PA"].sum()
            / nav
            * 100
        )
        sh["G10"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "1.2")
                & (parn["Investment Type"] != "FWD")
                & (parn["Investment Type"] != "SYTH")
            ]["Closing Exposure PA"].sum()
            / nav
            * 100
        )
        sh["H10"] = sh["F10"].value + sh["G10"].value

        # Net equity market value, row 11 of CS1 report
        sh["B11"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "3.1")
                & (parn["CCY"] == "ZAR")
            ]["End Market Value"].sum()
            / nav
            * 100
        )
        sh["C11"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "3.2")
                & (parn["CCY"] != "ZAR")
            ]["End Market Value"].sum()
            / nav
            * 100
        )
        sh["D11"] = sh["B11"].value + sh["C11"].value

        # Net equity effective exposure, row 11 of CS1 report
        sh["F11"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "3.1")
                & (parn["CCY"] == "ZAR")
            ]["Closing Exposure PA"].sum()
            / nav
            * 100
        )
        sh["G11"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "3.2")
                & (parn["CCY"] != "ZAR")
            ]["Closing Exposure PA"].sum()
            / nav
            * 100
        )
        sh["H11"] = sh["F11"].value + sh["G11"].value

        # Net property equity market value, row 12 of CS1 report
        sh["B12"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "4.1")
                & (parn["CCY"] == "ZAR")
            ]["End Market Value"].sum()
            / nav
            * 100
        )
        sh["C12"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "4.2")
                & (parn["CCY"] != "ZAR")
            ]["End Market Value"].sum()
            / nav
            * 100
        )
        sh["D12"] = sh["B12"].value + sh["C12"].value

        # Net property equity effective exposure, row 12 of CS1 report
        sh["F12"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "4.1")
                & (parn["CCY"] == "ZAR")
            ]["Closing Exposure PA"].sum()
            / nav
            * 100
        )
        sh["G12"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "4.2")
                & (parn["CCY"] != "ZAR")
            ]["Closing Exposure PA"].sum()
            / nav
            * 100
        )
        sh["H12"] = sh["F12"].value + sh["G12"].value

        # Net commodities market value, row 13 of CS1 report
        sh["B13"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "5.1")
                & (parn["CCY"] == "ZAR")
            ]["End Market Value"].sum()
            / nav
            * 100
        )
        sh["C13"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "5.2")
                & (parn["CCY"] != "ZAR")
            ]["End Market Value"].sum()
            / nav
            * 100
        )
        sh["D13"] = sh["B13"].value + sh["C13"].value

        # Net commodities effective exposure, row 13 of CS1 report
        sh["F13"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "5.1")
                & (parn["CCY"] == "ZAR")
            ]["Closing Exposure PA"].sum()
            / nav
            * 100
        )
        sh["G13"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "5.2")
                & (parn["CCY"] != "ZAR")
            ]["Closing Exposure PA"].sum()
            / nav
            * 100
        )
        sh["H13"] = sh["F13"].value + sh["G13"].value

        # Net bonds market value, row 14 of CS1 report
        sh["B14"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "2.1")
                & (parn["CCY"] == "ZAR")
            ]["End Market Value"].sum()
            / nav
            * 100
        )
        sh["C14"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "2.2")
                & (parn["CCY"] != "ZAR")
            ]["End Market Value"].sum()
            / nav
            * 100
        )
        sh["D14"] = sh["B14"].value + sh["C14"].value

        # Net bonds effective exposure, row 14 of CS1 report
        sh["F14"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "2.1")
                & (parn["CCY"] == "ZAR")
            ]["Closing Exposure PA"].sum()
            / nav
            * 100
        )
        sh["G14"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "2.2")
                & (parn["CCY"] != "ZAR")
            ]["Closing Exposure PA"].sum()
            / nav
            * 100
        )
        sh["H14"] = sh["F14"].value + sh["G14"].value

        # Net hedge funds market value, row 15 of CS1 report
        sh["B15"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "8.1")
                & (parn["CCY"] == "ZAR")
            ]["End Market Value"].sum()
            / nav
            * 100
        )
        sh["C15"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "8.2")
                & (parn["CCY"] != "ZAR")
            ]["End Market Value"].sum()
            / nav
            * 100
        )
        sh["D15"] = sh["B15"].value + sh["C15"].value

        # Net hedge funds effective exposure, row 15 of CS1 report
        sh["F15"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "8.1")
                & (parn["CCY"] == "ZAR")
            ]["Closing Exposure PA"].sum()
            / nav
            * 100
        )
        sh["G15"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "8.2")
                & (parn["CCY"] != "ZAR")
            ]["Closing Exposure PA"].sum()
            / nav
            * 100
        )
        sh["H15"] = sh["F15"].value + sh["G15"].value

        # Net private equity funds market value, row 16 of CS1 report
        sh["B16"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "9.1")
                & (parn["CCY"] == "ZAR")
            ]["End Market Value"].sum()
            / nav
            * 100
        )
        sh["C16"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "9.2")
                & (parn["CCY"] != "ZAR")
            ]["End Market Value"].sum()
            / nav
            * 100
        )
        sh["D16"] = sh["B16"].value + sh["C16"].value

        # Net private equity funds effective exposure, row 16 of CS1 report
        sh["F16"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "9.1")
                & (parn["CCY"] == "ZAR")
            ]["Closing Exposure PA"].sum()
            / nav
            * 100
        )
        sh["G16"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "9.2")
                & (parn["CCY"] != "ZAR")
            ]["Closing Exposure PA"].sum()
            / nav
            * 100
        )
        sh["H16"] = sh["F16"].value + sh["G16"].value

        # Net other assets market value, row 17 of CS1 report
        sh["B17"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:4] == "10.1")
                & (parn["CCY"] == "ZAR")
            ]["End Market Value"].sum()
            / nav
            * 100
        )
        sh["C17"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:4] == "10.2")
                & (parn["CCY"] != "ZAR")
            ]["End Market Value"].sum()
            / nav
            * 100
        )
        sh["D17"] = sh["B17"].value + sh["C17"].value

        # Net other assets effective exposure, row 17 of CS1 report
        sh["F17"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:4] == "10.1")
                & (parn["CCY"] == "ZAR")
            ]["Closing Exposure PA"].sum()
            / nav
            * 100
        )
        sh["G17"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:4] == "10.2")
                & (parn["CCY"] != "ZAR")
            ]["Closing Exposure PA"].sum()
            / nav
            * 100
        )
        sh["H17"] = sh["F17"].value + sh["G17"].value

        # Foreign FX market value, row 18 of CS1 report
        sh["B18"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "1.1")
                & (parn["CCY"] == "ZAR")
                & (parn["Investment Type"] == "FWD")
            ]["End Market Value"].sum()
            / nav
            * 100
        )
        sh["C18"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "1.2")
                & (parn["CCY"] == "ZAR")
                & (parn["Investment Type"] == "FWD")
            ]["End Market Value"].sum()
            / nav
            * 100
        )
        sh["D18"] = sh["B18"].value + sh["C18"].value

        # Foreign FX effective exposure, row 18 of CS1 report
        sh["F18"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "1.1")
                & (parn["CCY"] == "ZAR")
                & (parn["Investment Type"] == "FWD")
            ]["Closing Exposure PA"].sum()
            / nav
            * 100
        )
        sh["G18"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "1.2")
                & (parn["CCY"] == "ZAR")
                & (parn["Investment Type"] == "FWD")
            ]["Closing Exposure PA"].sum()
            / nav
            * 100
        )
        sh["H18"] = sh["F18"].value + sh["G18"].value

        # Foreign exchange market value, row 19 of CS1 report
        sh["B19"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "1.1")
                & (parn["CCY"] == "ZAR")
                & (parn["Investment Type"] == "FWD")
            ]["End Market Value"].sum()
            / nav
            * 100
        )
        sh["C19"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "1.2")
                & (parn["CCY"] != "ZAR")
                & (parn["Investment Type"] == "FWD")
            ]["End Market Value"].sum()
            / nav
            * 100
        )
        sh["D19"] = sh["B19"].value + sh["C19"].value

        # Foreign exchange effective exposure, row 19 of CS1 report
        sh["F19"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "1.1")
                & (parn["CCY"] == "ZAR")
                & (parn["Investment Type"] == "FWD")
            ]["Closing Exposure PA"].sum()
            / nav
            * 100
        )
        sh["G19"] = (
            parn[
                (parn["Reg 28 Classification"].astype(str).str[:3] == "1.2")
                & (parn["CCY"] != "ZAR")
                & (parn["Investment Type"] == "FWD")
            ]["Closing Exposure PA"].sum()
            / nav
            * 100
        )
        sh["H19"] = sh["F19"].value + sh["G19"].value

        # format the Asset allocation summary numbers
        for col in [chr(i) for i in range(ord("B"), ord("H") + 1)]:
            for row in range(
                item_row(sh, "synthetic_cash", 16), item_row(sh, "fx_foreign", 16) + 1
            ):
                sh[f"{col}{row}"].number_format = cell_format

        # calculate Asset allocation summary total market value and total effective exposure, row 20 of CS1 report
        for col in [
            "B",
            "C",
            "D",
            "F",
            "G",
            "H",
        ]:  # for each asset allocation summary column
            total = 0
            for row in range(9, 20):  # from 'Synthetic cash' row to 'FX foreign row'
                total += sh[f"{col}{row}"].value
                sh[f"{col}{item_row(sh, 'total_asset_allocation', 16)}"] = total

        # END OF (4) HEADING header_asset_allocation "Asset allocation summary"

        # (5) HEADING header_long_cover_info "Long cover information" =====
        # insert fund values on the sheet
        sh[f"C{item_row(sh, 'nav', 16)}"] = nav * zarusd
        sh[f"C{item_row(sh, 'long_cover', 16)}"] = (
            ailf
        ) * zarusd  # row long_cover, column "C"
        sh[f"D{item_row(sh, 'long_cover', 16)}"] = (
            (ailf) / nav * 100
        )  # row long_cover, column "D"
        # sh[f"C{item_row(sh, 'long_cover', 16)}"] = (ailf + other_UTs + other_ETFs) * zarusd       # row long_cover, column "C"
        # sh[f"D{item_row(sh, 'long_cover', 16)}"] = (ailf + other_UTs + other_ETFs) / nav * 100    # row long_cover, column "D"
        if (ftyp == "UCITS") or (ftyp == "ICAV"):
            sh[f"A{item_row(sh, 'ucits_note', 16)}"] = (
                f"*** the securities in the fund include deposits, transferable securities and money market \
instruments, all of which are eligible for purposes of derivative cover under EU UCITS regulations"  # row ucits_note, column "A"
            )
            sh.insert_rows(item_row(sh, "ucits_note", 16) + 1)
        else:
            sh[f"A{item_row(sh, 'ucits_note', 16)}"] = ""

        # create a dictionary of derivative descriptions and values
        # dervs = {'d': ['Negative MtM on TRSes',               trs_neg_mtm],
        #          'c': ['Bought futures or calls & sold puts', eqty_futs + bond_futs + eqty_fut_frgn_mtm + bond_fut_frgn_mtm +  max(0, crry_derv)],
        #          'b': ['Negative MtM OTC derivatives',        otcs, trs_neg_mtm],
        #          'a': ['Sold currency futures',               max(0, crry_derv)]}
        dervs = {
            "b": ["Negative MtM OTC derivatives", otcs, trs_neg_mtm],
            "a": [
                "Bought futures or calls & sold puts",
                eqty_futs
                + bond_futs
                + eqty_fut_frgn_mtm
                + bond_fut_frgn_mtm
                + max(0, crry_derv),
            ],
        }

        derv_rows = sum(
            [
                -max(0, crry_derv) != 0,
                otcs != 0,
                eqty_futs + bond_futs + eqty_fut_frgn_mtm + bond_fut_frgn_mtm != 0,
                trs_neg_mtm != 0,
            ]
        )

        # insert and populate a row below "Long cover used (net)" for each non-zero effective exposure derivative class
        for key, value in dervs.items():
            # print(key, value[1])
            if value[1] != 0:
                derv_row = (
                    item_row(sh, "total_long_cover_used", 16) + 1
                )  # row total_long_cover_used
                sh.insert_rows(derv_row, 1)
                sh[f"A{derv_row}"] = f"   {value[0]}"
                sh[f"C{derv_row}"] = value[1] * zarusd
                sh[f"D{derv_row}"] = value[1] / nav * 100

        # calculate long cover used (net) totals
        total = 0
        rw = item_row(sh, "total_long_cover_used", 16)
        for row in range(rw, rw + derv_rows + 1):
            total += sh[f"C{row}"].value if sh[f"C{row}"].value is not None else 0
        sh[f"C{rw}"] = (
            total  # row total_long_cover_used, column "C", long cover available in ZAR
        )
        sh[f"D{rw}"] = (
            total / nav / zarusd * 100
        )  # row total_long_cover_used, column "D", long cover available as a % of NAV

        # calculate long cover available
        sh[f"C{item_row(sh, 'long_cover_available', 16)}"] = (
            sh[f"C{item_row(sh, 'header_long_cover_info', 16) + 1}"].value
            - sh[f"C{item_row(sh, 'total_long_cover_used', 16)}"].value
        )  # row long_cover_available

        sh[f"D{item_row(sh, 'long_cover_available', 16)}"] = (
            sh[f"C{item_row(sh, 'long_cover_available', 16)}"].value
            / nav
            / zarusd
            * 100
        )

        # format a range border to the derivative long cover numbers
        if derv_rows != 0:
            range_border(
                sh,
                3,
                4,
                item_row(sh, "total_long_cover_used", 16) + 1,
                item_row(sh, "long_cover_available", 16) - 2,
            )  # from utilities.py

        # format the derivative long cover numbers
        for col in ["C", "D"]:
            for row in range(
                item_row(sh, "long_cover", 16), item_row(sh, "long_cover_available", 16)
            ):
                sh[f"{col}{row}"].number_format = cell_format

        # END OF (5) HEADING header_long_cover_info "Long cover information" =====

        # (6) HEADING header_derivatives "List of derivatives *" =====
        # list of derivatives
        d = item_row(sh, "derv_list", 16)
        # m = parn[(parn['Derivative'].notna()) & (parn['Investment Type'] != 'SYTH')]
        if cln == "Include CLNs":
            m = parn[(parn["Derivative"].notna()) & (parn["Investment Type"] != "SYTH")]
        else:
            m = parn[
                (parn["Derivative"] != "Credit-linked Note")
                & (parn["Derivative"].notna())
                & (parn["Investment Type"] != "SYTH")
            ]
        # if cln == 'Include CLNs':
        #     m = parn[(parn['Investment Type'] != 'SYTH')]
        # else:
        #     m = parn[(parn['Derivative'] != 'Credit-linked Note') & (parn['Investment Type'] != 'SYTH')]
        k = m["Primary Asset ID"].unique()  # unique IDs in subset of derivatives
        # sh.insert_rows(d + 1, len(k) - 1)
        sh.insert_rows(d + 1, len(k))

        # for i, pa_id in tqdm(enumerate(k, start = d)):
        for i, pa_id in enumerate(k, start=d):
            g = m[m["Primary Asset ID"] == pa_id]
            sh[f"A{i}"] = i - d + 1
            sh[f"B{i}"] = pa_id
            sh[f"C{i}"] = g["i Issue Name"].iloc[0]  # Issue Description
            sh[f"D{i}"] = g["Reg 28 Classification"].iloc[0]  # Reg28Code
            sh[f"E{i}"] = static[static["Reg 28 Classification"] == sh[f"D{i}"].value][
                "Reg 28 Description"
            ].item()  # Reg28Name
            sh[f"F{i}"] = "Local" if g["CCY"].iloc[0] == "ZAR" else "Foreign"  # Region
            sh[f"G{i}"] = (
                "Exchange"
                if g["Counterparty"].iloc[0] in ["Exchange", "JSE"]
                else "Bank"
            )  # Issuer Type
            sh[f"H{i}"] = g["Issuer"].iloc[0]  # Reg28 Issuer Name
            sh[f"I{i}"] = (
                "Exchange"
                if g["Counterparty"].iloc[0] in ["Exchange", "JSE"]
                else "OTC"
            )  # Exchange
            sh[f"J{i}"] = g["Counterparty"].iloc[0]  # Counterparty
            sh[f"K{i}"] = g["CCY"].iloc[0]  # Local / base currency
            sh[f"L{i}"] = "ZAR"  # Currency of report
            sh[f"M{i}"] = (
                g["Closing Exposure PA"].astype(float).sum() * zarusd
            )  # Effective Exposure in ZAR
            sh[f"N{i}"] = (
                g["Closing Exposure PA"].astype(float).sum() / nav * 100
            )  # Effective Exposure % of Portfolio
            sh[f"O{i}"] = g["Derivative"].iloc[0]  # Derivative type"

            # format alignmnent and numbers of list of derivatives
            sh[f"A{i}"].number_format = "#,##0"
            for col in ["M", "N"]:
                sh[f"{col}{i}"].number_format = cell_format
            for col in ["B", "C", "H", "J", "O"]:
                sh[f"{col}{i}"].alignment = Alignment(wrapText=True)

            for col in [chr(i) for i in range(ord("A"), ord("P") + 1)]:
                sh[f"{col}{i}"].border = cell_border

        # END OF (6) HEADING header_derivatives "List of derivatives *" =====

        if len(list_of_dervs) == 0:
            # delete locator column
            sh.delete_cols(16, 2)
        else:
            # (7) HEADING header_derivative_types "Derivative Counterparty Exposure **" =====

            # list of counterparties
            if cln == "Include CLNs":
                parn_cntpties = parn[
                    (parn["Counterparty"].notna())
                    & (parn["Derivative"].notna())
                    & (parn["Investment Type"] != "SYTH")
                ]
            else:
                parn_cntpties = parn[
                    (parn["Counterparty"].notna())
                    & (parn["Derivative"].notna())
                    & (parn["Investment Type"] != "SYTH")
                    & (parn["Derivative"] != "Credit-linked Note")
                ]

            counterparties = parn_cntpties[
                "Counterparty"
            ].unique()  # unique counterparties in subset of counterparties
            if len(counterparties) > 1:
                sh.insert_rows(
                    item_row(sh, "header_derivative_types", 16) + 1,
                    len(counterparties) - 1,
                )

            # populate the headings of the 'Derivative Counterparty Exposure **' columns
            for col_idx, derv in enumerate(list_of_dervs, start=4):
                # print(col_idx, derv)
                sh.cell(
                    row=item_row(sh, "header_derivative_types", 16),
                    column=col_idx,
                    value=derv,
                )
                sh.cell(
                    row=item_row(sh, "header_derivative_types", 16), column=col_idx
                ).alignment = Alignment(horizontal="center", vertical="top")
                sh.cell(
                    row=item_row(sh, "header_derivative_types", 16), column=col_idx
                ).font = Font(bold=True, size=10, name="Arial")
                sh.cell(
                    row=item_row(sh, "header_derivative_types", 16), column=col_idx
                ).border = cell_border
                sh.cell(
                    row=item_row(sh, "header_derivative_types", 16), column=col_idx
                ).fill = PatternFill(
                    start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid"
                )

                # per counterparty and per derivative totals
                for i, ctpty in enumerate(
                    counterparties,
                    start=item_row(sh, "header_derivative_types", 16) + 1,
                ):
                    # print(i, ctpty)
                    j = parn_cntpties[parn_cntpties["Counterparty"] == ctpty]
                    sh.cell(row=i, column=col_idx).value = (
                        abs(
                            j[j["Derivative"] == derv]["Closing Exposure PA"]
                            .astype(float)
                            .sum()
                        )
                        * zarusd
                    )
                    sh.cell(row=i, column=1, value=ctpty)
                    sh.cell(row=i, column=1).alignment = Alignment(wrapText=True)
                    sh.cell(row=i, column=2).value = (
                        abs(
                            j[j["Counterparty"] == ctpty]["Closing Exposure PA"]
                            .astype(float)
                            .sum()
                        )
                        * zarusd
                    )
                    sh.cell(row=i, column=3).value = (
                        abs(
                            j[j["Counterparty"] == ctpty]["Closing Exposure PA"]
                            .astype(float)
                            .sum()
                        )
                        / nav
                        * 100
                    )

            # total net counterparty exposure
            total = 0
            for row in range(
                item_row(sh, "header_derivative_types", 16) + 1,
                item_row(sh, "total_counterparties", 16),
            ):
                total += 0 if sh[f"B{row}"].value is None else sh[f"B{row}"].value
            sh[f"B{item_row(sh, 'total_counterparties', 16)}"] = (
                total  # net counterparty exposure in ZAR
            )
            sh[f"C{item_row(sh, 'total_counterparties', 16)}"] = (
                total / zarusd / nav * 100
            )  # net counterparty exposure as a % of NAV

            # format numbers and borders of 'Derivative Counterparty Exposure **'
            for col_idx in range(2, 4 + len(list_of_dervs)):
                for row_idx in range(
                    item_row(sh, "header_derivative_types", 16) + 1,
                    item_row(sh, "header_derivative_types", 16)
                    + 1
                    + len(counterparties),
                ):
                    sh.cell(row=row_idx, column=col_idx).number_format = cell_format
                    sh.cell(row=row_idx, column=col_idx).border = cell_border
                    sh.cell(row=row_idx, column=col_idx).alignment = Alignment(
                        horizontal="right"
                    )

            # END OF (7) HEADING header_derivative_types "Derivative Counterparty Exposure **" =====

            # (8) TOUCH-UPS to the sheet

            # apply font Calibri, size 11 to all cells on the sheet
            for row_idx in range(
                item_row(sh, "first_row", 16), item_row(sh, "last_row", 16) + 1
            ):
                for col_idx in range(1, 16):
                    sh.cell(row_idx, col_idx).font = cell_font_calibri_11

            # apply bold fint to each of the headings and totals on the sheet
            bold_titles = [
                "first_row",
                "investment_manager",
                "report_title",
                "nav",
                "header_asset_allocation",
                "total_asset_allocation",
                "title_long_cover_info",
                "header_long_cover_info",
                "list_of_derivatives",
                "header_derivatives",
                "derivative_counterparty_exposure",
                "header_derivative_types",
                "total_counterparties",
                "disclosures",
            ]
            for title in bold_titles:
                for col_idx in range(1, 16):
                    sh.cell(
                        item_row(sh, title, 16), col_idx
                    ).font = cell_font_calibri_11_bold

            # def rows_align_height(worksheet, row_from, row_to, col, normal_height, text_test_lengthcol_right, new_height):
            rows_align_height(sh, 1, sh.max_row, 3, 15.75, 30, 60)

            # delete locator column
            sh.delete_cols(16, 2)

            # END OF (8) TOUCH-UPS to the sheet

        # (9) SAVING THE FILE =====

        stop_time_cs1 = time.time()

        # add the (non-lookthrough) Reg28 classification sheet to the workbook
        start_time_r28 = time.time()
        shR28 = wb.create_sheet("Reg28")

        for col_idx, col_name in enumerate(
            parn.columns, 1
        ):  # write the dataframe headers for the Reg28 sheet
            shR28.cell(row=1, column=col_idx, value=col_name)

        start_row = 2  # write the R28 sheet contents; if headers were written to row 1, start data from row 2
        for r_idx, row in enumerate(parn.itertuples(index=False), start=start_row):
            for c_idx, value in enumerate(row, 1):
                shR28.cell(row=r_idx, column=c_idx, value=value)

        # save the CS1 file in the py test folder and close the openpyxl workbook
        # sh.delete_cols(16)
        wb.save(filename)
        wb.close
        print(
            f" {fund} ({'incl' if cln == 'Include CLNs' else 'excl'} CLNs): {timediff(start_time_cs1, stop_time_cs1)} CS1 sheet, \
{timediff(start_time_r28, time.time())} Reg28 sheet, {filename}\n"
        )

        # END OF (9) SAVING THE FILE =====

        if open == 1:
            open_xl_file(filename)

s = "s" if len(no_holdings) != 1 else ""
print(
    f"\nFund{s} with no holdings at \
{rptDate.strftime('%A %d %b %Y')} \
({len(no_holdings)}):\n {(',').join(no_holdings)}"
)

s = "s" if len(no_derivatives) != 1 else ""
print(
    f"Fund{s} with no derivatives at \
{rptDate.strftime('%A %d %b %Y')} \
({len(no_derivatives)}): {(',').join(no_derivatives)}\n"
)
# print(f'{timediff(start_time_fund_loop, time.time())} CS1 report roundtrip time for {len(funds)} funds at {rptDate.strftime("%d %b %Y")}')


# In[65]:


# combine PARN, Reg28, no_holdings, and no_derivatives dataframes into one workbook for review

start_time_combined_workbook = time.time()
print(
    f"Combining workbook for {len(funds)} funds at {rptDate.strftime('%d %b %Y')} ..."
)

combined_name = os.path.join(
    pthTest, f"Combined CS1 reports {rptDate.strftime('%d%b%Y')}.xlsx"
)  # assign the file name
writer = pd.ExcelWriter(
    combined_name, engine="xlsxwriter"
)  # instantiate a sheet writer with file name

holdings.to_excel(
    writer, index=False, sheet_name=f"Holdings ({len(holdings)})"
)  # write the summary sheet
navs.to_excel(
    writer, index=False, sheet_name=f"NAVs ({len(navs)})"
)  # write the fund holdings sheet
parn_r28.to_excel(
    writer, index=False, sheet_name=f"R28 ({len(parn_r28)})"
)  # write the derivative deltas sheet
pd.DataFrame(no_holdings, columns=["No Holdings"]).to_excel(
    writer, index=False, sheet_name=f"No holdings ({len(no_holdings)})"
)
pd.DataFrame(no_derivatives, columns=["No Derivatives"]).to_excel(
    writer, index=False, sheet_name=f"No derivatives ({len(no_derivatives)})"
)

writer.close()  # https://pandas.pydata.org/docs/reference/api/pandas.ExcelWriter.html   class for writing DataFrame objects into excel sheets

print(
    f"\n {timediff(start_time_combined_workbook, time.time())} combining workbook for {len(funds)} funds at {rptDate.strftime('%d %b %Y')}\n"
)

print(
    f" {timediff(start_time_fund_loop, time.time())} CS1 report roundtrip time for {len(funds)} funds at {rptDate.strftime('%d %b %Y')}\n"
)

# STEPS

# (1) FUND SET-UP =====
# (2) TEST FOR NO DERIVATIVES IN THE FUND =====
# (3) DERIVATIVE COVER METRICS from derv_checker_compiling_csv_new.ipynb =====
# (4) HEADING header_asset_allocation "Asset allocation summary" ====
# (5) HEADING header_long_cover_info "Long cover information" =====
# (6) HEADING header_derivatives "List of derivatives *" =====
# (7) HEADING header_derivative_types "Derivative Counterparty Exposure **" =====
# (8) TOUCH-UPS to the sheet
# (9) SAVING THE FILE
