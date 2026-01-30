#!/usr/bin/env python
# coding: utf-8

print("\n\n##########################")
print("#    START reg28_ib.py    #")
print("##########################\n\n")

import time
import pandas as pd
import os
import openpyxl
import copy  # "AttributeError: Style objects are immutable and cannot be changed. Reassign the style with a copy"
from openpyxl.styles import (
    NamedStyle,
    Alignment,
    Font,
)  # https://openpyxl.readthedocs.io/en/stable/styles.html
from openpyxl.styles.borders import Border, Side
from openpyxl.cell import (
    Cell,
)  # https://stackoverflow.com/questions/42215933/apply-wrap-text-to-all-cells-using-openpyxl
from tqdm import tqdm
from utilities import timediff, item_row
from constants import (
    supercats,
    superhens,
    pthPy,
    pthSttlmnt,
    pthReports,
    pthTest,
    pth_r28_lmts,
)


# get report data
def get_inputs():
    start_time = time.time()
    global funds, date, syth, nl, fund_list
    # print('Getting the Schedule IB report inputs ...')

    # fund long name lookup
    nl = pd.read_excel(
        pthSttlmnt, sheet_name="Funds", index_col=None, header=0, usecols="A,B"
    ).dropna(subset=["Fund Code"])
    funds = pd.read_excel(pthPy, sheet_name="r28_ib", usecols="A").dropna()
    funds["Funds"] = funds["Funds"].apply(
        str.upper
    )  # https://sparkbyexamples.com/pandas/pandas-convert-column-to-uppercase/
    fund_list = (",").join(funds["Funds"])
    date = pd.read_excel(pthPy, sheet_name="r28_ib", usecols="C", nrows=1).iloc[0, 0]
    syth = pd.read_excel(pthPy, sheet_name="r28_ib", usecols="D", nrows=1).iloc[
        0, 0
    ]  # indicator to show 'SYTH' or not
    # fund           = funds.iloc[0,0]

    s = "s" if len(funds) != 1 else ""
    print(
        f"Schedule IB report{s} as at {date.strftime('%A %d %b %Y')} for {len(funds)} fund{s}:\n {fund_list}\n"
    )
    # print('\n', f'{timediff(start_time, time.time())}: getting the Schedule IB report inputs with pd.read_excel() completed')


# function to get Reg 28 into a dataframe from which values will be looked up for the schedule
def r28_df():
    start_time = time.time()
    global rgAll, rg, r28, ctgs
    # print('Setting up dataframe of values for the schedule ...')

    # dataframe the Reg 28 classifications report
    rgAll = pd.read_excel(
        os.path.join(pthReports, f"{fund} Reg28 {date.strftime('%d%b%Y')}.xlsx")
    )

    # remove cash contra / synthetic cash rows in-place
    rg = rgAll[rgAll["Investment Type"] != "SYTH"]
    syth = len(rg[rg["Investment Type"] == "SYTH"])

    # sort reg28 by classification and issuer
    # https://stackoverflow.com/questions/33165734/update-index-after-sorting-data-frame
    # https://stackoverflow.com/questions/17141558/how-to-sort-a-pandas-dataframe-by-two-or-more-columns
    rg = rg.sort_values(
        by=["Reg 28 Classification", "Issuer", "Primary Asset ID"],
        ascending=[True, False, False],
        ignore_index=True,
    )

    # dataframe the complete list of categories and their limits
    limits = pd.read_excel(
        pth_r28_lmts, sheet_name="Static", usecols="A:C", index_col=None
    ).dropna()  # complete categories
    # len(limits)

    # merge the Reg 28 and the limits dataframes
    r28 = rg.merge(
        limits,
        left_on="Reg 28 Classification",
        right_on="Reg 28 Classification",
        how="left",
    )

    # add a combined instrument name column (ID + Description)
    # https://stackoverflow.com/questions/19377969/combine-two-columns-of-text-in-pandas-dataframe
    r28["Instr"] = r28["Primary Asset ID"] + " - " + r28["i Issue Name"]

    # get the unique item categories in Reg 28
    ctgs = r28["Reg 28 Classification"].unique()
    len(ctgs)

    # rename the columns of the merged dataframe
    headings = {
        "Reg 28 Classification": "R28C",
        "End Market Value": "EMV",
        "Percentage of Market Value": "PMV",
        "Closing Exposure PA": "CEPA",
        "Issuer Limit": "IL",
        "Aggr Limit": "AL",
    }
    r28 = r28.rename(columns=headings)

    # print('\n', f'{fund} on {date.strftime("%A %d %b %Y")}: {syth} contra{"s" if syth > 1 else ""} removed, ',
    # f'{len(ctgs)} sub-categor{"y" if len(ctgs) == 1 else "ies"}:', '\n', (', ').join(ctgs), '\n')

    # print(f'{timediff(start_time, time.time())}: setting up dataframe of values for the schedule completed: ')


# constants_2 - static for formatting security values and percentages

nmbr = '#,##0.00 ;-#,##0.00 ;"- "'  # https://support.microsoft.com/en-us/office/number-format-codes-5026bbd6-04bc-48cd-bf33-80f18b4eae68
fnt = "Calibri"
sz = 12

# cell styles for column 'C', 'Security Description', to be used in the workbook
cell_style_K = NamedStyle(
    name="cell_style_K"
)  # alignment style for column 'C' of the schedule
cell_style_K.alignment = Alignment(vertical="top", wrap_text=True, indent=3)

# cell styles for column 'D', 'Limit (%)', to be used in the workbook
cell_style_D = NamedStyle(
    name="cell_style_D"
)  # alignment style for column 'D', 'Limit (%)', of the schedule
cell_style_D.alignment = Alignment(vertical="top", horizontal="center")
cell_style_D.font = Font(name=fnt, size=sz)

# cell styles for column headings on the schedule
cell_style_heads = NamedStyle(
    name="cell_style_heads"
)  # cellstyle for column headings on the schedule
cell_style_heads.alignment = Alignment(vertical="top", horizontal="center")
cell_style_heads.font.bold = True
cell_style_heads.border = Border(bottom=Side(style="thin"))

# cell styles for numbers on the schedule
cell_style_numbers = NamedStyle(name="cell_style_numbers")
cell_style_numbers.number_format = nmbr
cell_style_numbers.alignment = Alignment(horizontal="right", vertical="top")
cell_style_numbers.font = Font(name=fnt, size=sz)

template_with_sheets = (
    0  # use the Schedule IB template which does not have other sheets
)


# cell styles for openpyxl workbooks
def open_wb():
    global wb, sh, cell_style_heads

    start_time = time.time()
    # print('Opening the SchIB template as a workbook ...')

    if template_with_sheets == 1:
        wb = openpyxl.load_workbook(pth_r28_lmts)  # open the Reg Schedule IB template
        del wb["Static"]  # delete unneccesary sheets from the template workbook
        del wb["Tbl2"]  # delete unneccesary sheets from the template workbook
        del wb["CS1"]  # delete unneccesary sheets from the template workbook
    else:
        wb = openpyxl.load_workbook(pth_r28_lmts)  # open the Reg Schedule IB template

    sh = wb["SchIB"]  # assign the sheet to be worked on
    sh.title = f"{fund} SchIB {date.strftime('%d%b%Y')}"  # set tab name of IB sheet

    # enter fund name and report date on IB sheet
    sh["A2"] = f"{nl[nl['Fund Code'] == fund].iat[0, 1]} ({fund})"  # fund long name
    sh["A4"] = (
        f"Assets held in compliance with Regulation 28 as at {date.strftime('%d %B %Y')}"
    )

    # add cell styles to be applied in the workbook
    wb.add_named_style(
        cell_style_K
    )  # add cell styles for column 'C', 'Security Description', to be used in the workbook
    # using 'cell_style_C' yields the error "ValueError: Style cell_style_C exists already"
    wb.add_named_style(
        cell_style_D
    )  # add cell styles for column 'D', 'Limit (%)', to be used in the workbook
    wb.add_named_style(
        cell_style_heads
    )  # add cell styles for the 'Fair Value' and the 'Limit (%)' headings
    wb.add_named_style(
        cell_style_numbers
    )  # add cell styles for the numbers in columns 'E', and 'F'

    # print(f'{timediff(start_time, time.time())}: opening the SchIB template as a workbook and assigning its cell styles completed')


# function to copy the schedule template and then update static values on it
def paste_nav():
    start_time = time.time()
    # print('Updating static values on the schedule ...')

    # enter fund NAV on the schedule
    sh["F8"] = r28["EMV"].sum()  # TOTAL NAV
    sh["F13"] = r28["EMV"].sum()  # TOTAL NAV
    sh["E" + str(item_row(sh, "TOTAL", 7))] = r28["EMV"].sum()  # TOTAL NAV
    sh["F" + str(item_row(sh, "TOTAL", 7))] = 100  # TOTAL NAV

    # print(f'{timediff(start_time, time.time())}: updating static values on the schedule completed')


# populate instruments for the 3(f), 3(g), 3(h), and 3(i) subtotals
def fghi():  # 3(f), 3(g), 3(h), and 3(i) subtotals and their constituents and remove the contras, if necessary
    start_time = time.time()

    # deal with the contras
    if (
        syth != "Market Value"
    ):  # if 'Effective Exposure' was seleteced, add the contra categories to be summed
        # https://phoenixnap.com/kb/python-add-to-dictionary
        superhens["1"]["1.1"].update({"1.1(a) contra"})
        superhens["1"]["1.2"].update({"1.2(a) contra"})
        supercats["3(h)"]["3(h) 1.1"].update({"1.1(a) contra"})
        supercats["3(i)"].update({"3(i) contra": {"1.2(a) contra"}})
    else:  # else if 'Market Value' was selected, delete the three contra line items on the schedule
        contras = ["1.1(a) contra", "1.2(a) contra"]
        for contra in contras:
            sh.delete_rows(
                item_row(sh, contra, 7), 4
            )  # delete the '1.1/1.2(a) contra' line items and three subsequent rows
        sh.delete_rows(
            item_row(sh, "3(i) contra", 7), 1
        )  # delete the '3(i) contra' line item, but only one row

    # sum and populate tthe 3(f,g,h,i) sub-totals
    # iterating through a nested dictionary - https://www.programiz.com/python-programming/nested-dictionary
    spacer2 = 1
    for (
        supercat,
        values,
    ) in supercats.items():  # key supercat = '3(f)', ... ; supercats.items() = {'3(f) 2.1(e)(ii)', '3(f) 3.1(b)', ...}
        # sum totals for super category 3(f,g,h,i) line items below TOTAL
        sigmumEMV = 0
        sigmumPMV = 0
        for (
            schd_row,
            lists,
        ) in (
            values.items()
        ):  # key schd_row = '3(f) 2.1(e)(ii)', ...; values = ['2.1(e)(ii), '...']
            # print(supercat + ':->' + schd_row)          # POINTER
            k = pd.DataFrame()  # https://stackoverflow.com/questions/16597265/appending-to-an-empty-dataframe-in-pandas
            sigmaEMV = 0
            sigmaPMV = 0
            for cat in lists:
                # print('\t\t' + cat)                     # POINTER
                k = pd.concat(
                    [k, r28[r28["R28C"] == cat]]
                )  # https://stackoverflow.com/questions/16597265/appending-to-an-empty-dataframe-in-pandas
                k.reset_index(
                    drop=True, inplace=True
                )  # re-index in place + drop the newly inserted index

                sigmaEMV += r28[r28["R28C"] == cat][
                    "EMV"
                ].sum()  # sum over each category within a line item below TOTAL
                sigmaPMV += r28[r28["R28C"] == cat]["PMV"].sum()
                # print('\nCategory :', cat, k)

            sh[item_row(sh, schd_row, 7)][4].value = sigmaEMV  # assign
            sh[item_row(sh, schd_row, 7)][5].value = sigmaPMV

            sigmumEMV += sigmaEMV  # cumulatively track sub-category sub-totals
            sigmumPMV += sigmaPMV

        sh[item_row(sh, supercat, 7)][
            4
        ].value = sigmumEMV  # assign 3(f,g,h,i) line item sub-totals
        sh[item_row(sh, supercat, 7)][5].value = sigmumPMV

        #     for row in k.iterrows():
        #         sh[item_row(sh, schd_row, 7)][4].value  = r28[r28['R28C'] == cat]['EMV'].sum()
        #         sh[item_row(sh, schd_row, 7)][5].value  = r28[r28['R28C'] == cat]['PMV'].sum()
        #     # for each category, add the market values and pecentage values
        # for row in k.iterrows():
        #     sh[item_row(sh, schd_row, 7)][4].value  = r28[r28['R28C'] == cat]['EMV'].sum()
        #     sh[item_row(sh, schd_row, 7)][5].value  = r28[r28['R28C'] == cat]['PMV'].sum()

        # print(f'item_row() for schd_row {schd_row} is {item_row(sh, schd_row, 7)}')
        # sh.insert_rows(item_row(sh, schd_row, 7), len(k)) # insert as many rows as length of dataframe k at the row item

        # for index, row in k.iterrows():                   # loop over each security
        #     print(supercat + ':' + schd_row, cat, index, len(k))
        #     rw = item_row(sh, schd_row, 7) + spacer2 + index    # loop sequentially over the empty rows, pasting securities along the way ...
        #     sh[rw][2].value  = row['Instr']
        #     sh[rw][4].value  = row['EMV']
        #     sh[rw][5].value  = row['PMV']
        #     sh[rw][4].border = Border(left  = Side(style = 'thin'))
        #     sh[rw][5].border = Border(right = Side(style = 'thin'))
        #     # sh[rw + len()][4].border = Border(top    = Side(style = 'thin'))
        #     # sh[rw + len()][5].border = Border(top    = Side(style = 'thin'))
        #     # sh[rw + len()][4].border = Border(bottom = Side(style = 'thin'))
        #     # sh[rw + len()][5].border = Border(bottom = Side(style = 'thin'))
        #     #sh.row_dimensions[rw].height = ht # https://stackoverflow.com/questions/37891149/openpyxl-auto-height-row
    # print(f'{timediff(start_time, time.time())}: populating the 3(f), 3(g), 3(h), and 3(i) line items completed')


# function to populate higher category sub-total line items
def paste_totals():
    start_time = time.time()
    # print('Summing and populating higher category sub-totals ...')
    for cat_1, values_1 in superhens.items():
        # print('\nLevel 1 Category:', cat_1)                                  # POINTER
        sigma_2_EMV = 0
        sigma_2_PMV = 0
        for cat_2, values_2 in values_1.items():
            # print('\tLevel 2 Category:', cat_2)                              # POINTER
            sigma_3_EMV = 0
            sigma_3_PMV = 0
            for cat_3, values_3 in values_2.items():  # alphabetic sub-category
                # print('\t\tAlphabetical Category:', cat_3)                   # POINTER
                sigma_4_EMV = 0
                sigma_4_PMV = 0
                for cat_4 in values_3:  # roman numeral sub-category
                    # print('\t\t\tRoman Category:', cat_4)                    # POINTER
                    sigma_4_EMV += r28[r28["R28C"] == cat_4][
                        "EMV"
                    ].sum()  # assign 'x.x(a)(i)' sub-totals
                    sigma_4_PMV += r28[r28["R28C"] == cat_4]["PMV"].sum()
                    # print('\t\t\t\t', cat_4 + ' :-> ' + str(sigma_4_PMV))    # POINTER
                sh[item_row(sh, cat_3, 7)][
                    4
                ].value = sigma_4_EMV  # assign 'x.x(a)' sub-totals
                sh[item_row(sh, cat_3, 7)][5].value = sigma_4_PMV
                sigma_3_EMV += (
                    r28[r28["R28C"] == cat_3]["EMV"].sum() + sigma_4_EMV
                )  # sum the alphabetical categories (cat_3 + sum of the romans)
                sigma_3_PMV += r28[r28["R28C"] == cat_3]["PMV"].sum() + sigma_4_PMV
                # print('\t\t\t', cat_3 + ' :-> ' + str(sigma_3_PMV))          # POINTER
            sh[item_row(sh, cat_2, 7)][4].value = sigma_3_EMV  # assign 'x.x' sub-totals
            sh[item_row(sh, cat_2, 7)][5].value = sigma_3_PMV
            sigma_2_EMV += (
                r28[r28["R28C"] == cat_2]["EMV"].sum() + sigma_3_EMV
            )  # sum the 'x.x' categories (cat_2 + sum of the alphabeticals)
            sigma_2_PMV += r28[r28["R28C"] == cat_2]["PMV"].sum() + sigma_3_PMV
            # print('\t\t\t', cat_2 + ' :-> ' + str(sigma_2_PMV))              # POINTER
        sh[item_row(sh, cat_1, 7)][4].value = sigma_2_EMV  # assign 'x' sub-totals
        sh[item_row(sh, cat_1, 7)][5].value = sigma_2_PMV

    # print(f'{timediff(start_time, time.time())}: summing and populating higher category sub-totals completed')


# function to delete categories without members
def delete_null_catgs():
    start_time = time.time()
    super_cats = [
        x for x in range(1, 11)
    ]  # complete list of super catgory (one-digit categories) items
    cats = set(
        [x[0] for x in list(ctgs)]
    )  # unique non-empty categories in the current fund
    empty_cats = [
        x for x in super_cats if str(x) not in cats
    ]  # list of empty super catgegories

    for empty_cat in empty_cats:
        # print(empty_cat, sh['G' + str(item_row(sh, str(empty_cat), 7))].value, item_row(sh, str(empty_cat + 1), 7) - item_row(sh, str(empty_cat), 7))
        sh.delete_rows(
            item_row(sh, str(empty_cat), 7),
            item_row(sh, str(empty_cat + 1), 7) - item_row(sh, str(empty_cat), 7),
        )

    # print(f'{timediff(start_time, time.time())}: deleting {len(empty_cats)} empty categories, {(", ").join(map(str, empty_cats))}, completed')


# paste instruments and then issuers onto the schedule
def paste_categories():
    start_time = time.time()
    # print('Pasting instruments and issuers for each of the categories ...')
    # update aggregate totals for the given item - https://openpyxl.readthedocs.io/en/stable/editing_worksheets.html

    for ctg in ctgs:
        # ctg     = ctgs[0]                    # set the next item / category to be populated on the sheet
        k = r28[r28["R28C"] == ctg]  # set of all securities for the item / category
        k.reset_index(
            drop=True, inplace=True
        )  # Saturn Cloud How to Reset Index in a Pandas Dataframe
        row_item = item_row(sh, ctg, 7)  # item row number

        # paste the aggregate security total fair value and total % fair value
        sh[row_item][4].value = k["EMV"].sum()
        sh[row_item][5].value = k["PMV"].sum()

        # static row numbers
        space1 = 1  # number of rows between category identifier in column 7 and first issuer in that category
        space2 = 1  # number of rows between issuers' securities in a category

        # paste code + description , limit, MV and %MV for each security in the category and uniformly format security row heights
        # delete the "No assets ..." and "Per issuer" rows then insert blanks rows on which to paste all the securities for the given category (= len(k))
        sh.delete_rows(
            row_item + space1, 2
        )  # sh.delete_rows(sheet_row_number, number_of_rows_to_be_deleted)
        sh.insert_rows(
            row_item + space1, len(k)
        )  # sh.insert_rows(first row location, number of rows to insert)

        for index, row in k.iterrows():  # loop over each security
            rw = (
                row_item + space1 + index
            )  # iterate sequentially over the empty rows, pasting securities along the way ...
            sh[rw][2].value = row["Instr"]
            sh[rw][4].value = row["EMV"]
            sh[rw][5].value = row["PMV"]
            sh[rw][4].border = Border(left=Side(style="thin"))
            sh[rw][5].border = Border(right=Side(style="thin"))
            sh[
                "C" + str(rw)
            ].style = "cell_style_K"  # apply wrap and then indent the security descriptions in the 'C' column

        # set of issuers and their sub-totals for the category - https://stackoverflow.com/questions/51971384/pandas-groupby-does-not-preserve-order
        m = (
            k.groupby(["Issuer", "IL"], as_index=False, sort=False)
            .agg({"CCY": "count", "EMV": "sum", "PMV": "sum"})
            .rename(columns={"CCY": "N"})
        )

        # paste code + description, issuer limit, MV and %MV for each issuer in the category
        rw = row_item + space1  # the active row for the following pastes
        for idx, row in m.iterrows():
            # insert number of blank rows on which to paste all the securities (=len(m)) for the given item
            sh.insert_rows(
                rw, 1
            )  # sh.insert_rows(first row location, number of rows to insert)
            sh[rw][2].value = row["Issuer"]
            sh[rw][3].value = row["IL"]
            sh[rw][4].value = row["EMV"]
            sh[rw][5].value = row["PMV"]
            sh[rw][4].border = Border(bottom=Side(style="thin"))
            sh[rw][5].border = Border(bottom=Side(style="thin"))
            sh[rw + row["N"]][4].border = Border(
                bottom=Side(style="thin"), left=Side(style="thin")
            )
            sh[rw + row["N"]][5].border = Border(
                bottom=Side(style="thin"), right=Side(style="thin")
            )
            sh.insert_rows(
                rw + row["N"] + 1, space2
            )  # sh.insert_rows(first row location, number of rows to insert)
            rw = (
                rw + row["N"] + 1 + space2
            )  # row['N'] contains the number of securities associated with the issuer

        # print(f'{timediff(start_time, time.time())}: pasting instruments and issuers for category {ctg} completed')
    # print(f'{timediff(start_time, time.time())}: pasting instruments and issuers for each of the {len(ctgs)} sub-categories completed')


# function to prettify the schedule
def prettify():
    start_time = time.time()
    # print(f'Prettifying the schedule ...')

    # format rows from top of schedule to the TOTAL row - # https://stackoverflow.com/questions/49525545/openpyxl-formatting-cell-with-decimal
    for row in range(item_row(sh, "Top", 7) - 1, item_row(sh, "SAFEX", 7)):
        cols = ["E", "F"]
        for col in cols:
            sh[col][row].number_format = nmbr
            sh[col][row].alignment = Alignment(horizontal="right", vertical="top")
            sh[col][row].font = Font(name=fnt, size=sz)

        # format the 'Limit (%)' column
        sh["D"][row].style = "cell_style_D"
        if sh["D"][row].value != 0.025:
            sh["D"][row].number_format = "0%"
        # if sh['D'][row].value == .025:
        else:  # format the '2.5%' limit indicators
            sh["D"][row].number_format = "0.0%"

    # format the 'Fair Value' headings
    fv_line_items = [
        "Item",
        "3(i) head",
        "3(f) head",
        "3(g) head",
        "3(h) head",
        "Exemptions",
    ]
    cols = ["E", "F"]
    for fv_line_item in fv_line_items:
        for col in cols:
            sh[col + str(item_row(sh, fv_line_item, 7))].style = cell_style_heads

    # format the 'Limit (%)' heading
    sh["D" + str(item_row(sh, "Limit (%)", 4))].style = cell_style_heads

    clls = ["F8", "F13"]
    for cll in clls:
        sh[cll].number_format = nmbr
        sh[cll].alignment = Alignment(horizontal="right", vertical="top")
        sh[cll].font = Font(name=fnt, size=sz)

    # format the 'Total Value' and 'TOTAL' value and percentage rows
    totals = ["Total Value", "TOTAL"]
    cols = ["E", "F"]
    for total in totals:
        for col in cols:
            sh[col + str(item_row(sh, total, 7))].font = Font(
                name=fnt, size=sz, bold=True
            )

    # set column widths
    widths = {"D": 8.57}
    for col in widths:
        sh.column_dimensions[col].width = widths[col]

    # print(f'{timediff(start_time, time.time())}: prettifying the schedule completed')


# delete the '3(h)' line items below TOTAL
def delete_3h():
    start_time = time.time()
    sh.delete_rows(
        item_row(sh, "3(h) head", 7),
        item_row(sh, "3(h)", 7) - item_row(sh, "3(h) head", 7) + 1,
    )
    # delete the '3(i) contra' line item, but only one row
    # print(f'{timediff(start_time, time.time())}: deleted 3(h) line item below TOTAL')


# delete supporting but extraneous columns
def delete_columns():
    start_time = time.time()
    sh.delete_cols(7, 24)
    # print(f'{timediff(start_time, time.time())}: columns "G" (7) through "T" (20) deleted')


# function to save the file
def save():
    start_time = time.time()
    # print(f'Saving {fund} Reg28 SchIB {date.strftime("%d%b%Y")}.xlsx in {pthTest} ...')
    wb.save(
        os.path.join(pthTest, f"{fund} Reg28 SchIB {date.strftime('%d%b%Y')}.xlsx")
    )  # save the completed Schedule IB in the Test folder
    wb.close


# function to append classifications sheet to existing workbook
# https://saturncloud.io/blog/how-to-append-existing-excel-sheet-with-new-dataframe-using-python-pandas/
def append_classifications_sheet():
    start_time = time.time()
    file_name = f"{fund} Reg28 SchIB {date.strftime('%d%b%Y')}.xlsx"
    file_save = os.path.join(pthTest, file_name)
    book = openpyxl.load_workbook(file_save)

    with pd.ExcelWriter(file_save, engine="openpyxl", mode="a") as writer:
        rgAll.to_excel(
            writer, sheet_name=f"{fund} Reg28 {date.strftime('%d%b%Y')}", index=False
        )

    # print(timediff(time.time(), start_time), 'to append the classifications sheet')


# BUG: ExcelWriter with mode='a' corrupts file #39576
# https://github.com/pandas-dev/pandas/issues/39576

# call and run all the functions
start_time0 = time.time()
get_inputs()

# loop
doneReg28 = []
noReg28 = []
for fund in tqdm(funds["Funds"]):
    if os.path.isfile(
        os.path.join(pthReports, f"{fund} Reg28 {date.strftime('%d%b%Y')}.xlsx")
    ):
        # print(fund)                     # test
        r28_df()  # create a dataframe, for what follows, from the Reg 28 categorisations sheet
        open_wb()  # load the SchIB teamplate as a workbook and assign cell styles to the workbook
        paste_nav()  # paste summary category subtotals on the schedule
        fghi()  # populate instruments for the 3(f), 3(g), 3(h), and 3(i) subtotals
        paste_totals()  # sum and populate the higher category sub-totals on the schedule
        delete_null_catgs()  # delete empty categories from the schedule
        paste_categories()  # paste instrument and issuer subtotals for each category on the schedule
        prettify()  # prettify the schedule
        delete_3h()  # 3(h) needs an instrument aggregation
        delete_columns()  # delete previously supporting, extraneous columns
        save()  # save the individual fund schedule
        append_classifications_sheet()  # append classification sheet
        doneReg28.append(fund)  # keep track of number of schedules completed
    else:
        pass
        noReg28.append(
            fund
        )  # keep track of funds which did not have a classifcation sheet to work from

    # print(f'Schedule IBs for {date.strftime("%d %b %Y")} done: {len(doneReg28)} - {(", ").join(doneReg28)}', '\n\n', \
    # f'Fund{"" if len(noReg28) == 1 else "s"} without a Reg 28 sheet: {len(noReg28)} - {(", ").join(noReg28)}')

# open the test folder where the reports are stored
# os.startfile(r'P:\Working Folders\Hilton\W\Reg_Tests')
# print('Opening the saved schedule')
# open_xl_file(os.path.join(pthTest,f'{fund} Reg28 SchIB {date.strftime("%d%b%Y")}.xlsx'))
os.startfile(pthReports)
os.startfile(pthTest)

print(
    f"Fund{'' if len(noReg28) == 1 else 's'} without a \
Reg 28 sheet for {date.strftime('%d %b %Y')} ({len(noReg28)}):\n",
    (", ").join(noReg28),
)
print(
    f"\n{timediff(start_time0, time.time())}: roundtrip \
time for {len(funds) - len(noReg28)} \
fund{'' if len(funds) - len(noReg28) == 1 else 's'}"
)

print("#########################")
print("#      END reg28ib      #")
print("#########################")
