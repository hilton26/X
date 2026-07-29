#!/usr/bin/env python
# coding: utf-8

# # Assigning issuer names to securities in a portfolio

# Source: Eagle 'Reg 28 Report - Incl Effective Exposure' with nine columns: 0. 'Entity Name', 1. 'Investment Type', 2. 'i Issue Name', 3. 'Primary Asset ID', 4. 'CCY', 5. 'Reg28 Classification', 6. 'End Market Value', 7. 'Percentage', 8. 'Closing Exposure PA'
#
#
# On "TypeError: This COM object can not automate the makepy process - please run makepy manually for this object", \
# close the Excel dialogue box and run this script again.')


print("\n\n#################################")
print("#                               #")
print("#      START issuers_1.py       #")
print("#                               #")
print("#################################\n\n")


# libraries, libraries!
print("Importing libraries for issuers_1 ...\n")
import time

start_time_issuers_1 = time.time()
start_time = time.time()

# general libraries
from datetime import datetime
import pandas as pd  # for dataframes

# pd.options.mode.chained_assignment = None  # default='warn'
# # https://stackoverflow.com/questions/20625582/how-to-deal-with-settingwithcopywarning-in-pandas
import numpy as np  # for np.NaN
import re  # for regex
from re import search  # for regex
from datetime import datetime  # for script run durations
import os  # for BESA folder contents
import sys
from tqdm import tqdm
import subprocess
from constants import (
    pthPy,
    pthBESA,
    pth_struct,
    pthCLNs,
    pthMedCirc2026,
    pthSttlmnt,
    pthTest,
    yll,
    iss_1,
    issuers_2,
    issuers_3,
)
from utilities import timediff

print(f" {timediff(start_time, time.time())} importing libraries for issuers_1\n")

# (1) read in the lookthrough holdings file, from the 'arc' sheet of py_reports.xlsm, as a dataframe

start_time = time.time()
print("Reading in classifier input file ...")

# check if a url was given in the py_reports file and then ...
df_check = pd.read_excel(pthPy, sheet_name="arc", usecols="V", nrows=7)
url = df_check.iloc[6, 0].replace('"', "")
rpt = df_check.iloc[2, 0]
if url == url:  # ... if so, use the py_reports url ...
    # py_input = pd.read_excel(url, engine = 'openpyxl', usecols = 'A:J')
    # py_input = pd.read_excel(url, sheet_name = 'All', usecols = 'A:J')
    py_input = pd.read_excel(url, usecols="A:J")
    if isinstance(list(py_input)[9], str):
        rptDate = datetime.strptime(list(py_input)[9], "%d %b %Y")
    else:
        rptDate = list(py_input)[9]
else:  # ... prompt for a valid url
    print('Please provide a valid URL in cell "L2" of the "classifier" tab')

print(f" {url if isinstance(url, str) else 'No url to look-through holdings'}\n")

fnds = py_input["Entity Name"].unique()
funds = (", ").join(fnds)
s = "" if len(fnds) == 1 else "s"
print(
    f"{rptDate.strftime('%a %d %b %Y')} instrument classifications for {len(fnds)} fund{s}:\n {funds}"
)

# get BESA data
besa_fnames = [
    int(s[re.search(r"\d{8}", s).span()[0] : re.search(r"\d{8}", s).span()[0] + 8])
    for s in os.listdir(pthBESA)
    if "." in s
]
besa_fdate = max(besa_fnames)  # BESA file date as an integer
bsaDate = datetime.strptime(str(besa_fdate), "%Y%m%d")  # BESA file date as a datetime

print(f" Report date    : {rptDate.strftime('%a %d %b %Y')}")
print(f" Report type    : {rpt}")
print(f" BESA file date : {bsaDate.strftime('%a %d %b %Y')}")
print(
    f" {len(py_input) - 1:,} securities, {len(py_input['Primary Asset ID'].unique()):,} \
(1/{(len(py_input) - 1) / len(py_input['Primary Asset ID'].unique()):.1f} times or \
{len(py_input['Primary Asset ID'].unique()) / (len(py_input) - 1) * 100:.1f}%) of which are unique"
)

print(f"\n {timediff(start_time, time.time())} reading in classifier input file\n")

# get input data
start_time = time.time()
print(
    "Reading in input data incl regex, CLNs, med schemes, settlement, BESA, accruals, realty, and margins ..."
)

res = list(filter(lambda x: str(besa_fdate) in x, os.listdir(pthBESA)))[0]
bsa = os.path.join(pthBESA, res)
besa_data = pd.read_csv(
    bsa, skiprows=4, usecols=["Bond Code"]
).dropna()  # BESA-listed security codes
lstds = pd.read_excel(
    pth_struct, sheet_name="listed", usecols=["Bond Code"]
).dropna()  # Listed debt override
dfrgx = pd.read_excel(
    pth_struct,
    sheet_name="dates",
    usecols=["date_regex", "format", "alternative format"],
).dropna(subset=["format"])  # regex formats
sa_hols = pd.read_excel(
    pth_struct, sheet_name="hols", usecols=["sa_hols"]
).dropna()  # import SA holidays
accr = pd.read_excel(pth_struct, sheet_name="accr", usecols=["accruals"]).dropna()
fnd_typ = pd.read_excel(pth_struct, sheet_name="fnd", usecols=["description"]).dropna()
marg = pd.read_excel(pth_struct, sheet_name="accr", usecols=["margins"]).dropna()
issrgx = pd.read_excel(
    pth_struct,
    sheet_name="issuers",
    usecols=[
        "description",
        "id",
        "issuer name",
    ],
).dropna(subset=["issuer name"])
indx1 = pd.read_excel(
    pth_struct, sheet_name="issuers", usecols=["description", "issuer name", "ticker"]
).dropna(subset=["issuer name"])
bx_re = pd.read_excel(
    pth_struct, sheet_name="issuers", usecols=["issuer name", "property"]
).dropna(subset=["issuer name"])
realty = bx_re[bx_re["property"] == "P"].drop("property", axis=1).reset_index(drop=True)
bond_data = pd.read_excel(
    pth_struct, sheet_name="guar", usecols=["Bond Code", "Guarantee Type"]
).dropna()  # govt guarantee
clns = pd.read_excel(
    pthCLNs,
    sheet_name="CLN",
    usecols=["Code", "CLN?", "Counterparty Long Name", "Issuer Long Name"],
).dropna(subset=["Code"])
med_circ = pd.read_excel(
    pthMedCirc2026, sheet_name="ListedDebtDec2025", usecols=["Bond Code", "2026Circ7"]
).dropna(subset=["Bond Code"])
sttlmnt = pd.read_excel(
    pthSttlmnt, sheet_name="Sttlmnt", usecols=["Fund", "Custodian", "SAFEX"]
).dropna(subset=["Fund"])
indx = indx1[indx1["ticker"] == "Index"].drop("ticker", axis=1).reset_index(drop=True)

accr_list = [
    x for e in accr.values.tolist() for x in e
]  # https://bobbyhadz.com/blog/python-remove-square-brackets-from-list
margin_list = [x for e in marg.values.tolist() for x in e]

print(
    f" {timediff(start_time, time.time())} reading \
in input data incl regex, CLNs, med schemes, \
settlement, BESA, accruals, realty, and margins\n"
)


# 'descr' + 'id' columns - https://stackoverflow.com/questions/19377969/combine-two-columns-of-text-in-pandas-dataframe
start_time = time.time()
print(
    'Joining "description" and "id" columns from \
the "issuers" tab of pth_struct.xlsm into a dataframe ...'
)

# convert "description" and "id" columns to string type and join them with a pipe delimiter
cols_to_convert = ["description", "id"]
for col in cols_to_convert:
    issrgx[col] = issrgx[col].fillna("").astype(str)

print(issrgx[["description", "id"]].info())

# issrgx["descid"] = issrgx[["description", "id"]].agg("|".join, axis=1)
issrgx["descid"] = issrgx[["description", "id"]].agg(
    lambda row: "|".join(v for v in row if v), axis=1
)
# issrgx.drop(['id', 'description'], axis = 1, inplace = True)

print(
    f' {timediff(start_time, time.time())} joining \
"description" and "id" columns from the "issuers" \
tab of pth_struct.xlsm into a dataframe\n'
)


# Determine funds with nil effective exposure
start_time = time.time()
print(
    "Determining funds with zero effective exposure \
and funds with no settlement account"
)

names = fnds
zero_EE = []
for name in names:
    if py_input[py_input["Entity Name"] == name]["Closing Exposure PA"].sum() == 0:
        zero_EE.append(name)

print(
    f" {len(zero_EE)} funds with empty effective exposure \
column: \n  {(', ').join(list(zero_EE))} \n"
)

# confirm all funds have a settlement bank account by merging df and sttlmnt
no_sttlmnt = py_input.merge(sttlmnt, left_on="Entity Name", right_on="Fund", how="left")
fnds_0 = no_sttlmnt[no_sttlmnt["Custodian"].isna()]["Entity Name"].unique()
s1 = "s" if len(fnds_0) != 1 else ""
print(
    f" {len(fnds_0)} fund{s1} with no settlement account: \n  {(', ').join(list(fnds_0))}"
)

# determine securities with integer IDs (SA government securities without "R" prefixed)
is_int_mask = py_input["Primary Asset ID"].apply(lambda x: isinstance(x, int))
print(f'{len(py_input[is_int_mask])} SA government securities without "R" prefixed')

py_input["Primary Asset ID"] = py_input["Primary Asset ID"].apply(
    lambda x: "R" + str(x) if isinstance(x, int) else x
)

print(
    f"{timediff(start_time, time.time())} determining funds with zero effective exposure and funds with no settlement account \n"
)


# In[7]:


# create text pattern functions

start_time = time.time()
print("Setting up functions ...")


def cln(txt):
    if txt in list(clns["Code"]):
        return clns[clns["Code"] == txt]["CLN?"].iloc[0]
    else:
        return None


# def fnd(txt):
#     for pattern in fnd_type['description']:
#         if re.search(pattern, str(txt).upper()):
#             return fnd_type.loc[fnd['description'] == pattern].iat[0,1] # note 'break' within the for loop
#             break


# function to identify FRNs and NCDs
def frn(txt):
    pattern = "STEP UP|STEP-UP|STEP_UP|STEPUP|NCD|FRN"
    if re.search(pattern, str(txt).upper()):
        return 1


# function to identify ILBs
def ilb(txt):
    pattern = "INFLATION|CPI|ILB"
    if re.search(pattern, str(txt).upper()):
        return 1


# function to identify funds
def fnd(txt):
    import re
    from re import search

    pattern1 = r"\b(?:FUND(?!\s*MANAGE)|FUND(?!S)|UCIT|ETF|ISHARES)\b"
    patternc = r"\b(PHYSICAL GOLD|GOLD ETC|COMMODITY|PHYSICAL SILVER|SILVER ETC|PLATINUM|PALLADIUM)\b"
    patternm = r"\b(MONEY)\b"
    patternd = r"\b(BONDS|$ TIP|$TIP|DURATION|YIELD|INCOME|INTEREST|POSITIVE RETURN)\b"
    patterne = r"\b(EQUITY|WORLD|FEEDER|\sPLUS\s|MSCI|SMALL CAP|LARGE CAP|GLOBAL|OPPORTUNITY|BIN YUAN|BALANCED|INTL)\b"
    patternp = r"\b(\sREAL\s|REALTY|REIT|PROPERTY|REAL ESTATE|HOMES|FAIRVEST|HYPROP)\b"
    if re.search(pattern1, str(txt).upper()) and re.search(patternc, str(txt).upper()):
        return "fc"
    elif re.search(pattern1, str(txt).upper()) and re.search(
        patternm, str(txt).upper()
    ):
        return "fm"
    elif re.search(pattern1, str(txt).upper()) and re.search(
        patternd, str(txt).upper()
    ):
        return "fd"
    elif re.search(pattern1, str(txt).upper()) and re.search(
        patterne, str(txt).upper()
    ):
        return "fe"
    elif re.search(pattern1, str(txt).upper()) and re.search(
        patternp, str(txt).upper()
    ):
        return "fp"
    else:
        return None


def property(txt):
    import re
    from re import search

    pattern = r"\b(\sREAL\s|REALTY|REIT|PROPERTY|REAL ESTATE|HOMES|FAIRVEST|HYPROP)\b"
    if re.search(pattern, str(txt).upper()):
        return "p"


# function to discern property tickers; ## Claude, 9 Mar 2026
def classify_re(row):
    isin_match = row["ISIN"] in isins["ISIN"].values
    return "P" if isin_match or name_match else np.nan


# function to identify pribvate equity funds
def pef(txt):
    import re
    from re import search

    pattern1 = r"\b(PRIVATE EQUITY)\b"
    pattern2 = r"\b(PRIVATE EQUITY FUND OF FUNDS|PRIVATE FOF)\b"
    if re.search(pattern1, str(txt).upper()):
        return "pef"
    elif re.search(pattern2, str(txt).upper()):
        return "pefof"
    else:
        return None


# function to identify pribvate equity funds
def hf(txt):
    import re
    from re import search

    pattern1 = r"HEDGE FUND(?! OF FUNDS)"
    pattern2 = r"\b(HEDGE FUND OF FUNDS|HEDGE FOF)\b"
    if re.search(pattern1, str(txt).upper()):
        return "hf"
    elif re.search(pattern2, str(txt).upper()):
        return "hfof"
    else:
        return None


def commodity(txt):
    import re
    from re import search

    pattern1 = r"\b(?:PHYSICAL GOLD|GOLD ETC)\b"
    pattern2 = r"\b(?:COMMODITY|PHYSICAL SILVER|SILVER ETC|PLATINUM|PALLADIUM)\b"
    if re.search(pattern1, str(txt).upper()):
        return "au"
    elif re.search(pattern2, str(txt).upper()):
        return "xu"
    else:
        return None


# function to find index
def inx(txt):
    if (indx["description"].eq(txt)).any():
        if (
            indx.loc[indxbond_data["Bond Code"] == txt].iat[0, 1]
            == "GOVERNMENT GUARANTEE"
        ):
            return 1


# function to derive margin account
def mrg(
    txtA, txtB
):  # https://stackoverflow.com/questions/13331698/how-to-apply-a-function-to-two-columns-of-pandas-dataframe
    pattern1 = r"\sMARGIN|(?!\d\d)MARG(?!\d\d)"  # pattern to test txtA in 'i Issue Name' column
    pattern2 = "VARMAR"  # pattern to test txtB in 'Primary Asset ID' column
    if re.search(pattern1, str(txtA).upper()) or re.search(
        pattern2, str(txtB).upper()
    ):  # https://www.regular-expressions.info/lookaround.html
        return 1


# # TEST THIS mrg(,) FUNCTION
# txt1 = 23 # should get AttributeError: 'int' object has no attribute 'upper'
# txt2 = '23'
# print(mrg(txt1, txt2))


# function to indicate if a security is government guaranteed
def gvg(
    txt,
):  # "vlookup" https://www.statology.org/pandas-check-if-column-contains-string/ exact string
    if (bond_data["Bond Code"].eq(txt)).any():
        if (
            bond_data.loc[bond_data["Bond Code"] == txt].iat[0, 1]
            == "GOVERNMENT GUARANTEE"
        ):
            return 1


# function that indicates if bond security is BESA-listed
def besa(
    txt,
):  # "vlookup" https://www.statology.org/pandas-check-if-column-contains-string/ exact string
    if (besa_data["Bond Code"].eq(txt)).any() or (lstds["Bond Code"].eq(txt)).any():
        return "B"


# function to identify strings of text starting with 3 or 4 capital letters and ending with two or three digits
# to identify candidate BESA-listed securities
def besa_maybe(txt):
    pattern = r"^[A-Z]{3,4}\d{2,3}$"
    if re.search(pattern, str(txt).upper()):
        return 1


# function to identify a repo
def repo(txt):
    pattern = "RPCO|RPMT|RPCA"
    if re.search(pattern, str(txt).upper()):
        return (
            txt[0:3].upper()
        )  # https://thispointer.com/python-how-to-get-first-n-characters-in-a-string/


# function to get days remaining to maturity # https://www.geeksforgeeks.org/python-datetime-strptime-function/
def term(date_string):
    try:
        if len(date_string) == 0:
            return None
        else:
            return datetime.strptime(date_string, "%d%b%Y") - rptDate

    # manage exceptions:
    except TypeError as te:
        return None
    except ValueError as ve:
        return None


# function to assign med scheme category based on current Circular 11 of 2024 from the CMS
def medcirc(txt):
    if (
        med_circ["Bond Code"].eq(txt).any()
    ):  # https://www.statology.org/pandas-check-if-column-contains-string/
        return med_circ.loc[med_circ["Bond Code"] == txt].iat[0, 1]


# utility function to check if a string contains an element in a given list
def res(
    t_list, t_string
):  # https://www.geeksforgeeks.org/python-test-if-string-contains-element-from-list/
    return bool([ele for ele in t_list if (ele in t_string)])


# function that extracts date from string
def datex(txt):
    try:
        for pattern in dfrgx["date_regex"]:
            if re.search(pattern, txt.title()):
                break
        return datetime.strptime(
            re.search(pattern, txt.title()).group(),
            dfrgx.loc[dfrgx["date_regex"] == pattern].iat[0, 1],
        ).strftime("%d%b%Y")

    # manage exceptions:
    except ValueError as ve:
        # print(f'ValueError {ve}')
        return None
    except TypeError as te:
        # print(f'TypeError {te}')
        return None
    except AttributeError as ae:
        # print(f'AttributeError {ae}')
        return None


# function to derive issuer name based on search for a pattern in instrument description and instrument id
def issuer_did(txt):
    for pattern in issrgx["descid"]:
        if re.search(pattern, str(txt).upper()):
            return (
                issrgx.loc[issrgx["descid"] == pattern]
                .reset_index(drop=True)
                .loc[0, "issuer name"]
            )
            break


# function to identify a derivative counterparty
def counterparty(row):
    if row["Investment Type"] == "SYTH":
        return np.nan
    elif (
        clns["Code"].eq(row["Primary Asset ID"])
    ).any():  # if cln is included in list then ...
        return clns.loc[clns["Code"] == row["Primary Asset ID"]].iat[
            0, 2
        ]  # ... look up bank counterparty
    elif (
        (row["Investment Type"] == "OP")
        and (
            (row["Derivative"] in "Structured Note")
            or (row["Derivative"] in "Linked Note")
        )
        and (row["CCY"] == "ZAR")
    ):  # SARB-registered bank-ssued option
        return issuer_did(row["i Issue Name"])
    elif ((row["Investment Type"] == "OP") or (row["Investment Type"] == "FT")) and (
        row["CCY"] == "ZAR"
    ):  # JSE-listed option
        return "JSE"
    elif ((row["Investment Type"] == "OP") or (row["Investment Type"] == "FT")) and (
        row["CCY"] != "ZAR"
    ):  # foreign exchange-listed option
        return "Exchange"
    elif (row["Investment Type"] == "EQ") or (
        row["Investment Type"] == "DERV"
    ):  # Equity issues and TRSes
        return row["Issuer"]
    else:
        return row["Issuer"]


# function to identify a derivative for Reg 28 CS1
def derivative(row):
    pattern1 = "STRUCTURED"
    pattern2 = "LINKED"
    if row["Investment Type"] == "SYTH":
        return np.nan
    elif re.search(
        pattern1, row["i Issue Name"].upper()
    ):  # TypeError: string indices must be integers, not 'str'
        return "Structured Note"
    elif re.search(pattern2, row["i Issue Name"].upper()):
        return "Linked Note"
    elif row["CLN"] == 1:
        return "Credit-linked Note"
    elif row["repo"] == "RPC":
        return "Repo Trade"
    elif row["Investment Type"] == "DERV":
        return "Swap"
    elif row["Investment Type"] == "FT":
        return "Futures"
    elif row["Investment Type"] == "OP":
        return "Listed Option"
    elif row["Investment Type"] == "FWD":
        return "Currency Forward"


# function to derive index name name based on search for a pattern in instrument description
def dexin(txt):
    for pattern in indx["description"]:
        if re.search(pattern, str(txt).upper()):
            return indx.loc[indx["description"] == pattern].iat[
                0, 1
            ]  # note 'break' within the for loop
            break


# # To test repo(), cln(), and property() functions
# txt1 = 'NN109U sadf structuREd asdfs rpmtb attacq'
# print(f' {txt1} derivative()  : {derivative(txt1)}')
# print(f' repo({txt1})  = {repo(txt1)}')
# print(f' cln({txt1})   = {cln(txt1)}')
# print(property(txt1))

print(f" {timediff(start_time, time.time())} setting up functions", "\n")


# In[8]:


# function to assign an issuer: https://towardsdatascience.com/create-new-column-based-on-other-columns-pandas-5586d87de73d
start_time = time.time()
print("Setting up issuer identifier function ...")


def classify1(row):
    t = issuer_did(row["i Issue Name"])  # temp, so function only gets called once
    g = issuer_did(row["Primary Asset ID"])  # temp, so function only gets called once

    if cln(row["Primary Asset ID"]) == 1:  # if cln is included in list then ...
        return clns.loc[clns["Code"] == row["Primary Asset ID"]].iat[
            0, 1
        ]  # ... look up reference entity

    elif repo(row["Primary Asset ID"]) == "RPC":  # repo, bank legs RPCO and RPCA
        return "Absa Bank Ltd"

    elif repo(row["Primary Asset ID"]) == "RPMT":  # repo, government bond leg
        return "Republic of South Africa"

    elif (row["Investment Type"] == "OP") and (
        (row["Derivative"] in "Structured Note") or (row["Derivative"] in "Linked Note")
    ):  # repo, gov leg
        return dexin(row["i Issue Name"])

    elif isinstance(g, str):  # INSTRUMENT ID is a hit ...
        return g

    elif isinstance(t, str):  # INSTRUMENT DESCRIPTION is a hit ...
        return t

    else:  # signifies no issuer assigned
        return "-xxx-"


print(
    f" {timediff(start_time, time.time())} setting up issuer identifier function", "\n"
)

# # usage of classify1() https://pandas.pydata.org/docs/reference/api/pandas.DataFrame.html
# dict   = {'i Issue Name': 'Absa Bank Ltd JB3+150 ABFN29 290825', 'Primary Asset ID': 'ABFN29', 'Investment Type': 'FI' , 'Derivative': np.nan}
# idx    = [0]
# df     = pd.DataFrame(data = dict, index = idx)
# df
# print(classify1(df.loc[0]))


# In[9]:


# (2) remove blank Market Value rows and zero-value Effective Exposure rows
start_time = time.time()
print("Removing NaN and zero value rows ...")

before = py_input.groupby("Entity Name")[
    ["End Market Value", "Percentage of Market Value", "Closing Exposure PA"]
].sum()
before["Diff"] = before["End Market Value"] - before["Closing Exposure PA"]
# https://stackoverflow.com/questions/43102734/format-a-number-with-commas-to-separate-thousands

rowsNaN = len(py_input[py_input["End Market Value"].isnull()])
py_input = py_input[py_input["End Market Value"].notnull()]  # remove NaN MV column rows
df_input = py_input[
    (round(py_input["End Market Value"], 2) != 0)
    | (round(py_input["Closing Exposure PA"], 2) != 0)
]

# drop the last column
df_input = df_input.drop(df_input.columns[9], axis=1)

# delete rows where MV and EE are zero to two decimals
print(
    f" {rowsNaN:,} NaN rows and {len(py_input) - len(df_input):,} zero effective exposure rows \
removed, {len(df_input):,} rows remain"
)

after = df_input.groupby("Entity Name")[
    ["End Market Value", "Percentage of Market Value", "Closing Exposure PA"]
].sum()
after["Diff"] = after["End Market Value"] - after["Closing Exposure PA"]
# https://stackoverflow.com/questions/43102734/format-a-number-with-commas-to-separate-thousands

print(f"{timediff(start_time, time.time())} removing NaN and zero value rows", "\n")


# In[10]:


# (3) change the 'Percentage of Market Value' column
start_time = time.time()
print('Changing "Percentage of Market Value" column ...')

navs = df_input.groupby("Entity Name")[
    "End Market Value"
].sum()  # (column N) this has type 'pandas.core.series.Series'
nav = navs.to_dict()  # nav series changed to dictionary to make it lookupable

# recalc the '% of Total Market Value' column per fund and ...
newTMV = []
for i, row in tqdm(
    df_input.iterrows(), total=df_input.shape[0]
):  # https://stackoverflow.com/questions/47087741/use-tqdm-progress-bar-with-pandas
    if nav[row["Entity Name"]] == 0:
        fndpct = 100
    else:
        fndpct = row["End Market Value"] / nav[row["Entity Name"]] * 100
    newTMV.append(fndpct)

# ... replace 'Percentage of Market Value' with the values in the new list
df_input["Percentage of Market Value"] = newTMV

print(
    f'{timediff(start_time, time.time())} changing "Percentage of Market Value" column',
    "\n",
)


# In[11]:


# (4) save 'df_input' dataframe including ALL funds as a workbook to be used later
start_time = time.time()
print('Saving "df_input" dataframe as a workbook called "yall" ...')

with pd.ExcelWriter(yll, engine="xlsxwriter") as writer:
    df_input.to_excel(writer, index=False, sheet_name="all")  # assigned attributes
writer.close()

print(
    f' {timediff(start_time, time.time())} saving "df_input" dataframe as a workbook called "yall"\n'
)


# In[12]:


uniques = df_input.drop_duplicates(subset="Primary Asset ID", keep="first")
print(list(uniques), uniques.shape)


# In[13]:


# (5) find unique instruments and identify their instrument attributes
start_time = time.time()
print("Isolating unique securities ...")

# unique securities - https://pandas.pydata.org/docs/reference/api/pandas.DataFrame.drop_duplicates.html
uniques = df_input.drop_duplicates(subset="Primary Asset ID", keep="first")

# drop fund name column - https://pandas.pydata.org/docs/reference/api/pandas.DataFrame.drop.html
uniques.drop(["Entity Name"], axis=1, inplace=True, errors="ignore")

# drop accrual and margin Investment Type rows
uniques = uniques[
    ~uniques["Primary Asset ID"].isin(accr_list)
    & ~uniques["Primary Asset ID"].isin(margin_list)
]

# reset index - https://pandas.pydata.org/docs/reference/api/pandas.DataFrame.reset_index.html
uniques.reset_index(
    drop=True, inplace=True
)  # drop original index, overwrite the original dataframe

print(f" {len(uniques)} unique securities")

print(f"   {timediff(start_time, time.time())} isolating unique securities\n")


# In[14]:


# (6) identify the unique instruments' attributes
start_time = time.time()
print("Appending instrument attributes to the unique securities ...")

# add new columns to uniques dataframe
uniques[["FundX", "PEFX", "HFX", "CommodityX"]] = None

# uniques['fnd']        = uniques['i Issue Name'    ].map(fnd)         # NEW
# uniques['CLN']        = uniques['i Issue Name'    ].map(cln)
uniques["CLN"] = uniques["Primary Asset ID"].map(cln)
uniques["FRN"] = uniques["i Issue Name"].map(frn)
uniques["ILB"] = uniques["i Issue Name"].map(ilb)
uniques["Date"] = uniques["i Issue Name"].map(datex)  # security maturity date
uniques["FundX"] = uniques["i Issue Name"].map(fnd)
uniques["PropertyX"] = uniques["i Issue Name"].map(property)
uniques["PEFX"] = uniques["i Issue Name"].map(pef)
uniques["HFX"] = uniques["i Issue Name"].map(hf)
uniques["CommodityX"] = uniques["i Issue Name"].map(commodity)  # commodity
uniques["MedCirc"] = uniques["Primary Asset ID"].map(medcirc)
uniques["GovGuar"] = uniques["Primary Asset ID"].map(gvg)
uniques["repo"] = uniques["Primary Asset ID"].map(repo)
uniques["BESA"] = uniques["Primary Asset ID"].map(besa)
uniques["BESA_MAYBE"] = uniques["Primary Asset ID"].map(besa_maybe)
uniques["margin"] = uniques.apply(
    lambda x: mrg(x["i Issue Name"], x["Primary Asset ID"]), axis=1
)
uniques["Derivative"] = uniques.apply(derivative, axis=1)
uniques["Term"] = (pd.to_datetime(uniques["Date"], format="%d%b%Y") - rptDate).dt.days
# https://stackoverflow.com/questions/26763344/convert-pandas-column-to-datetime
# https://stackoverflow.com/questions/37840812/pandas-subtracting-two-date-columns-and-the-result-being-an-integer

print(
    f" {timediff(start_time, time.time())} appending instrument attributes to the unique securities\n"
)


# In[16]:


# (7) identify issuers
start_time = time.time()
print(
    f"Identifying issuers over {len(uniques):,} unique securities \
for {len(fnds)} fund{s} at {rptDate.strftime('%a %d %b %Y')} using \
{len(issrgx['issuer name'].dropna()):,} regex patterns ..."
)

issuers = []
for index, row in tqdm(uniques.iterrows(), total=uniques.shape[0]):
    issuers.append(classify1(row))

# add new 'Issuer' column
uniques["Issuer"] = issuers

print(f" {timediff(start_time, time.time())} identifying issuers\n")


# In[24]:


# (8) identify derivative counterparties for Reg 28 CS1 of 2023
start_time = time.time()
print("Identifying derivative counterparties ...")

uniques["Counterparty"] = uniques.apply(counterparty, axis=1)

print(f" {timediff(start_time, time.time())} identifying derivative counterparties\n")


# In[25]:


# (9) identify securities with absent issuers
start_time = time.time()
print("Identifying securities with absent issuers ...")

no_issuer = uniques[
    (uniques["Issuer"] == "-xxx-") | uniques["Issuer"].isnull()
].drop_duplicates(subset="Primary Asset ID", keep="first")

no_CLN_issuer = uniques[
    (uniques["CLN"] == 1) & (uniques["Issuer"].isna())
].drop_duplicates(subset="Primary Asset ID", keep="first")

print(f" Unallocated issuers : {str(len(no_issuer.Issuer))}")
print(f" Unallocated CLNs    : {str(len(no_CLN_issuer.CLN))}")

print(
    f" {timediff(start_time, time.time())} identifying securities with absent issuers\n"
)


# In[26]:


# (10) write the dataframe to review it as a workbook
start_time = time.time()
print("Writing the dataframe to a sheet for review ...")

iss1_xl = pd.ExcelWriter(iss_1, engine="xlsxwriter")  #!pip install xlsxwriter
uniques.to_excel(iss1_xl, index=False, sheet_name="uniques")  # no attributes
no_issuer.to_excel(
    iss1_xl, index=False, sheet_name=f"no issuers ({len(no_issuer)})"
)  # no issuers
no_CLN_issuer.to_excel(
    iss1_xl, index=False, sheet_name=f"no CLN issuers ({len(no_CLN_issuer)})"
)  # no CLN issuers

iss1_xl.close()

print(
    f" {timediff(start_time, time.time())} writing the dataframe to a sheet for review\n"
)


# In[27]:


# (11) prettify the sheets using openpyxl
start_time = time.time()
print("Giving the review sheet structure with openpyxl ...")

# utilise openpyxl tools to add excel features to results sheet
import openpyxl as px  # for adding sort filters to the excel sheet
from openpyxl.cell import Cell  # to format cells
from openpyxl.styles import (
    Alignment,
    Color,
    PatternFill,
    Font,
    Border,
)  # to format cells

# add filters to all columns https://stackoverflow.com/questions/51566349/openpyxl-how-to-add-filters-to-all-columns, and then
# freeze entire header row in openpyxl https://stackoverflow.com/questions/25588918/how-to-freeze-entire-header-row-in-openpyxl
# iterate over worksheets https://stackoverflow.com/questions/18495672/how-to-iterate-over-worksheets-in-workbook-openpyxl
# wrap text in sheet headers https://stackoverflow.com/questions/42215933/apply-wrap-text-to-all-cells-using-openpyxl
# fill colour cells https://stackoverflow.com/questions/30484220/fill-cells-with-colors-using-openpyxl
wb = px.load_workbook(pthTest + r"\issuers_1.xlsx")
for sheet in wb.worksheets:
    sheet.auto_filter.ref = sheet.dimensions
    sheet.freeze_panes = sheet["J2"]
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.fill = PatternFill(
                start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
            )

# add a hyperlink to the 'pth_struct' sheet
# https://stackoverflow.com/questions/46162147/python-openpyxl-change-font-to-bold
# https://toricode.com/python-openpyxl-font-underline-for-excel-cells/
# https://stackoverflow.com/questions/8440284/setting-styles-in-openpyxl
# https://stackoverflow.com/questions/22986725/how-can-i-align-text-in-a-cell-to-the-top-with-openpyxl
sht = wb[f"no issuers ({len(no_issuer)})"]
sht["P1"].hyperlink = pth_struct
sht["P1"].value = "pth_struct.xlsm"
sht["P1"].font = Font(bold=True, underline="single", color="0000EE")
sht["P1"].alignment = Alignment(horizontal="left", vertical="top")

wb.save(pthTest + r"\issuers_1.xlsx")  # save the file to the W folder
wb.close()

print(
    f" {timediff(start_time, time.time())} giving the review sheet structure with openpyxl",
    "\n",
)
print(
    f"\n {timediff(start_time_issuers_1, time.time())} ISSUERS_1 COMPLETED\n===============================\n"
)


print("\n\n###############################")
print("#                             #")
print("#      END issuers_1.py       #")
print("#                             #")
print("###############################\n\n")


# (12) run issuers_2, and _3.ipynb if all securities have an assigned issuer, else open issuers_1.xlsx
start_time = time.time()
print(
    "Running issuers_2, and _3.ipynb if all securities have an assigned issuer, else opening issuers_1.xlsx, ..."
)

print(
    f"Exceptions: \n {len(no_issuer)} unnamed issuers, and,\n {len(no_CLN_issuer)} unnamed CLN issuers"
)

if (len(no_issuer.Issuer) == 0) and (len(no_CLN_issuer.CLN) == 0):
    subprocess.run([sys.executable, issuers_2])  ## when running issuers_2.py
    subprocess.run([sys.executable, issuers_3])  ## when running issuers_3.py
else:
    os.startfile(iss_1)
    print(r"Check the issuers and CLNs in the \issuers_1.xlsx file")

print(
    f"\n{timediff(start_time_issuers_1, time.time())} running issuers_1, _2, and _3.ipynb\n"
)

# os.startfile(iss_1)
# os.startfile(iss_2)

print("\n\n##########################################")
print("#                                        #")
print("#      END issuers_1, _2, and _3.py      #")
print("#                                        #")
print("##########################################\n\n")
