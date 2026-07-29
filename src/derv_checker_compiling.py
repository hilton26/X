#!/usr/bin/env python
# coding: utf-8

# # Compile the Derivative Cover Check Sheets

print("\n\n##################################################")
print("#                                                #")
print("#      START 2/4 derv_checker_compiling.py  X    #")
print("#                                                #")
print("##################################################\n\n")

# Libraries, libraries!

import time

start_time = time.time()
start_time_derv_compiling = start_time
print("Importing libraries and setting paths ...")

import pandas as pd
import os, sys, shutil, re
from datetime import timedelta
from tqdm import tqdm
from constants import pthSttlmnt, pthEXPORTS, pthLOCAL, derv_tmpl
from utilities import timediff, parn_de

print(
    f" {timediff(start_time, time.time())} importing libraries \
and setting paths\n"
)

# Get report date and selected summary sheet option
start_time = time.time()
print(
    "Getting the reporting date and latest downloaded \
holdings and derivatives files...\n"
)

fPARN, fDE, funds, rptDate, summ_yn, dervthreshold = parn_de()

# check if the required files have been downloaded, else continue
if not os.path.exists(fPARN) or not os.path.exists(fDE):
    sys.exit(
        f"Stopping: missing expected download(s):\n"
        f"  {fPARN} which {'exists' if os.path.exists(fPARN) else 'does not exist'}\n"
        f"  {fDE} which {'exists' if os.path.exists(fDE) else 'does not exist'}\n"
    )
# df = pd.read_excel(pthPy, sheet_name="arc", header=None, usecols="A,E").dropna(
#     subset=[0]
# )
# k = df.iloc[2, 1]
# rptDate = (
#     k if isinstance(k, datetime) else prior_working_day(datetime.today())
# )  # prior working day or report date override; has type datetime()
# summ_yn = df.iloc[3, 1]
# funds = df[0].iloc[1:]

# # get paths to the expected holdings and derivative metric files
# fPARN = os.path.join(
#     pth_dl,
#     f"PARN ({len(funds)}) {rptDate.strftime('%d%b%Y')}.csv",
# )
# fDE = os.path.join(
#     pth_dl,
#     f"DERV ({len(funds)}) {rptDate.strftime('%d%b%Y')}.csv",
# )

# # check if the required files have been downloaded, else exit
# if not os.path.exists(fPARN) or not os.path.exists(fDE):
#     sys.exit(
#         f"\n\nStopping: missing expected download(s):\n"
#         f"  {fPARN} which {'exists' if os.path.exists(fPARN) else 'does not exist'}\n"
#         f"  {fDE} which {'exists' if os.path.exists(fDE) else 'does not exist'}\n\n"
#     )

print(
    f" {rptDate.strftime('%A %d %b %Y')} for {len(funds)} funds:\n",
    f"{(', ').join(funds.tolist())}",
)
print(f" A summary sheet is{' not' if summ_yn == 'No' else ''} required", "\n")
print(
    f" {timediff(start_time, time.time())} getting the reporting \
date and latest downloaded holdings and derivatives files\n",
)

# Dataframe the holdings and deltas csv files and convert numerical columns from str to float

start_time = time.time()
print("Creating holdings and deltas dataframes ...")

# dataframe of the holdings and deltas csv files
# wbH = pd.read_csv(parN_nm)
wbH = pd.read_csv(fPARN)
wbD = pd.read_csv(fDE)

# get the fund names and fund codes from the holdings dataframe
fnames = wbH["Entity Name"].unique()
fcodes = wbH["Entity ID"].unique()

# convert derivative dataframe columns from str to float
# https://stackoverflow.com/questions/55557004/getting-attributeerror-float-object-has-no-attribute-replace-error-while
headsD = ["Nominal Holding", "Delta", "Market Value", "Effective Exposure"]
for head in headsD:
    wbD[head] = [
        str(x).replace(",", "").replace("-", "-") for x in wbD[head]
    ]  # KeyError: 'Nominal Holding'
    wbD[head] = wbD[head].astype(float)

# convert holdings dataframe columns from str to float
headsH = [
    "Current Exposure",
    "Sum of Market Value Income",
    "Original Nominal",
    r"Market Price /Yield",
]
for head in headsH:
    wbH[head] = [str(x).replace(",", "").replace("-", "-") for x in wbH[head]]
    wbH[head] = wbH[head].astype(float)

# holdings date column from type string to type datetime
wbH["i Position Effective Date"] = pd.to_datetime(wbH["i Position Effective Date"])

# maturity date column from type object to type datetime and 'NaT' to a long date in datetime format
wbH["Maturity Date"] = pd.to_datetime(wbH["Maturity Date"])


# https://stackoverflow.com/questions/38509538/numpy-checking-if-a-value-is-nat
def convert_NaT_to_report_date(dte):
    if str(dte) == "NaT":
        # return datetime(2099, 12, 31)
        return rptDate
    else:
        return dte


# rename 'UNKNOWNs' as 'SWAPS' where they are not SYTH or empty portfolio holdings
# https://stackoverflow.com/questions/36909977/update-row-values-where-certain-condition-is-met-in-pandas
unknowns_filter = (
    (wbH["Valuation First Level"] == "UNKNOWN")
    & (wbH["Sub Security Type"] == "TRS")
    & (wbH["Investment Type"] != "SYTH")
    & pd.notna(wbH["Investment Type"])
)
wbH.loc[unknowns_filter, ["Valuation First Level", "Valuation Second Level"]] = "SWAPS"
unknowns = wbH.loc[unknowns_filter]
print(
    f'  {len(unknowns)} "UNKNOWN" securities: {(", ").join(unknowns["PrimaryAssetID"].tolist())}'
)

# convert date columns to datetime format
date_cols = ["i Position Effective Date", "Maturity Date", "Next Coupon Date"]
for date_col in date_cols:
    wbH[date_col] = pd.to_datetime(wbH[date_col])

# convert holdings numerical columns to numbers
num_col_names = [
    "Original Nominal",
    "Clean Book Value",
    "Clean Market Value",
    "Accrued Income",
    "Dividend Receivable",
    "Sum of Market Value Income",
    "Market Price /Yield",
    r"% of Total Market Value",
    "Coupon",
    "Duration",
    "Modified Duration",
    "NACA Yield",
    "NACM Yield",
    "Weighted Avg NACA Yield",
    "Weighted Avg NACM Yield",
    "Market Value %",
    "Current Exposure",
    "Current Exposure %",
    "Weighted Average NACS Yield",
    "Weighted Average Coupon",
    "Weighted Modified Duration",
]
for num_col_name in num_col_names:
    wbH[num_col_name] = wbH[num_col_name].astype(str).str.replace(",", "").astype(float)

# convert deltas numerical columns to numbers
num_col_names_dervs = ["Nominal Holding", "Delta", "Market Value", "Effective Exposure"]
for num_col_name_derv in num_col_names_dervs:
    wbD[num_col_name_derv] = (
        wbD[num_col_name_derv].astype(str).str.replace(",", "").astype(float)
    )

print(
    f" {timediff(start_time, time.time())} creating \
holdings and deltas dataframes",
    "\n",
)

# Change the '% of Total Market Value" column
start_time = time.time()

# change the '% of Total Market Value" column (N) to the fund-
# specific % based on 'Sum of Market Value Income' column (M)
navs = wbH.groupby("Entity ID")[
    "Sum of Market Value Income"
].sum()  # (column N) this has type 'pandas.core.series.Series'
nav = navs.to_dict()  # nav series changed to dictionary to make it lookupable

# recalc the '% of Total Market Value' column per fund and then ...
newTMV = []
for i, row in tqdm(
    wbH.iterrows(),
    desc="Recalculating Total Market Value percentages ...",
    total=wbH.shape[0],
):
    if nav[row["Entity ID"]] == 0:
        fndpct = 100
    else:
        fndpct = row["Sum of Market Value Income"] / nav[row["Entity ID"]] * 100
    newTMV.append(fndpct)

# ... replace the '% of Total Market Value' with the values in the new list
wbH["% of Total Market Value"] = newTMV

# recalc the 'Current Exposure %' column per fund and then ...
newCEp = []  # new Current Exposure % column
for i, row in tqdm(wbH.iterrows(), total=wbH.shape[0]):
    if nav[row["Entity ID"]] == 0:
        currentexposurepct = 1
    else:
        currentexposurepct = row["Current Exposure"] / nav[row["Entity ID"]] * 100
    newCEp.append(currentexposurepct)

# ... replace the '% of Total Market Value' with the values in the new list
wbH["Current Exposure %"] = newCEp

print(
    f" {timediff(start_time, time.time())} recalculating \
and saving fund Total Market Value percentages\n"
)

# Determine which derivative cover reports have already completed

start_time = time.time()
print(
    f"Determining the derivative cover reports already \
completed for {rptDate.strftime('%a %d %b %Y')} ..."
)

# determine which funds have completed derivative cover reports
pattern = (
    r"^(?!TEST).*Derv Calc " + f"{rptDate.strftime('%d%b%Y')}" + r"\.xlsx$"
)  # file name does not start with "TEST "
fcodes_compl = [
    item.replace(f" Derv Calc {rptDate.strftime('%d%b%Y')}.xlsx", "")
    for item in os.listdir(pthLOCAL)
    if re.search(pattern, item)
]  # codes of completed derivative calculation sheets in Path.home()\Documents\TestDervFiles

fnames_compl = [
    wbH[wbH["Entity ID"] == fundcode.replace("~$", "")].iloc[0, 0]
    for fundcode in fcodes_compl
]
# names of completed derivative calculation sheets
# replace() removes the leading '~$' which happens when a file with name is in use in Excel

# get incomplete funds as the difference between the funds and completed list of fund names
# https://www.askpython.com/python/list/difference-between-two-lists-unique-entries#:~:
# \text=In%20Python%2C%20to%20find%20the,unique%20entries%20from%20both%20lists.
# incompl = (set(fnames) - set(fnames_compl)).union(set(fnames_compl) - set(fnames))
fcodes_incompl = (set(fcodes) - set(fcodes_compl)).union(
    set(fcodes_compl) - set(fcodes)
)
fnames_incompl = (set(fnames) - set(fnames_compl)).union(
    set(fnames_compl) - set(fnames)
)
print(
    "\n",
    f"{len(fcodes_compl)} completed:",
    "\n",
    f" {(',').join(sorted(fcodes_compl))}",
)
print(
    "\n",
    f"{len(fcodes_incompl)} remaining:",
    "\n",
    f" {(',').join(sorted(fcodes_incompl))}",
    "\n",
)
print(
    f" {timediff(start_time, time.time())} determining the derivative cover reports already completed for \
{rptDate.strftime('%a %d %b %Y')}: {len(fcodes_compl)} completed, {len(fcodes_incompl)} remaining",
    "\n",
)

start_time_compiling = time.time()

### TEST ###
# choose a specifc fund
# fcode = "PIMBAL"
# fname = wbH[wbH["Entity ID"] == fcode]["Entity Name"].iloc[0]
### TEST ###

import openpyxl

# create "td" to find maturities > 13 months
td = rptDate + timedelta(days=397)

# create a lookup table for fund UT status and investment team
twoA = pd.read_excel(pthSttlmnt, sheet_name="Funds", usecols="A, D:E")

#
fnames_incompl = sorted(fnames_incompl)
for index, fname in enumerate(
    tqdm(
        fnames_incompl,
        desc=f"Compiling derivative calculation files for {len(fnames_incompl)} funds at {rptDate.strftime('%d %b %Y')} ...",
    ),
    start=1,
):
    start_time = time.time()

    fcode = wbH[wbH["Entity Name"] == fname].iloc[0, 36]
    delt = wbD[wbD["Entity Name"] == fname]
    hold = wbH[wbH["Entity Name"] == fname]
    wb = openpyxl.load_workbook(derv_tmpl)  # open the template
    sh = wb["Summary"]  # assign the sheet to be worked on
    # sh.title = f'{fund} SchIB {date.strftime("%d%b%Y")}'  # set tab name of IB sheet

    # create a lookup table for fund UT status and investment team
    ftyp = twoA[twoA["Fund Code"] == fcode].iloc[0, 2]

    # enter fund name and report date on template
    sh["A1"] = f"{fname} ({fcode})"  # fund long name
    sh["A2"] = f"Derivative Cover {rptDate.strftime('%A %#d %B %Y')}"  # report date
    sh["A3"] = ftyp  # UT or not a UT

    dervs = (
        (hold["Valuation First Level"] == "DERIVATIVES").sum()
        + (hold["Valuation First Level"] == "FORWARDS").sum()
        + (hold["Valuation First Level"] == "FORWARD RATE AGREEMENT").sum()
        + (hold["Security Type"] == "SWP").sum() / 3
        + (hold["Security Type"] == "SWAP").sum() / 3
    )  # number of derivatives, catering for the three legs of a FRA

    nav = hold["Sum of Market Value Income"].sum()  # fund NAV

    cash = hold[
        (hold["Valuation First Level"] == "CASH")
        & (hold["Security Type"] != "CIS")
        & (hold["Sub Security Type"] != "REPO")
        & (~hold["i Issue Name"].str.upper().str.contains("MARGIN"))
    ][
        "Current Exposure"
    ].sum()  # cash incl accruals, excl MMFs, excl margins, excl repo P/L

    mmfs = hold[
        (hold["PrimaryAssetID"] == "PRMFB3") | (hold["PrimaryAssetID"] == "PCMMB3")
    ]["Current Exposure"].sum()  # + \
    # hold[hold['PrimaryAssetID'] == 'PCMMB3']['Current Exposure'].sum() # money market funds, i.e., PMMF and PCMMF

    mmis = (
        hold[
            (hold["Valuation First Level"] == "MONEY MARKET")
            & (hold["Maturity Date"] < td)
            # #TypeError: Cannot compare Timestamp with datetime. Use ts == pd.Timestamp(date) or ts.date() == date instead.
            & (hold["Sub Security Type"] != "CLN")
        ]["Current Exposure"].sum()
        if ftyp == "UT" or ftyp == "ETF"
        else hold[hold["Valuation First Level"] == "MONEY MARKET"][
            "Current Exposure"
        ].sum()
    )  # money market instruments

    bonds = (
        hold[
            (hold["Valuation First Level"] == "BONDS")
            & (hold["Maturity Date"] < td)
            & (hold["Sub Security Type"] != "CLN")
        ]["Current Exposure"].sum()
        if ftyp == "UT" or ftyp == "ETF"
        else hold[hold["Valuation First Level"] == "BONDS"]["Current Exposure"].sum()
    )  # bonds excl CLNs, excl > 13 months, else incl CLNs

    marg_jse = (
        hold[hold["PrimaryAssetID"] == "SAFEX"]["Current Exposure"].sum()
        + hold[hold["PrimaryAssetID"] == "VARMARG"]["Current Exposure"].sum()
    )  # JSE SAFEX and JSE VARMARG

    marg_otc = hold[
        (hold["i Issue Name"].str.upper().str.contains("MARGIN"))
        & (hold["PrimaryAssetID"] != "SAFEX")
        & (hold["PrimaryAssetID"] != "VARMARG")
    ]["Current Exposure"].sum()  # other margins, non-JSE margin accounts

    repo = (
        hold[hold["PrimaryAssetID"].str.upper().str.startswith("RPCO", na=False)][
            "Current Exposure"
        ].sum()
        + hold[hold["PrimaryAssetID"].str.upper().str.startswith("RPCA", na=False)][
            "Current Exposure"
        ].sum()
    )

    repo_gain = max(0, repo)  # net profit on repos

    repo_loss = min(0, repo)  # net loss on repos

    crry_derv = hold[
        hold["Valuation Second Level"].str.upper().str.contains("CURRENCY DERIVATIVES")
    ]["Current Exposure"].sum()

    fwds = min(
        0,
        hold[hold["Valuation Second Level"].str.upper().str.contains("FORWARDS")][
            "Current Exposure"
        ].sum(),
    )

    other_UTs = (
        0
        if twoA[twoA["Fund Code"] == fcode]["UT"].iloc[0] == "UT"
        else hold[
            (hold["Security Type"] == "CIS")
            & (hold["Sub Security Type"] != "CSH")
            & (~hold["PrimaryAssetID"].isin(["PIMEVOA", "PIMIDFA"]))
        ]["Current Exposure"].sum()
    )  # other, non-MMF, UTs

    count_other_UTs = (
        0
        if twoA[twoA["Fund Code"] == fcode]["UT"].iloc[0] == "UT"
        else hold[
            (hold["Security Type"] == "CIS")
            & (hold["Sub Security Type"] != "CSH")
            & (~hold["PrimaryAssetID"].isin(["PIMEVOA", "PIMIDFA"]))
        ]["Current Exposure"].count()
    )  # other, non-MMF, UTs

    other_ETFs = (
        0
        if twoA[twoA["Fund Code"] == fcode]["UT"].iloc[0] == "UT"
        else hold[
            (hold["Security Type"] == "ETF") & (hold["Sub Security Type"] != "CSH")
        ]["Current Exposure"].sum()
    )  # other, non-MMF, ETFs

    count_other_ETFs = (
        0
        if twoA[twoA["Fund Code"] == fcode]["UT"].iloc[0] == "UT"
        else hold[
            (hold["Security Type"] == "ETF") & (hold["Sub Security Type"] != "CSH")
        ]["Current Exposure"].count()
    )  # other, non-MMF, ETFs

    fras = min(
        0,
        hold[hold["Valuation First Level"] == "FORWARD RATE AGREEMENT"][
            "Original Nominal"
        ]
        .fillna(0)
        .dot(
            hold[hold["Valuation First Level"] == "FORWARD RATE AGREEMENT"][
                r"Market Price /Yield"
            ].fillna(0)
        ),
    )
    # FRAs as a dot product of nominals and prices

    ailf = (
        cash + mmfs + mmis + bonds + repo_gain + marg_jse + marg_otc
    )  # total assets in liquid form

    eqty_fut_frgn_mtm = -hold[
        (hold["Valuation Second Level"] == "Equity Derivatives")
        & (hold["CCY"] != "ZAR")
    ]["Sum of Market Value Income"].sum()  # mark-to-market p&l on equity futures

    bond_fut_frgn_mtm = -hold[
        (hold["Valuation Second Level"] == "Bond Derivatives") & (hold["CCY"] != "ZAR")
    ]["Sum of Market Value Income"].sum()  # mark-to-market p&l on bond futures

    trs_neg_mtm = -min(
        0,
        hold[(hold["Sub Security Type"] == "TRS")]["Sum of Market Value Income"].sum(),
    )  # net negative mtm on OTC derivativesotcs      = fwds + min(0, trs_neg_mtm) + repo_loss + fras                              # total OTC derivatives

    eqty_futs = max(
        0,
        hold[(hold["Valuation Second Level"] == "Equity Derivatives")][
            "Current Exposure"
        ].sum(),
    )  # equity futures

    bond_futs = max(
        0,
        hold[(hold["Valuation Second Level"] == "Bond Derivatives")][
            "Current Exposure"
        ].sum(),
    )  # bond futures

    frgn_futs = 0

    otcs = fwds + min(0, trs_neg_mtm) + repo_loss + fras

    # frgn_trs  = hold[(hold['Sub Security Type'] == 'TRS') & (hold['CCY'] != 'ZAR')]['Current Exposure'].sum() # profit or loss on total return swaps

    lstd_drvs = (
        eqty_futs + bond_futs + eqty_fut_frgn_mtm + bond_fut_frgn_mtm + trs_neg_mtm
    )

    eqty_fut_frgn_mtm = -hold[
        (hold["Valuation Second Level"] == "Equity Derivatives")
        & (hold["CCY"] != "ZAR")
    ]["Sum of Market Value Income"].sum()  # mark-to-market p&l on equity futures

    bond_fut_frgn_mtm = -hold[
        (hold["Valuation Second Level"] == "Bond Derivatives") & (hold["CCY"] != "ZAR")
    ]["Sum of Market Value Income"].sum()  # mark-to-market p&l on bond futures

    # leverage
    excl = ["CASH", "MONEY MARKET", "UNKNOWN", "SYTH"]
    lvg_g = (
        hold[~hold["Valuation First Level"].isin(excl)]["Current Exposure"].abs().sum()
        / nav
        * 100
    )  # leverage gross
    lvg_c = (
        hold[~hold["Valuation First Level"].isin(excl)]["Current Exposure"].sum()
        / nav
        * 100
    )  # leverage commitment (net)

    net_eff_exp = -min(
        0, hold[hold["Investment Type"] == "SYTH"]["Current Exposure"].sum() / nav
    )
    short_cash = -min(
        0,
        hold[
            (
                (hold["Valuation First Level"] == "CASH")
                & (~hold["PrimaryAssetID"].isin(["PRMFB3", "PCMMB3"]))
            )
        ]["Current Exposure"].sum()
        / nav,
    )  # cash excl MMFs < 0

    sh["E16"] = net_eff_exp
    sh["E17"] = short_cash
    sh["E18"] = sh["E16"].value + sh["E17"].value
    sh["E19"] = lvg_g
    sh["E20"] = lvg_c

    # assign cell values
    sh["B6"] = ailf
    sh["B15"] = -max(0, crry_derv)  # total short currency derivatives
    sh["B18"] = otcs
    sh["B24"] = ailf - max(0, crry_derv) + otcs
    sh["B26"] = eqty_futs + bond_futs + eqty_fut_frgn_mtm + bond_fut_frgn_mtm
    sh["B34"] = ailf - max(0, crry_derv) + otcs - lstd_drvs
    sh["B38"] = ailf - max(0, crry_derv) + otcs - lstd_drvs + other_UTs + other_ETFs
    sh["A38"] = (
        f"{'A' if sh['B38'].value > 0 else 'Ina'}dequate cash cover for derivatives"
    )

    sh["E1"] = dervs
    sh["F1"] = f"derivative{'' if dervs == 1 else 's'}"

    sh[
        "G1"
    ].hyperlink = (
        rf"P:\Investment Operations\GRC\Compliance\Client Mandates\{fcode} Rules.docx"
    )
    sh["G1"].value = "Mandate"  # Optional: Change the display text
    sh[
        "G2"
    ].value = rf"P:\Investment Operations\GRC\Compliance\Derivative Cover\{fcode} Derv Calc {rptDate.strftime('%d%b%Y')}.xlsx"

    sh["B4"] = nav
    sh["B7"] = cash
    sh["B8"] = mmfs
    sh["B9"] = mmis
    sh["A9"] = f"Money market instruments{
        ' excl CLNs and excl > \
13 month bonds'
        if ftyp == 'UT'
        else ''
    }"
    sh["B10"] = bonds
    sh["B11"] = repo_gain
    sh["B12"] = marg_jse
    sh["B13"] = marg_otc
    sh["B16"] = -max(0, crry_derv)
    sh["B19"] = fwds
    sh["B20"] = min(0, trs_neg_mtm)
    sh["B21"] = repo_loss
    sh["B22"] = fras
    sh["A24"] = f"{'A' if ailf - crry_derv + otcs > 0 else 'Ina'}dequate \
cash cover for currency derivatives"
    sh["B27"] = eqty_futs
    sh["B28"] = bond_futs
    sh["B29"] = eqty_fut_frgn_mtm + bond_fut_frgn_mtm
    sh["B31"] = trs_neg_mtm
    sh["B32"] = trs_neg_mtm
    sh["B36"] = 0 if ftyp == "UT" else other_UTs + other_ETFs
    sh["A36"] = (
        ""
        if ftyp == "UT"
        else f"Cash from the {count_other_UTs + count_other_ETFs} \
non-MMF underlying \
UT{'s' if count_other_UTs + count_other_ETFs != 1 else ''}"
    )

    # iterate through cells in the specified column C to
    # give corresponding percentage values
    for row in range(6, 39):  # from cell "C6" to cell "C38"
        # print(fcode, "B" + str(row), type("B" + str(row)),sh["B" + str(row)].value)
        if (
            isinstance(sh["B" + str(row)].value, float)
            and sh["B" + str(row)].value is not None
        ):
            sh["C" + str(row)].value = sh["B" + str(row)].value / nav * 100
            # print(fcode, "B" + str(row), type("B" + str(row)),sh["B" + str(row)].value)

    # derivative summaries
    n = 9  # column 'H' + 1
    r = 1  # 'H2' SA equity futures
    futs_eq_sa = hold[
        (hold["Valuation Second Level"] == "Equity Derivatives")
        & (hold["CCY"] == "ZAR")
    ]
    sh.cell(r, n - 1).value = (
        "SA equity futures (" + str(len(futs_eq_sa["i Issue Name"])) + ")"
    )
    if len(futs_eq_sa) > 0:
        sh.cell(r + 1, n - 1).value = futs_eq_sa["Current Exposure %"].sum() / 100
        sh["E13"] = 0 if sh["H2"].value is None else min(0, sh["H2"].value)
        for k in range(n, len(futs_eq_sa) + n):
            sh.cell(r, k).value = futs_eq_sa["i Issue Name"].iloc[k - n]
            sh.cell(r + 1, k).value = futs_eq_sa["Current Exposure %"].iloc[k - n] / 100
    sh["E14"] = sh["E13"].value + sh["E12"].value
    sh["F14"] = f"{'A' if sh['E14'].value >= 0 else 'Ina'}dequate short \
    SA equity futures cover [BN90 16(1)(a) & (b)]"

    r = 4  # 'H5' ex-SA equity futures
    futs_eq_wo = hold[
        (hold["Valuation Second Level"] == "Equity Derivatives")
        & (hold["CCY"] != "ZAR")
    ]
    sh.cell(r, n - 1).value = (
        "Ex-SA equity futures (" + str(len(futs_eq_wo["i Issue Name"])) + ")"
    )
    if len(futs_eq_wo) > 0:
        sh.cell(r + 1, n - 1).value = futs_eq_wo["Current Exposure %"].sum() / 100
        sh["E9"] = 0 if sh["H5"].value is None else min(0, sh["H5"].value)
        for k in range(n, len(futs_eq_wo) + n):
            sh.cell(r, k).value = futs_eq_wo["i Issue Name"].iloc[k - n]
            sh.cell(r + 1, k).value = futs_eq_wo["Current Exposure %"].iloc[k - n] / 100
    sh["E10"] = sh["E9"].value + sh["E8"].value
    sh["F10"] = f"{'A' if sh['E10'].value >= 0 else 'Ina'}dequate \
    short foreign equity futures cover [BN90 16(1)(a) & (b)]"

    r = 7  # 'H8' SA bond futures
    futs_bd_sa = hold[
        (hold["Valuation Second Level"] == "Bond Derivatives") & (hold["CCY"] == "ZAR")
    ]
    sh.cell(r, n - 1).value = (
        "SA bond futures (" + str(len(futs_bd_sa["i Issue Name"])) + ")"
    )
    if len(futs_bd_sa) > 0:
        sh.cell(r + 1, n - 1).value = futs_bd_sa["Current Exposure %"].sum() / 100
        sh["E5"] = 0 if sh["H8"].value is None else min(0, sh["H8"].value)
        for k in range(n, len(futs_bd_sa) + n):
            sh.cell(r, k).value = futs_bd_sa["i Issue Name"].iloc[k - n]
            sh.cell(r + 1, k).value = futs_bd_sa["Current Exposure %"].iloc[k - n] / 100
    sh["E6"] = sh["E5"].value + sh["E4"].value
    sh["F6"] = f"{'A' if sh['E6'].value >= 0 else 'Ina'}dequate short \
    SA bond futures cover [BN90 16(1)(a) & (b)]"

    r = 10  # 'H11' ex-SA bond futures
    futs_bd_wo = hold[
        (hold["Valuation Second Level"] == "Bond Derivatives") & (hold["CCY"] != "ZAR")
    ]
    sh.cell(r, n - 1).value = (
        "Ex-SA bond futures (" + str(len(futs_bd_wo["i Issue Name"])) + ")"
    )
    if len(futs_bd_wo) > 0:
        sh.cell(r + 1, n - 1).value = futs_bd_wo["Current Exposure %"].sum() / 100
        for k in range(n, len(futs_bd_wo) + n):
            sh.cell(r, k).value = futs_bd_wo["i Issue Name"].iloc[k - n]
            sh.cell(r + 1, k).value = futs_bd_wo["Current Exposure %"].iloc[k - n] / 100

    r = 13  # 'H14' currency futures
    futs_cr = hold[hold["Valuation Second Level"] == "Currency Derivatives"]
    # futs_cr = hold[hold["Valuation Second Level"] == "Currency Derivatives"]
    # # minus added 16Feb 2026 to align wit5h manual report calc
    sh.cell(r, n - 1).value = (
        "Currency futures (" + str(len(futs_cr["i Issue Name"])) + ")"
    )
    if len(futs_cr) < 0:
        sh.cell(r + 1, n - 1).value = futs_cr["Current Exposure %"].sum() / 100
        for k in range(n, len(futs_cr) + n):
            sh.cell(r, k).value = futs_cr["i Issue Name"].iloc[k - n]
            sh.cell(r + 1, k).value = futs_cr["Current Exposure %"].iloc[k - n] / 100

    r = 16  # 'H17' currency forwards
    fwds_cr = hold[hold["Valuation Second Level"] == "Forwards"]
    sh.cell(r, n - 1).value = (
        "Currency forwards (" + str(len(fwds_cr["i Issue Name"])) + ")"
    )
    if len(fwds_cr) > 0:
        sh.cell(r + 1, n - 1).value = fwds_cr["Current Exposure %"].sum() / 100
        for k in range(n, len(fwds_cr) + n):
            sh.cell(r, k).value = fwds_cr["i Issue Name"].iloc[k - n]
            sh.cell(r + 1, k).value = fwds_cr["Current Exposure %"].iloc[k - n] / 100

    r = 19  # 'H20' SA equity TRSes
    swps_eq_sa = hold[
        (hold["Valuation Second Level"] == "SWAPS")
        & (hold["Sub Security Type"] == "TRS")
        & (hold["CCY"] == "ZAR")
    ]
    sh.cell(r, n - 1).value = (
        "SA equity TRSes (" + str(int(len(swps_eq_sa["i Issue Name"]) / 3)) + ")"
    )
    if len(swps_eq_sa) > 0:
        sh.cell(r + 1, n - 1).value = swps_eq_sa["Current Exposure %"].sum() / 100
        for k in range(n, len(swps_eq_sa) + n):
            sh.cell(r, k).value = swps_eq_sa["i Issue Name"].iloc[k - n]
            sh.cell(r + 1, k).value = swps_eq_sa["Current Exposure %"].iloc[k - n] / 100

    r = 22  # 'H23' ex-SA equity TRSes
    swps_eq_wo = hold[
        (hold["Valuation Second Level"] == "SWAPS")
        & (hold["Sub Security Type"] == "TRS")
        & (hold["CCY"] != "ZAR")
    ]
    sh.cell(r, n - 1).value = (
        "Ex-SA equity TRSes (" + str(int(len(swps_eq_wo["i Issue Name"]) / 3)) + ")"
    )
    if len(swps_eq_wo) > 0:
        sh.cell(r + 1, n - 1).value = swps_eq_wo["Current Exposure %"].sum() / 100
        for k in range(n, len(swps_eq_wo) + n):
            sh.cell(r, k).value = swps_eq_wo["i Issue Name"].iloc[k - n]
            sh.cell(r + 1, k).value = swps_eq_wo["Current Exposure %"].iloc[k - n] / 100

    r = 25  # 'H26' repo trades
    repo = hold[
        (hold["PrimaryAssetID"].str.startswith("RPCO"))
        | (hold["PrimaryAssetID"].str.startswith("RPCA"))
    ]
    sh.cell(r, n - 1).value = "Repos (" + str(int(len(repo) / 2)) + ")"
    if len(repo) > 0:
        sh.cell(r + 1, n - 1).value = repo["Current Exposure %"].sum() / 100
        for k in range(n, len(repo) + n):
            sh.cell(r, k).value = repo["i Issue Name"].iloc[k - n]
            sh.cell(r + 1, k).value = repo["Current Exposure %"].iloc[k - n] / 100

    # save the file and close openpyxl
    filename = os.path.join(
        pthLOCAL, f"{fcode} Derv Calc {rptDate.strftime('%d%b%Y')}.xlsx"
    )
    wb.save(filename)
    wb.close
    # os.startfile(filename)  # open the file for review

print(
    f" {timediff(start_time_compiling, time.time())} compiling \
derivative calculation files for {len(fnames_incompl)} funds \
at {rptDate.strftime('%d %b %Y')}"
)

# Save the fund derivative cover calc files to the Derivative Cover folder
start_time = time.time()
ked = os.listdir(pthLOCAL)

# create and save the derivative cover summary file in the Exports folder
# and then clear the temporary local folder copy the locally (C:\) stored
# derivative calc files to a PIM network folder
for file in tqdm(
    ked,
    desc=f"Saving the {len(ked)} fund derivative calculation \
files to the Derivative Cover folder ...",
):
    shutil.copy(os.path.join(pthLOCAL, file), pthEXPORTS)

print(
    f"\n {timediff(start_time, time.time())} saving the {len(ked)} fund \
derivative calculation files to Derivative Cover folder\n"
)

# Delete contents of the temporary local folder
start_time = time.time()

local_folder_delete = "yes"
if local_folder_delete == "yes":
    for filename in tqdm(
        os.listdir(pthLOCAL), desc="Deleting contents of the temporary folder ..."
    ):
        file_path = os.path.join(pthLOCAL, filename)
        try:
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)
        except Exception as e:
            print(f"Couldn't delete {file_path} because {e}")

print(
    f" {timediff(start_time, time.time())} deleting \
contents of the temporary folder completed"
)

print(
    f"\n\n{timediff(start_time_derv_compiling, time.time())} \
total time to compile and save {len(fnames)} \
derivative calculation files for {rptDate.strftime('%d %b %Y')}"
)

print("\n\n#################################################")
print("#                                               #")
print("#       END 2/4 derv_checker_compiling.py X     #")
print("#                                               #")
print("#################################################\n\n")
