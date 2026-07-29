#!/usr/bin/env python
# coding: utf-8

# # Utility Functions for Use in Other Python Scripts

# ## To use any of the functions in this notebool add this line when importing libraries:
#
# from utilties import timediff, osprey, last_working_day, report_types_dict (, etc.) which worls well in .py and in .ipynb files
#
# %run utilities.ipynb works only in .ipynb files


# https://stackoverflow.com/questions/44116194/import-a-function-from-another-ipynb-file
# !pip install ipynb
# !pip show openpyxl

import sys
import time
from datetime import datetime


def timediff(timestamp_ante: float, timestamp_post: float) -> str:
    """
    Calculates the difference between two timestamps and returns it in days, hours, minutes, and seconds.

    Args:
        timestamp_ante (float): The earlier timestamp in float type
        timestamp_post (float): The later timestamp in float type

    Returns:
        str: a string of the form ..d ..hr ..min ..sec
    """

    # # Extract components
    total_seconds = abs(
        timestamp_post - timestamp_ante
    )  # int(time_difference.total_seconds())

    days = total_seconds // (24 * 3600)
    hours = (
        total_seconds % (24 * 3600)
    ) // 3600  # remaining_seconds_after_days = total_seconds % (24 * 3600)
    minutes = (
        (total_seconds % (24 * 3600)) % 3600
    ) // 60  # remaining_seconds_after_hours = remaining_seconds_after_days % 3600
    seconds = ((total_seconds % (24 * 3600)) % 3600) % 60

    if days == 0 and hours == 0 and minutes == 0:
        return f"{seconds:.1f}sec"
    elif days == 0 and hours == 0:
        return f"{minutes:.0f}min {seconds:.1f}sec"
    elif days == 0:
        return f"{hours:.0f}hr {minutes:.0f}min {seconds:.1f}sec"
    else:
        return f"{days:.0f}d {hours:.0f}hr {minutes:.0f}min {seconds:.1f}sec"


# # TEST
# timediff(39125.2356365, 310127.1254689) # timediff_EX(39125.2356365, 310127.1254689, 2) yielded '75hr 4516min 41.89sec'


# utility function to open an excel file, .xls or .xlsx
def open_xl_file(file_name_and_path):
    import win32com.client as win32  # library to convert xls to xlsx

    excel = win32.gencache.EnsureDispatch("Excel.Application")
    excel.DisplayAlerts = False  # suppress the warning dialogue
    excel.Workbooks.Open(file_name_and_path)
    excel.DisplayAlerts = True  # unsuppress the warning dialogue


# function to find Excel sheet row number of string "a", on openpyxl sheet "worksheet", in column "col"
def item_row(
    worksheet, a: str, col: int
):  # where "worksheet" is an openpyxl worksheet, "a" is a string, "col" is the column number in which to search
    for row in worksheet.iter_rows(max_col=col):
        for cell in row:
            if cell.value == a:
                return cell.row


def range_border(worksheet, col_left, col_right, row_top, row_bottom):
    # import openpyxl
    from openpyxl.styles.borders import Border, Side  # import openpyxl attributes

    thin_border = Side(style="thin", color="000000")  # black color
    for cl in range(col_left, col_right + 1):
        for rw in [row_top - 1, row_bottom]:
            worksheet.cell(rw, cl).border = Border(bottom=thin_border)
    for rw in range(row_top, row_bottom + 1):
        for cl in [col_left, col_right + 1]:
            worksheet.cell(rw, cl).border = Border(left=thin_border)

    # fix lines undone by previous actions
    worksheet.cell(row_bottom, col_left).border = Border(
        left=thin_border, bottom=thin_border
    )  # fix bottom left corner
    worksheet.cell(row_bottom, col_right).border = Border(
        right=thin_border, bottom=thin_border
    )  # fix bottom right corner
    worksheet.cell(row_top, col_left).border = Border(
        left=thin_border, top=thin_border
    )  # fix top left corner
    worksheet.cell(row_top, col_right).border = Border(
        right=thin_border, top=thin_border
    )  # fix top right corner

    # specific case where range is over a single row
    if row_top == row_bottom:
        worksheet.cell(row_bottom, col_left).border = Border(
            left=thin_border, bottom=thin_border, top=thin_border
        )  # fix left corner
        worksheet.cell(row_bottom, col_right).border = Border(
            right=thin_border, bottom=thin_border, top=thin_border
        )  # fix right corner

    # specific case where range is over a single column
    if col_left == col_right:
        worksheet.cell(row_top, col_left).border = Border(
            left=thin_border, right=thin_border, top=thin_border
        )  # fix top corner
        worksheet.cell(row_bottom, col_left).border = Border(
            left=thin_border, right=thin_border, bottom=thin_border
        )  # fix bottom corner

    # specific case where range is over a single cell
    if (col_left == col_right) and (row_bottom == row_top):
        worksheet.cell(row_bottom, col_left).border = Border(
            left=thin_border, right=thin_border, bottom=thin_border, top=thin_border
        )


def rows_align_height(
    worksheet,
    row_from,
    row_to,
    col,
    height_normal,
    text_length_threshold,
    height_if_text_is_long,
):
    # import openpyxl
    for row_index in range(row_from, row_to + 1):
        # Determine height based on content in row_index
        if isinstance(worksheet.cell(row=row_index, column=col).value, float):
            worksheet.row_dimensions[row_index].height = height_normal
        elif (worksheet.cell(row=row_index, column=col).value) and (
            len(worksheet.cell(row=row_index, column=col).value) > text_length_threshold
        ):
            worksheet.row_dimensions[row_index].height = height_if_text_is_long


# function to convert .xls to .xlsx using win32 given a file path to the xls file
def xlsToXlsx(filepathInclXls):
    import win32com.client as win32  # library to convert xls to xlsx

    excel = win32.gencache.EnsureDispatch("Excel.Application")
    excel.DisplayAlerts = False  # suppress the warning dialogue
    wb = excel.Workbooks.Open(filepathInclXls)
    wb.SaveAs(
        filepathInclXls + "x", FileFormat=51
    )  # FileFormat = 51 (56) for .xlsx (.xlx) extension
    wb.Close()
    excel.DisplayAlerts = True  # unsuppress Excel warning dialogue
    return print(" ", filepathInclXls + "x")


# function to identify property securities; same function is used in zips.ipynb
# import re
# from re import search


def property(txt):
    import re
    from re import search

    pattern = r"\b(\sREAL\s|PROPERT|REALTY|PPTY|\sESTATE|REIT|ATTACQ|BURSTONE|MAS PLC|NEPI ROCKCASTLE|OCTODEC|FAIRVEST|HYPROP)\b"
    if re.search(pattern, str(txt).upper()):
        return "P"


# ### TEST
# txt = 'asd nepi rockcastle sadf'
# property(txt)

# utility function to save downloaded file with reporting date in cell 'J1'


def fnd(txt):
    import re
    from re import search

    pattern = r"\b(?:FUND(?!\s*MANAGE)|UCIT|ETF|ISHARES)\b"
    if re.search(pattern, str(txt).upper()):
        return "f"


# ### TEST
# txt = 'epic fund managers'
# property(txt)


def dater(folder_path, fund_name, dte):
    start_time = time.time()

    # from pathlib import Path
    import os

    fls = os.listdir(folder_path)
    a = max(
        [os.path.abspath(os.path.join(folder_path, fl)) for fl in fls if r28N in fl],
        key=os.path.getmtime,
    )
    z = "xlsx" if a[len(a) - 3 :] == "lsx" else "xls"
    wb = excel.Workbooks.Open(a)
    wb.Worksheets("Reg 28 Report - Incl Effective ").Range("J1").Value = dte.strftime(
        "%d%b%Y"
    )
    wb.SaveAs(
        os.path.abspath(
            os.path.join(
                folder_path, f"{fund_name} lookthrough {dte.strftime('%d%b%Y')}.{z}"
            )
        )
    )
    wb.Close()

    # print(f"    Downloaded file found and saved in
    # {timediff(start_time, time.time())}")  # time to get file name


# function to get latest file with specified extension
# in a given folder and change that file's name
def latest_file(folder_path, suffix="csv", new_file_name="newt"):
    import glob, os
    from pathlib import Path

    folder_path = str(Path.home() / "Downloads")
    # https://datatofish.com/latest-file-python/
    files = glob.glob(folder_path + r"\*" + suffix)
    try:
        # latest file name
        max_file = max(files, key=os.path.getctime)
        # rename the latest created file in the folder  https://www.squash.io/how-to-rename-a-file-with-python/
        path_new = os.path.join(folder_path, new_file_name) + "." + suffix
        if os.path.exists(path_new):
            os.remove(path_new)
        os.rename(max_file, path_new)

    except ValueError as ve:
        max_file = f"No .{suffix} files in the folder. Error: {str(ve)}."
    # print(f"{max_file} -> {path_new}")
    # return


# Function to report date and selected summary sheet option
def parn_de():

    import pandas as pd
    from constants import pthPy, pth_dl
    from utilities import prior_working_day
    from datetime import datetime
    import os

    print("Getting the reporting date and names derivative files ...")

    df = pd.read_excel(pthPy, sheet_name="arc", header=None, usecols="A,E").dropna(
        subset=[0]
    )
    k = df.iloc[2, 1]
    rptDate = (
        k if isinstance(k, datetime) else prior_working_day(datetime.today())
    )  # prior working day or report date override; has type datetime()
    summ_yn = df.iloc[3, 1]
    dervthreshold = df.iloc[4, 1] * 100
    funds = df[0].iloc[1:]

    # derive holdings and derivatives file paths
    fPARN = os.path.join(
        pth_dl,
        f"PARN ({len(funds)}) {rptDate.strftime('%d%b%Y')}.csv",
    )

    fDE = os.path.join(
        pth_dl,
        f"DERV ({len(funds)}) {rptDate.strftime('%d%b%Y')}.csv",
    )

    return fPARN, fDE, funds, rptDate, summ_yn, dervthreshold


# An Eagle report lookup function, given seven parameters
def osprey(rpt_type, funds, d_from, d_to, name, sfx="csv"):
    """
    Function:
        To download a report from the online fund accounting
        system for a specified report type, funds, format, and dates

    Args:
        rpt_type: A report name under Queries of the Eagle
                  online application, including r28i, parn,
                  derv, trad, scty, dflw, utps, fnav, tcrf, and cact
        funds:    A comma-, but without spaces, separated string of
                  fund codes, including the ones appended
                  with "_C", .e.g., 'PABS,PPSBAL_C,SMMAIF'
        d_from:   A start date for the report in datetime
                  format, e.g., datetime(2025,5,1)
        d_to:     An  end date for the report in datetime
                  format, e.g., datetime(2025,5,30)
        name:     A descriptive name to be added to the downloaded
                  report to make it more identifiable
        sfx:      A file name extension specifying the report
                  format, i.e., 'xls' or 'csv'
        al:       The user name for the online application
        xe:       The user credential for the online application

    Returns:
        A downloaded report in the local Downloads folder
        renamed to identify it
    """

    # (1) present expected file name and whether it already exists in the Downloads folder, before downloading the report
    to_date = f" to {d_to.strftime('%d%b%Y')}" if d_to != d_from else ""

    new_file_name = (
        f"{rpt_type.upper()} {name}({len(funds.split(','))}) "
        f"{d_from.strftime('%d%b%Y')}{to_date}"
    )

    # (2) load libraries
    print("(2) loading libraries")
    import time

    start_time_osprey = time.time()

    # from datetime import datetime  # , timedelta
    from utilities import timediff, latest_file
    import os, re
    import pandas as pd
    from constants import pth_dl, eagle_default, report_types_dict

    file_exists = (pth_dl / f"{new_file_name}.{sfx}").exists()
    print(
        f"Expected file name based on input values:\n"
        f"   {new_file_name}.{sfx}\n"
        f"which {'already exists' if file_exists else 'does not yet exist'} "
        f"in the Downloads folder.\n"
    )

    # access environment variables
    from dotenv import load_dotenv
    # to access environment variables from .env file

    load_dotenv()  # take environment variables from .env file

    # selenium suite of tools
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys

    # from selenium.webdriver.support.select import Select
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import NoSuchElementException
    from selenium.common.exceptions import (
        NoAlertPresentException,
    )  # to handle the eagleportal.prescient.co.za alerts
    from selenium.common.exceptions import TimeoutException

    # print(funds)

    print(
        "(2a) Checking if rpt_type is 'fnav', \
in which case, replacing '_C' in fund name"
    )
    # (2a) for fnav rpt_type, first remove "_C" from the list of funds else the FNAV report will return "No data returned for the input criteria."
    if rpt_type == "fnav":
        sfx = "csv"
        lkup = pd.DataFrame(
            pd.Series(funds.split(",")), columns=["funds_ante"]
        )  # create a dataframe to look up before and after fund codes
        lkup["funds_post"] = lkup["funds_ante"].apply(
            lambda x: x.replace("_C", "") if x.endswith("_C") else x
        )
        funds = ",".join(lkup["funds_post"].astype(str))

    print("(3) setting report suffix for .xls vs .csv")
    # (3) set the report suffix
    t = "0" if sfx == "csv" else "4"  # report format: DXI4[0] for .xls[.csv]

    max_retries = 5
    batch_succeeded = False
    for attempt in range(1, max_retries + 1):
        print(f"(4) Importing the webdriver (attempt {attempt}/{max_retries})")
        # (4) assign the browser driver
        from selenium import webdriver

        driver = webdriver.Firefox()

        print("(5) Starting the web driver and opening the browser ...")
        # (5) open the browser on the default web page
        # eagle_default = r"https://eagleportal.prescient.co.za/Default.aspx"
        driver.get(eagle_default)  # default page
        wait = WebDriverWait(driver, 10)
        # https://selenium-python.readthedocs.io/waits.html,
        # max wait for elements to appear

        print("(6) Browser, open sesame! ...")
        # (6) log in
        driver.find_element(
            By.CSS_SELECTOR, "#LoginCtrl_MainLoginControl_UserName"
        ).send_keys(os.getenv("EAGLE_UN"))
        driver.find_element(
            By.CSS_SELECTOR, "#LoginCtrl_MainLoginControl_Password"
        ).send_keys(os.getenv("EAGLE_PW"))
        driver.find_element(
            By.CSS_SELECTOR, "#LoginCtrl_MainLoginControl_LoginButton"
        ).click()

        print("(7) having logged in, opening the selected report page ...")
        # (7) having logged in, open the selected report page
        report_link = report_types_dict[rpt_type][1]
        driver.get(
            report_link
        )  # a hyperlink for the report page selected in the function osprey()

        try:
            print(
                "(7a) testing for the presence of an alert and accepting it if it exists"
            )
            # (7(a)) test for the presence of an alert
            # this solution from Gemini prompt 17 Sep 2025: "python selenium test
            # for the presence of alert text"
            try:
                WebDriverWait(driver, 3).until(EC.alert_is_present())
                alert = driver.switch_to.alert
                alert.accept()  # Or alert.dismiss()
            except TimeoutException:
                # print("No alert present per TimeoutException.")
                pass
            except NoAlertPresentException:
                # print("No alert present per NoAlertPresentException.")
                pass

            print("(8) switching to the query page ...")
            # (8) switch to the query page
            driver.find_element(By.CSS_SELECTOR, "#ModifyLinkLabel").click()

            print("(9) updating the FROM calendar ...")
            # (9) update the FROM calendar
            date_selector_fr = driver.find_element(
                By.CSS_SELECTOR,
                'input[id="ctl00_c_qc_QueryInputs_QueryInputsPopup_DATE1_DateCtrl_From_I"]',
            )  # FROM date element
            driver.execute_script(
                f'arguments[0].value = "{d_from.strftime("%#m/%#d/%Y")}";',
                date_selector_fr,
            )  # FROM date without leading zeroes
            driver.find_element(
                By.CSS_SELECTOR,
                'input[id="ctl00_c_qc_QueryInputs_QueryInputsPopup_DATE1_DateCtrl_From_I"]',
            ).click()  # click inside FROM calendar
            driver.find_element(
                By.CSS_SELECTOR,
                'td[id="ctl00_c_qc_QueryInputs_QueryInputsPopup_DATE1_DateCtrl_From_B-1"]',
            ).click()  # update the FROM calendar

            print("(10) updating the TO calendar, if it exists ...")
            # (10) if it exists, update the TO calendar
            try:  # https://stackoverflow.com/questions/38022658/selenium-python-handling-no-such-element-exception
                date_selector_to = driver.find_element(
                    By.CSS_SELECTOR,
                    'input[id="ctl00_c_qc_QueryInputs_QueryInputsPopup_DATE1_DateCtrl_To_I"]',
                )  # calendar
                driver.execute_script(
                    f'arguments[0].value = "{d_to.strftime("%m/%d/%Y")}";',
                    date_selector_to,
                )
                driver.find_element(
                    By.CSS_SELECTOR,
                    'input[id="ctl00_c_qc_QueryInputs_QueryInputsPopup_DATE1_DateCtrl_To_I"]',
                ).click()  # click inside TO calendar
                driver.find_element(
                    By.CSS_SELECTOR,
                    'td[id="ctl00_c_qc_QueryInputs_QueryInputsPopup_DATE1_DateCtrl_To_B-1"]',
                ).click()  # update the TO calendar
            except (
                NoSuchElementException
            ):  # in the event that the selected report does not have a "to" calendar
                pass

            print(
                "(11) getting the web element for the FUND LIST and assigning values to it ..."
            )
            # (11) get the web element for the FUND LIST and assign values to it
            ct100_SelectedIds = (
                'input[id="ctl00_c_qc_QueryInputs_QueryInputsPopup_FUND0_SelectedIds"]'
            )
            fund_selector = driver.find_element(
                By.CSS_SELECTOR,
                ct100_SelectedIds,
            )
            driver.execute_script(f'arguments[0].value = "{funds}";', fund_selector)

            print("(12) clicking 'Entity ID' to trigger the report generation ...")
            # (12) click the table header where "Entity ID" resides
            ct100_FUND0 = 'table[id="ctl00_c_qc_QueryInputs_QueryInputsPopup_FUND0_SelectedItemsGrid_DXHeaderTable"]'
            WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, ct100_FUND0))
            ).click()  # fund code banner
            time.sleep(5)  # arbitrary 5 second wait

            print(
                "(12a) testing for the presence of an authentication \
alert after clicking Submit ..."
            )
            # (12a) test for the presence of an authentication alert after submitting the report query
            try:
                WebDriverWait(driver, 10).until(EC.alert_is_present())
                alert = driver.switch_to.alert
                alert.send_keys(
                    os.getenv("EAGLE_UN") + Keys.TAB + os.getenv("EAGLE_PW")
                )
                alert.accept()
            except TimeoutException:
                pass
            except NoAlertPresentException:
                pass

            print(
                "(13) getting the web element of the 'Submit' \
button and then clicking it ..."
            )
            # (13) get the web element of the 'Submit' button and then click it
            submit_button = driver.find_element(
                By.CSS_SELECTOR,
                'input[id="ctl00_c_qc_QueryInputs_QueryInputsPopup_RunBtn"]',
            )
            submit_button.click()

            # added by Claude Code Interpreter on 17 Sep 2025 in response to
            # "python selenium test for the presence of alert text"
            print(
                "(13a) testing for the presence of an authentication \
alert after clicking Submit ..."
            )
            # (13a) test for the presence of an authentication alert after submitting the report query
            try:
                WebDriverWait(driver, 10).until(EC.alert_is_present())
                alert = driver.switch_to.alert
                alert.send_keys(
                    os.getenv("EAGLE_UN") + Keys.TAB + os.getenv("EAGLE_PW")
                )
                alert.accept()
            except TimeoutException:
                pass
            except NoAlertPresentException:
                pass

            print("(14) waiting for and then clicking the export button ...")
            # (14) Wait for and then click the export button and then the xls download button
            # https://stackoverflow.com/questions/56085152/selenium-python-error-element-could-not-be-scrolled-into-view
            start_time_14 = time.time()
            WebDriverWait(driver, 1000).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, 'a[id="DistrBtn"]'))
            ).click()

            # timeout_14 = 1000
            # is_clickable_14 = EC.element_to_be_clickable(
            #     (By.CSS_SELECTOR, 'a[id="DistrBtn"]')
            # )
            # export_button = False
            # while not export_button:
            #     export_button = is_clickable_14(driver)
            #     sys.stdout.write(
            #         f"\r  ... waited {timediff(start_time_14, time.time())} "
            #         "so far for the export format button to become clickable"
            #     )
            #     sys.stdout.flush()
            #     if not export_button:
            #         if time.time() - start_time_14 > timeout_14:
            #             raise TimeoutException(
            #                 f"export format button not clickable after {timeout_14}s"
            #             )
            #         time.sleep(0.5)
            # sys.stdout.write("\n")
            # export_button.click()
            print(
                f"  Waited {timediff(start_time_14, time.time())} for \
the export format button to be clickable and then clicked it ..."
            )

            start_time_14a = time.time()
            print(
                "(14a) Waiting for and then clicking the export \
format button ..."
            )
            # (14a) Waiting for and then clicking the export format button ...
            WebDriverWait(driver, 1000).until(
                EC.element_to_be_clickable(
                    (By.CSS_SELECTOR, f'td[id="ExportMnu_DXI{t}_T"]')
                )
            ).click()
            print(
                f"  Waited {timediff(start_time_14a, time.time())} for \
the download button to be clickable and then clicked it ..."
            )

            print("(14b) waiting for the download to complete ...")
            # (14b) waiting for the download to complete ...
            start_time_dl = time.time()
            t_dl = 30  # downloaded within the past t_dl seconds
            tz = 60  # max wait time of 1 minutes for download to complete
            # and moving on to the next batch, to avoid getting stuck
            # on a batch if the download gets stuck for some reason
            dl_pattern = re.compile(
                rf"{re.escape(report_types_dict[rpt_type][0])}.*\.{sfx}$"
            )
            while True:
                candidates = [
                    f
                    for f in pth_dl.iterdir()
                    if dl_pattern.search(f.name)
                    and time.time() - f.stat().st_mtime <= t_dl
                ]  # report type, suffix and downloaded within the past 30 seconds
                if not list(pth_dl.glob("*.part")) and candidates:
                    break
                if time.time() - start_time_dl > tz:
                    print(
                        f"  Warning: timed out waiting for download \
after {tz} seconds"
                    )
                    break
                time.sleep(1)
            print(f"  Download completed after {timediff(start_time_dl, time.time())}")

            print("(14c) renaming the downloaded file ...")
            #   (14c) rename the downloaded file
            folder_path = str(pth_dl)

            latest_file(folder_path, sfx, new_file_name)

            print(f"  Downloaded and renamed to: {new_file_name}")

            if rpt_type == "fnav":
                filen = os.path.join(folder_path, new_file_name + f".{sfx}")
                fnav = pd.read_csv(filen)
                fnav = fnav.merge(
                    lkup, how="left", left_on="NAV Entity ID", right_on="funds_post"
                )
                fnav["NAV Entity ID"] = fnav["funds_ante"]
                fnav.drop(columns=["funds_ante", "funds_post"], axis=1, inplace=True)
                fnav.to_csv(filen, index=False)

            print("(14d) closing the web driver ...")
            # (14d) close the driver now that the file is fully downloaded and renamed
            driver.quit()
            batch_succeeded = True
            break  # success — exit the retry loop

        except Exception as e:
            print(f"Exception in steps 7a onwards: {e}")
            try:
                driver.close()
                driver.quit()
            except Exception:
                pass
            if attempt < max_retries:
                print(
                    f"  Retrying {new_file_name}.{sfx} from step \
4 (attempt {attempt + 1}/{max_retries}) ..."
                )
            else:
                print(
                    f"  All {max_retries} {new_file_name}.{sfx} \
attempts failed for attempt {max_retries}."
                )

    print(f"\nDownloaded {new_file_name}.{sfx}")
    return new_file_name


# calendar functions for prior month end and for most recent working day

from datetime import timedelta
import calendar  # https://stackoverflow.com/questions/42950/get-the-last-day-of-the-month
import holidays  #!pip install holidays

za_holidays = holidays.ZA()
ie_holidays = holidays.IE()
us_holidays = holidays.US()
eu_holidays = holidays.ECB()
ny_holidays = holidays.NYSE()

# example usage
list(holidays.ZA(years=2025))

# calendar functions for prior month end and for most recent working day


# function for prior month end
def prior_month_end(given_date: datetime = datetime.today()) -> datetime:
    # https://stackoverflow.com/questions/2489669/how-do-python-functions-handle-the-types-of-parameters-that-you-pass-in
    last_month = given_date.replace(day=1) - timedelta(
        days=1
    )  # first day of current month minus 1 day
    return last_month


# function for prior working day per copilot 11 Dec 2024
def prior_working_day(given_date: datetime = datetime.today()) -> datetime:
    # initialise South African public holidays
    za_holidays = holidays.ZA()
    # find the previous business day
    prior_day = given_date - timedelta(days=1)  # given date minus one now
    while prior_day.weekday() >= 5 or prior_day in za_holidays:
        prior_day -= timedelta(days=1)
    return prior_day


# function for last working day of the month of a given date
def last_working_day(given_date: datetime = datetime.today()) -> datetime:
    from datetime import datetime, timedelta
    import calendar  # https://stackoverflow.com/questions/42950/get-the-last-day-of-the-month
    import holidays  #!pip install holidays

    # initialise South African public holidays
    za_holidays = holidays.ZA()
    # find last day of the month
    date_lwdom = datetime(
        given_date.year,
        given_date.month,
        calendar.monthrange(given_date.year, given_date.month)[1],
    )
    # https://stackoverflow.com/questions/42950/get-the-last-day-of-the-month
    while date_lwdom.weekday() >= 5 or date_lwdom in za_holidays:
        date_lwdom -= timedelta(days=1)
    return date_lwdom


# # example usages of the prior_working_day(), prior_month_end(), and prior_working_day(prior_month_end()) functions

# print(f"                          prior_working_day()  = {prior_working_day()}")

# user_date = datetime(2024, 12, 2).date()
# print(f"                prior_working_day({user_date})  = {prior_working_day(user_date)}")

# user_date = datetime(2023, 4, 13).date()
# print(f"                  prior_month_end({user_date})  = {prior_month_end(user_date)}")

# user_date = datetime(2023, 12, 27).date()
# print(f"prior_working_day(prior_month_end({user_date})) = {prior_working_day(prior_month_end(user_date))}")

# user_date = datetime(2025, 12, 2).date()
# last_working_day(prior_month_end(datetime.today()))
# print(f"last_working_day(prior_month_end({user_date}))  = {last_working_day(prior_month_end(user_date))}")

# # no date specified (in which case the functions default to toay's date as a parameter:
# print(f"                           prior_working_day() = {prior_working_day().date()}")
# print(f"                             prior_month_end() = {prior_month_end().date()}")

# print(prior_working_day(), type(prior_working_day()))
# print(last_working_day(), type(last_working_day()))

"""
Gemini

Query: "get name of latest file in folder"

Response: This Python script first constructs a list of 
full file paths within the specified folder. It then uses 
max() with os.path.getmtime as the key to find the file 
with the latest modification time. Finally, 
os.path.basename() extracts just the file name.
"""


# def latest_file_in_folder(fldr_pth: str = Path.home() / 'Downloads') -> str:
def latest_file_in_folder(folder_path: str) -> str:
    import os

    files = [
        os.path.join(folder_path, f)
        for f in os.listdir(folder_path)
        if os.path.isfile(os.path.join(folder_path, f))
    ]
    if files:
        # return os.path.basename(max(files, key=os.path.getmtime)) #  long file name; type = NoneType
        return max(files, key=os.path.getmtime)  # short file name; type = NoneType
    else:
        print("No files found in the folder.")


def tmv(
    df,
    by_col="Entity Name",
    num_cols_curr=["End Market Value", "Closing Exposure PA"],
    num_cols_pct=["Percentage of Market Value", ""],
):
    """
    Function:
        change the '% of Total Market Value" column per fund

    Args:
        df                  : the dataframe to have its columns converted to type float of type pandas.core.frame.DataFrame
        by_col              : the column heading containing the fund codes or names; type is 'str'
        numeric_columns_dict: a dictionary of columns to be converted to per fund percentages
        ee_col              : name of the effective exposure column
        create_new_pctCE_col: boolean True or False to create a new percent effective exposure column

    Returns:
        dataframe with percentage columns converted on a per fund level
    """
    import pandas as pd

    # # TEST for presence of inputs
    # print(list(df[by_col].unique()))
    # print(by_col)
    # print(num_cols_curr)
    # print(num_cols_pct)
    # print('\n\n')

    # (1) create variables
    funds = list(df[by_col].unique())
    pct_ee_col = num_cols_pct[1] == ""
    pct_ee = (
        "% Current Exposure" if pct_ee_col else num_cols_pct[1]
    )  # create a 'percent effective exposure' heading if not already present
    pairs = {
        num_cols_curr[0]: num_cols_pct[0],
        num_cols_curr[1]: pct_ee,
    }  # dictionary of currency and associated percentage columns

    # (2) convert currency columns from dtype object to dtype float
    for pair in pairs:
        if df[pair].dtype == "float":
            pass
        else:
            df[pair] = (
                df[pair].str.replace(",", "").astype(float)
            )  # convert currency columns to float

    # (3) get fund NAVs and divide by 100 in anticpation of percent calculation later
    navs = df.groupby(by_col, as_index=False)[list(pairs.keys())].sum()
    # print(navs, '\n\n')
    for pair in pairs:
        navs[pair] = navs[pair] / 100
    # print(navs, '\n\n')

    # (4) add the effective exposure percentage column to the dataframe if not there already
    df[pct_ee] = df[num_cols_pct[0]]

    # (5) calculate percent columns
    for fund in funds:
        for pair, pct_col in pairs.items():
            nav = (
                navs.loc[navs[by_col] == fund][pair].item()
            )  # series.item() to convert an array of size 1 (only) to a Python scalar
            df.loc[df[by_col] == fund, pct_col] = df.loc[df[by_col] == fund, pair] / nav

    # print(df.groupby(by_col, as_index = False)[list(pairs.values())].sum())

    # print(df.groupby(by_col, as_index = False)[[num_cols_curr[0], num_cols_pct[0], num_cols_curr[1], pct_ee]].sum())

    # if pct_ee_col:
    df.drop([pct_ee], axis=1, inplace=True) if pct_ee_col else df

    # # delete NaN market value rows and zero effective exposure rows
    # len_before = len(df)
    # rowsNaN  = len(df[df[num_cols_curr[0]].isnull()])
    # df = df[df[num_cols_curr[0]].notnull()]                                              # delete NaN MV column rows
    # len_after = len(df)
    # df = df[(round(df[num_cols_curr[0]],2) != 0) | (round(df[num_cols_curr[1]],2) != 0)] # delete rows where MV and EE are zero to two decimals
    # len_after = len(df)

    return df


# # TEST tmv()
# tmv(df, 'Entity Name', num_cols_curr, num_cols_pct)


def rem_zeroes(
    df, by_col="Entity Name", cols=["End Market Value", "Closing Exposure PA"]
):
    """to remove NaN market value and zero value effective exposure rows from a dataframe of fund holdings ..."""
    import pandas as pd

    len_before = len(df)
    rowsNaN = len(df[df[cols[0]].isnull()])
    df = df[df[cols[0]].notnull()]  # delete NaN MV column rows
    len_after = len(df)
    df = df[
        (round(df[cols[0]], 2) != 0) | (round(df[cols[1]], 2) != 0)
    ]  # delete rows where MV and EE are zero to two decimals
    len_after = len(df)

    return df
    # print(f' {len_after:,} rows remain after {rowsNaN:,} NaN rows and {len_before - len_after:,} zero effective exposure rows \
    # removed from {len_before} inital rows')


# get the reporting date and list of funds to process from the py_reports.xlsm file

########## TEST
# col_ref = "derv_portfolio"  # column reference in py_reports.xlsm to get the report date and funds to process
########## TEST


def rpt_parms(col_ref="dervs", sht_name="qin"):
    import pandas as pd
    from datetime import datetime
    from constants import pthPy
    from utilities import prior_working_day

    try:
        # find the position of the column with the py script name
        col_heads = pd.read_excel(pthPy, sheet_name=sht_name, header=0, nrows=0)
        py_name = col_heads.columns.get_loc(col_ref)
        # print(f"Position of column '{col_ref}': {py_name}")
    except:
        return print(f"\n\n'{col_ref}' column not found in sheet '{sht_name}'\n\n")
    else:
        # find the 'fx' column
        fx = col_heads.columns.get_loc("fx")
        # print(f"Position of column 'fx': {fx}")

        # get report funds and dates
        df = pd.read_excel(
            pthPy,
            sheet_name=sht_name,
            usecols=[0, py_name, fx + 2, fx + 4],
            header=None,
        )
        df = df.dropna(subset=[df.columns[0]])
        # print(df)

        # get report date
        # k= pd.read_excel(pthPy, sheet_name="qin", usecols=[fx+2], header=None, nrows = 1).iloc[0]
        k = df.iloc[0, 2]
        rptDate = (
            k.date()
            if isinstance(k, datetime)
            else prior_working_day(datetime.today()).date()
        )  # prior working day or report date override; has type datetime()
        # print(rptDate)

        # if generic search, get the 'to' date from sheet 'qin'
        if col_ref == "eagle_gen":
            k2 = df.iloc[0, 3]
            date_to = (
                k2.date()
                if isinstance(k2, datetime)
                else prior_working_day(datetime.today()).date()
            )  # prior working day or report date override; has type datetime()
            # print(date_to)

        # delete two date columns and the 'Count --->' row
        df = df[[0, py_name]]
        df = df.drop(df.index[[0, 1]]).reset_index(drop=True)

        # print(df)

        # funds
        funds = df[df.iloc[:, -1] == 1].iloc[:, 0].reset_index(drop=True).tolist()

        return funds, rptDate


########## TEST the ouput of this function
# if __name__ == "__main__":
#     g, h = rpt_parms("derv_portfolio")
#     print(
#         "\n",
#         g,
#         "\n",
#         type(g),
#         "\n",
#         h,
#         "\n",
#         type(h),
#         "\n",
#     )
########## TEST the ouput of this function

# import time
# import os
# import shutil
# import datetime
# from tqdm import tqdm
# from utilities import timediff


def move_files_by_modified_date(source_directory, destination_directory, target_date):
    """
    Moves files from a source directory to a destination directory
    if their modification date is older than or equal to the target_date.

    Args:
        source_directory (str): The path to the directory containing the files.
        destination_directory (str): The path to the directory where files will be moved.
        target_date (datetime.date): The date to compare against file modification dates.
    """

    if not os.path.exists(source_directory):
        print(f"Error: Source directory '{source_directory}' does not exist.")
        return

    if not os.path.exists(destination_directory):
        os.makedirs(destination_directory)
        print(f"Created destination directory: '{destination_directory}'")

    start_time_move_files_by_modified_date = time.time()
    for filename in tqdm(os.listdir(source_directory)):
        source_path = os.path.join(source_directory, filename)

        if os.path.isfile(source_path):
            # Get the modification timestamp of the file
            mod_timestamp = os.path.getmtime(source_path)
            mod_date = datetime.date.fromtimestamp(mod_timestamp)

            if mod_date <= target_date:
                destination_path = os.path.join(destination_directory, filename)
                try:
                    shutil.move(source_path, destination_path)
                    print(f"Moved '{filename}' to '{destination_directory}'")
                except Exception as e:
                    print(f"Error moving '{filename}': {e}")
    print(
        f"{timediff(start_time_move_files_by_modified_date, time.time())} moving \
    files modified up to {target_date.strftime('%d %b %Y')} to folder {destination_directory}"
    )


def batch_list(items, batch_size=10):  # source ChatGPT
    return [items[i : i + batch_size] for i in range(0, len(items), batch_size)]


def r_classifier(report_type, url_input, report_date=datetime.today()):
    """
    receives a link to a sheet in excel format then
    runs issuers_1.ipynb using the given inputs

    """
    import time

    start_time_classifier = time.time()

    import subprocess
    import xlwings as xw
    from constants import pthPy, issuers_1
    from utilities import timediff

    rptDate = report_date
    rptType = "CS1 format only" if report_type == "cs1" else "Reg 28 and Reg 30 only"
    url = url_input.replace('"', "")

    # updating "arc" sheet
    # xw.Book(pthPy).sheets("classifier").range("M1").value = rptType
    xw.Book(pthPy).sheets("arc").range("V4").value = rptType
    # xw.Book(pthPy).sheets("classifier").range("L2").value = url
    xw.Book(pthPy).sheets("arc").range("V8").value = url
    xw.Book(pthPy).save()
    xw.Book(pthPy).close()

    # run issuers_1.ipynb
    subprocess.run([sys.executable, issuers_1])

    return print(
        f"{timediff(start_time_classifier, time.time())} \
            executing issuers_1.ipynb"
    )
